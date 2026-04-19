"""
Coordinacion module: incidencias, seguimientos, derivaciones, reuniones, charlas
Routes prefixed with /api/coordinacion
"""
import uuid
import logging
import jwt
import asyncio
import cloudinary.uploader
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from pydantic import BaseModel, Field
from fastapi import APIRouter, Depends, HTTPException, Query
from .core import db, get_current_user, resolve_user_from_token, require_role, JWT_SECRET, safe_create_index

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/coordinacion", tags=["coordinacion"])

# ══════════════════════════════════════════════════════════════════════════════
# ROLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

COORD_VIEW_ROLES = ["owner", "admin", "director", "coordinator", "psicologo"]
COORD_WRITE_ROLES = ["coordinator", "admin", "owner"]
COORD_DELETE_ROLES = ["admin", "owner"]

# ══════════════════════════════════════════════════════════════════════════════
# ENUMS
# ══════════════════════════════════════════════════════════════════════════════

INCIDENCIA_TYPES = [
    "conducta_disruptiva", "falta_respeto", "agresion_verbal",
    "agresion_fisica", "incumplimiento_normas", "conflicto_companeros",
    "ausencias_reiteradas", "incumplimiento_academico", "observacion_preventiva"
]
INCIDENCIA_SEVERITIES = ["baja", "media", "alta", "critica"]
INCIDENCIA_STATUSES = [
    "nueva", "en_revision", "en_seguimiento", "citacion_programada",
    "derivada", "resuelta", "cerrada"
]
PARENT_INVOLVEMENT_OPTIONS = ["ninguna", "informada", "presente", "comprometida"]

DERIVACION_AREAS = ["psicologia", "direccion", "tutoria", "orientacion_familiar", "externa"]
DERIVACION_STATUSES = ["pendiente", "en_proceso", "resuelta", "cancelada"]
DERIVACION_PRIORITIES = ["baja", "media", "alta", "urgente"]

DERIVACION_AREA_LABELS = {
    "psicologia": "Psicologia",
    "direccion": "Direccion",
    "tutoria": "Tutoria",
    "orientacion_familiar": "Orientacion familiar",
    "externa": "Derivacion externa"
}

# Map area to role for auto-assignment
AREA_TO_ROLE = {
    "psicologia": "psicologo",
    "direccion": "director",
    "tutoria": "teacher",
}

INCIDENCIA_TYPE_LABELS = {
    "conducta_disruptiva": "Conducta disruptiva",
    "falta_respeto": "Falta de respeto",
    "agresion_verbal": "Agresion verbal",
    "agresion_fisica": "Agresion fisica",
    "incumplimiento_normas": "Incumplimiento de normas",
    "conflicto_companeros": "Conflicto entre companeros",
    "ausencias_reiteradas": "Ausencias reiteradas",
    "incumplimiento_academico": "Incumplimiento academico",
    "observacion_preventiva": "Observacion preventiva"
}

SEVERITY_LABELS = {
    "baja": "Baja", "media": "Media", "alta": "Alta", "critica": "Critica"
}

STATUS_LABELS = {
    "nueva": "Nueva", "en_revision": "En revision", "en_seguimiento": "En seguimiento",
    "citacion_programada": "Citacion programada", "derivada": "Derivada",
    "resuelta": "Resuelta", "cerrada": "Cerrada"
}

# ══════════════════════════════════════════════════════════════════════════════
# AUDIT LOG (parametrized from psychology pattern)
# ══════════════════════════════════════════════════════════════════════════════

async def log_coordinacion_audit(user_id, action, resource_type, resource_id, student_id, school_id):
    try:
        await db.coordinacion_audit_log.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "action": action,
            "resource_type": resource_type,
            "resource_id": resource_id,
            "student_id": student_id,
            "school_id": school_id,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
    except Exception as e:
        logger.error(f"Coordinacion audit log error: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# INDEXES (called from server.py startup)
# ══════════════════════════════════════════════════════════════════════════════

async def ensure_coordinacion_indexes():
    try:
        await safe_create_index(db.coordinacion_incidencias, [("school_id", 1), ("status", 1), ("severity", -1)])
        await safe_create_index(db.coordinacion_incidencias, [("school_id", 1), ("student_id", 1), ("occurred_at", -1)])
        await safe_create_index(db.coordinacion_incidencias, [("school_id", 1), ("grade_id", 1), ("section_id", 1)])
        await safe_create_index(db.coordinacion_incidencias, [("school_id", 1), ("assigned_to", 1), ("status", 1)])
        await safe_create_index(db.coordinacion_incidencias, [("school_id", 1), ("deleted_at", 1)])
        await safe_create_index(db.coordinacion_seguimientos, [("school_id", 1), ("incidencia_id", 1), ("entry_date", -1)])
        await safe_create_index(db.coordinacion_seguimientos, [("school_id", 1), ("student_id", 1), ("entry_date", -1)])
        await safe_create_index(db.coordinacion_derivaciones, [("school_id", 1), ("incidencia_id", 1)])
        await safe_create_index(db.coordinacion_derivaciones, [("school_id", 1), ("status", 1), ("to_area", 1)])
        await safe_create_index(db.coordinacion_derivaciones, [("school_id", 1), ("to_user_id", 1), ("status", 1)])
        await safe_create_index(db.coordinacion_charlas, [("school_id", 1), ("scheduled_at", -1)])
        await safe_create_index(db.coordinacion_charlas, [("school_id", 1), ("status", 1)])
        await safe_create_index(db.coordinacion_reuniones, [("school_id", 1), ("student_id", 1), ("scheduled_at", -1)])
        await safe_create_index(db.coordinacion_reuniones, [("school_id", 1), ("status", 1)])
        logger.info("Coordinacion indexes created successfully")
    except Exception as e:
        logger.error(f"Error creating coordinacion indexes: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# PYDANTIC MODELS
# ══════════════════════════════════════════════════════════════════════════════

class IncidenciaCreate(BaseModel):
    student_id: str
    grade_id: str
    section_id: str
    type: str
    severity: str
    title: str = Field(max_length=140)
    description: str = Field(max_length=4000)
    occurred_at: str
    assigned_to: Optional[str] = None
    initial_action: Optional[str] = None
    confidential: bool = False
    notify_parents: bool = False
    tags: List[str] = []

class IncidenciaUpdate(BaseModel):
    type: Optional[str] = None
    severity: Optional[str] = None
    status: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    assigned_to: Optional[str] = None
    initial_action: Optional[str] = None
    confidential: Optional[bool] = None
    notify_parents: Optional[bool] = None
    tags: Optional[List[str]] = None

class SeguimientoCreate(BaseModel):
    observation: str = Field(max_length=4000)
    commitment: Optional[str] = None
    student_response: Optional[str] = None
    parent_involvement: str = "ninguna"
    next_steps: Optional[str] = None
    next_review_at: Optional[str] = None
    new_status: str

class DerivacionCreate(BaseModel):
    incidencia_id: str
    to_area: str
    priority: str = "media"
    reason: str = Field(max_length=4000)
    notes: Optional[str] = None
    to_user_id: Optional[str] = None

class DerivacionUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None
    resolution_notes: Optional[str] = None
    to_user_id: Optional[str] = None

REUNION_STATUSES = ["programada", "confirmada", "realizada", "cancelada", "no_asistio"]

class ReunionCreate(BaseModel):
    student_id: str
    incidencia_id: Optional[str] = None
    scheduled_at: str
    location: str = "Oficina de Coordinacion"
    agenda: str = Field(max_length=4000)
    parent_ids: List[str] = []
    notes: Optional[str] = None

class ReunionUpdate(BaseModel):
    status: Optional[str] = None
    scheduled_at: Optional[str] = None
    location: Optional[str] = None
    agenda: Optional[str] = None
    notes: Optional[str] = None
    outcome: Optional[str] = None
    commitments: Optional[str] = None

# ── CHARLA MODELS ─────────────────────────────────────────────────────────────

CHARLA_STATUSES = ["programada", "en_curso", "realizada", "cancelada"]

class CharlaCreate(BaseModel):
    title: str = Field(max_length=200)
    description: str = Field(max_length=4000)
    scheduled_at: str
    duration_minutes: int = 60
    location: str = "Auditorio"
    target_grades: List[str] = []
    target_sections: List[str] = []
    topics: List[str] = []
    notes: Optional[str] = None

class CharlaUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    scheduled_at: Optional[str] = None
    duration_minutes: Optional[int] = None
    location: Optional[str] = None
    target_grades: Optional[List[str]] = None
    target_sections: Optional[List[str]] = None
    topics: Optional[List[str]] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class MaterialAdd(BaseModel):
    type: str  # image, pdf, video, link
    url: str
    public_id: Optional[str] = None
    name: str
    size_bytes: Optional[int] = None

class AsistenciaUpdate(BaseModel):
    attendance: List[dict]  # [{"student_id": str, "present": bool}]

# ══════════════════════════════════════════════════════════════════════════════
# ENUMS ENDPOINT
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/enums")
async def get_enums(current_user=Depends(require_role(COORD_VIEW_ROLES))):
    return {
        "types": [{"id": k, "label": v} for k, v in INCIDENCIA_TYPE_LABELS.items()],
        "severities": [{"id": k, "label": v} for k, v in SEVERITY_LABELS.items()],
        "statuses": [{"id": k, "label": v} for k, v in STATUS_LABELS.items()],
        "parent_involvement": PARENT_INVOLVEMENT_OPTIONS,
        "derivacion_areas": [{"id": k, "label": v} for k, v in DERIVACION_AREA_LABELS.items()],
        "derivacion_statuses": DERIVACION_STATUSES,
        "derivacion_priorities": DERIVACION_PRIORITIES,
    }

# ══════════════════════════════════════════════════════════════════════════════
# GRADES, SECTIONS, STUDENTS (for incidencia form)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/grades")
async def get_coordinacion_grades(user=Depends(require_role(COORD_VIEW_ROLES))):
    """Get grades for the coordinator's school"""
    school_id = user["school_id"]
    grades = await db.grades.find(
        {"school_id": school_id, "activo": True},
        {"_id": 0, "id": 1, "nombre": 1, "nivel_id": 1, "orden": 1}
    ).sort("orden", 1).to_list(100)
    return grades

@router.get("/sections")
async def get_coordinacion_sections(
    grade_id: str,
    user=Depends(require_role(COORD_VIEW_ROLES))
):
    """Get sections for a specific grade"""
    school_id = user["school_id"]
    sections = await db.sections.find(
        {"school_id": school_id, "grado_id": grade_id, "activo": True},
        {"_id": 0, "id": 1, "nombre": 1, "grado_id": 1}
    ).sort("nombre", 1).to_list(50)
    return sections

@router.get("/students")
async def get_coordinacion_students(
    section_id: str,
    user=Depends(require_role(COORD_VIEW_ROLES))
):
    """Get students for a specific section"""
    school_id = user["school_id"]
    students = await db.users.find(
        {
            "school_id": school_id,
            "seccion_id": section_id,
            "role": "student",
            "student_status": {"$in": ["enrolled", "active"]}
        },
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "photo_url": 1}
    ).sort([("last_name", 1), ("name", 1)]).to_list(100)
    return students

# ══════════════════════════════════════════════════════════════════════════════
# INCIDENCIAS CRUD
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/incidencias")
async def list_incidencias(
    status: Optional[str] = None,
    severity: Optional[str] = None,
    grade_id: Optional[str] = None,
    section_id: Optional[str] = None,
    student_id: Optional[str] = None,
    assigned_to: Optional[str] = None,
    q: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    user=Depends(require_role(COORD_VIEW_ROLES))
):
    school_id = user["school_id"]
    query = {"school_id": school_id, "deleted_at": None}
    if status:
        query["status"] = status
    if severity:
        query["severity"] = severity
    if grade_id:
        query["grade_id"] = grade_id
    if section_id:
        query["section_id"] = section_id
    if student_id:
        query["student_id"] = student_id
    if assigned_to:
        query["assigned_to"] = assigned_to
    if q:
        query["$or"] = [
            {"title": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}}
        ]

    total = await db.coordinacion_incidencias.count_documents(query)
    skip = (page - 1) * page_size
    items = await db.coordinacion_incidencias.find(
        query, {"_id": 0}
    ).sort("occurred_at", -1).skip(skip).limit(page_size).to_list(page_size)

    # Enrich with student names
    student_ids = list({i["student_id"] for i in items if i.get("student_id")})
    if student_ids:
        students = await db.users.find(
            {"id": {"$in": student_ids}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1, "photo_url": 1}
        ).to_list(500)
        student_map = {s["id"]: s for s in students}
        for item in items:
            s = student_map.get(item.get("student_id"), {})
            item["student_name"] = f"{s.get('name', '')} {s.get('last_name', '')}".strip()
            item["student_photo"] = s.get("photo_url")

    # Enrich with assigned_to names
    assigned_ids = list({i["assigned_to"] for i in items if i.get("assigned_to")})
    if assigned_ids:
        assignees = await db.users.find(
            {"id": {"$in": assigned_ids}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(50)
        assignee_map = {u["id"]: f"{u.get('name', '')} {u.get('last_name', '')}".strip() for u in assignees}
        for item in items:
            if item.get("assigned_to"):
                item["assigned_to_name"] = assignee_map.get(item["assigned_to"], "")

    return {"items": items, "page": page, "page_size": page_size, "total": total}


@router.get("/incidencias/{incidencia_id}")
async def get_incidencia(incidencia_id: str, user=Depends(require_role(COORD_VIEW_ROLES))):
    school_id = user["school_id"]
    inc = await db.coordinacion_incidencias.find_one(
        {"id": incidencia_id, "school_id": school_id, "deleted_at": None}, {"_id": 0}
    )
    if not inc:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")

    # Enrich student info
    student = await db.users.find_one(
        {"id": inc["student_id"]},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "photo_url": 1, "email": 1}
    )
    if student:
        inc["student_name"] = f"{student.get('name', '')} {student.get('last_name', '')}".strip()
        inc["student_photo"] = student.get("photo_url")

    # Enrich grade/section names
    if inc.get("grade_id"):
        grade = await db.grades.find_one({"id": inc["grade_id"]}, {"_id": 0, "nombre": 1})
        inc["grade_name"] = grade.get("nombre", "") if grade else ""
    if inc.get("section_id"):
        section = await db.sections.find_one({"id": inc["section_id"]}, {"_id": 0, "nombre": 1})
        inc["section_name"] = section.get("nombre", "") if section else ""

    # Get seguimientos count
    seg_count = await db.coordinacion_seguimientos.count_documents(
        {"incidencia_id": incidencia_id, "school_id": school_id, "deleted_at": None}
    )
    inc["seguimientos_count"] = seg_count

    await log_coordinacion_audit(user["id"], "view", "incidencia", incidencia_id, inc["student_id"], school_id)
    return inc


@router.post("/incidencias")
async def create_incidencia(data: IncidenciaCreate, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]

    if data.type not in INCIDENCIA_TYPES:
        raise HTTPException(status_code=400, detail=f"Tipo invalido: {data.type}")
    if data.severity not in INCIDENCIA_SEVERITIES:
        raise HTTPException(status_code=400, detail=f"Severidad invalida: {data.severity}")

    # Verify student belongs to school
    student = await db.users.find_one(
        {"id": data.student_id, "school_id": school_id, "role": "student"},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1}
    )
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")

    now = datetime.now(timezone.utc).isoformat()
    inc_id = str(uuid.uuid4())

    incidencia = {
        "id": inc_id,
        "school_id": school_id,
        "student_id": data.student_id,
        "grade_id": data.grade_id,
        "section_id": data.section_id,
        "type": data.type,
        "severity": data.severity,
        "status": "nueva",
        "title": data.title,
        "description": data.description,
        "occurred_at": data.occurred_at,
        "reported_by": user["id"],
        "assigned_to": data.assigned_to or user["id"],
        "evidence": [],
        "initial_action": data.initial_action,
        "confidential": data.confidential,
        "notify_parents": data.notify_parents,
        "tags": data.tags,
        "created_at": now,
        "updated_at": now,
        "created_by": user["id"],
        "updated_by": user["id"],
        "deleted_at": None,
    }

    await db.coordinacion_incidencias.insert_one(incidencia)
    del incidencia["_id"]

    await log_coordinacion_audit(user["id"], "create", "incidencia", inc_id, data.student_id, school_id)
    logger.info(f"Incidencia created: {inc_id} by {user['id']} school={school_id}")

    return incidencia


@router.patch("/incidencias/{incidencia_id}")
async def update_incidencia(incidencia_id: str, data: IncidenciaUpdate, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]
    inc = await db.coordinacion_incidencias.find_one(
        {"id": incidencia_id, "school_id": school_id, "deleted_at": None}, {"_id": 0, "id": 1, "student_id": 1}
    )
    if not inc:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")

    update = {}
    if data.type is not None:
        if data.type not in INCIDENCIA_TYPES:
            raise HTTPException(status_code=400, detail=f"Tipo invalido: {data.type}")
        update["type"] = data.type
    if data.severity is not None:
        if data.severity not in INCIDENCIA_SEVERITIES:
            raise HTTPException(status_code=400, detail=f"Severidad invalida: {data.severity}")
        update["severity"] = data.severity
    if data.status is not None:
        if data.status not in INCIDENCIA_STATUSES:
            raise HTTPException(status_code=400, detail=f"Estado invalido: {data.status}")
        update["status"] = data.status
    if data.title is not None:
        update["title"] = data.title
    if data.description is not None:
        update["description"] = data.description
    if data.assigned_to is not None:
        update["assigned_to"] = data.assigned_to
    if data.initial_action is not None:
        update["initial_action"] = data.initial_action
    if data.confidential is not None:
        update["confidential"] = data.confidential
    if data.notify_parents is not None:
        update["notify_parents"] = data.notify_parents
    if data.tags is not None:
        update["tags"] = data.tags

    if not update:
        raise HTTPException(status_code=400, detail="No hay campos para actualizar")

    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["updated_by"] = user["id"]

    await db.coordinacion_incidencias.update_one({"id": incidencia_id}, {"$set": update})
    await log_coordinacion_audit(user["id"], "update", "incidencia", incidencia_id, inc["student_id"], school_id)

    updated = await db.coordinacion_incidencias.find_one({"id": incidencia_id}, {"_id": 0})
    return updated


@router.delete("/incidencias/{incidencia_id}")
async def delete_incidencia(incidencia_id: str, user=Depends(require_role(COORD_DELETE_ROLES))):
    school_id = user["school_id"]
    inc = await db.coordinacion_incidencias.find_one(
        {"id": incidencia_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1, "student_id": 1}
    )
    if not inc:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")

    now = datetime.now(timezone.utc).isoformat()

    # Soft-delete cascade: incidencia + seguimientos + derivaciones
    await db.coordinacion_incidencias.update_one(
        {"id": incidencia_id}, {"$set": {"deleted_at": now, "updated_by": user["id"]}}
    )
    await db.coordinacion_seguimientos.update_many(
        {"incidencia_id": incidencia_id, "school_id": school_id},
        {"$set": {"deleted_at": now}}
    )
    await db.coordinacion_derivaciones.update_many(
        {"incidencia_id": incidencia_id, "school_id": school_id},
        {"$set": {"deleted_at": now}}
    )

    await log_coordinacion_audit(user["id"], "delete", "incidencia", incidencia_id, inc["student_id"], school_id)
    logger.info(f"Incidencia soft-deleted: {incidencia_id} by {user['id']}")

    return {"message": "Incidencia eliminada correctamente"}


# ══════════════════════════════════════════════════════════════════════════════
# SEGUIMIENTOS CRUD
# ══════════════════════════════════════════════════════════════════════════════


@router.get("/seguimientos")
async def list_seguimientos_global(
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    status: Optional[str] = None,
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    student_id: Optional[str] = None,
    grade_id: Optional[str] = None,
    user=Depends(require_role(COORD_VIEW_ROLES)),
):
    """Global chronological view of all seguimientos for the school."""
    school_id = user["school_id"]
    now_str = datetime.now(timezone.utc).isoformat()

    # Build base query
    query = {"school_id": school_id, "deleted_at": None}
    if student_id:
        query["student_id"] = student_id

    # Date range on next_review_at
    if from_date or to_date:
        date_filter = {}
        if from_date:
            date_filter["$gte"] = from_date
        if to_date:
            date_filter["$lte"] = to_date + "T23:59:59"
        query["next_review_at"] = date_filter

    # Status filter: pendiente (future), vencido (past, incidencia not closed), completado (incidencia closed)
    # We'll compute this after fetching, but filter by grade requires a lookup first
    grade_student_ids = None
    if grade_id:
        grade_students = await db.users.find(
            {"school_id": school_id, "grado_id": grade_id, "role": "student"},
            {"_id": 0, "id": 1}
        ).to_list(500)
        grade_student_ids = [s["id"] for s in grade_students]
        query["student_id"] = {"$in": grade_student_ids}

    # Count for summary (before pagination)
    all_segs = await db.coordinacion_seguimientos.find(
        {"school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1, "next_review_at": 1, "incidencia_id": 1}
    ).to_list(5000)

    # Get closed incidencia ids for summary
    closed_statuses = ["resuelta", "cerrada", "archivada"]
    closed_inc_ids = set()
    inc_ids = list({s["incidencia_id"] for s in all_segs if s.get("incidencia_id")})
    if inc_ids:
        closed_incs = await db.coordinacion_incidencias.find(
            {"id": {"$in": inc_ids}, "status": {"$in": closed_statuses}},
            {"_id": 0, "id": 1}
        ).to_list(5000)
        closed_inc_ids = {i["id"] for i in closed_incs}

    # Week range
    from datetime import date as date_type
    today = date_type.today()
    week_start = today.isoformat()
    week_end = (today + timedelta(days=7)).isoformat()

    summary = {"total": len(all_segs), "pendientes": 0, "vencidos": 0, "esta_semana": 0}
    for s in all_segs:
        review = s.get("next_review_at", "")
        is_closed = s.get("incidencia_id") in closed_inc_ids
        if is_closed:
            pass  # completed, skip for pendiente/vencido counts
        elif review and review < now_str:
            summary["vencidos"] += 1
        elif review and review >= now_str:
            summary["pendientes"] += 1
        if review and week_start <= review[:10] <= week_end:
            summary["esta_semana"] += 1

    # Paginated query with sort by next_review_at asc
    total_filtered = await db.coordinacion_seguimientos.count_documents(query)
    items = await db.coordinacion_seguimientos.find(
        query, {"_id": 0}
    ).sort("next_review_at", 1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)

    # Enrich with student, incidencia, and grade data
    student_ids = list({s["student_id"] for s in items if s.get("student_id")})
    inc_ids_page = list({s["incidencia_id"] for s in items if s.get("incidencia_id")})

    student_map = {}
    if student_ids:
        students = await db.users.find(
            {"id": {"$in": student_ids}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1, "grado_id": 1, "seccion_id": 1}
        ).to_list(200)
        # Get grade/section names
        g_ids = list({s.get("grado_id") for s in students if s.get("grado_id")})
        s_ids = list({s.get("seccion_id") for s in students if s.get("seccion_id")})
        grade_map = {}
        section_map = {}
        if g_ids:
            grades = await db.grades.find({"id": {"$in": g_ids}}, {"_id": 0, "id": 1, "nombre": 1}).to_list(50)
            grade_map = {g["id"]: g["nombre"] for g in grades}
        if s_ids:
            sections = await db.sections.find({"id": {"$in": s_ids}}, {"_id": 0, "id": 1, "nombre": 1}).to_list(100)
            section_map = {s["id"]: s["nombre"] for s in sections}
        for st in students:
            gname = grade_map.get(st.get("grado_id"), "")
            sname = section_map.get(st.get("seccion_id"), "")
            grade_label = f"{gname} {sname}".strip() if gname else ""
            student_map[st["id"]] = {
                "name": f"{st.get('name','')} {st.get('last_name','')}".strip(),
                "grade": grade_label,
            }

    inc_map = {}
    if inc_ids_page:
        incs = await db.coordinacion_incidencias.find(
            {"id": {"$in": inc_ids_page}},
            {"_id": 0, "id": 1, "type": 1, "severity": 1, "status": 1}
        ).to_list(200)
        type_labels = INCIDENCIA_TYPE_LABELS
        inc_map = {i["id"]: {"title": type_labels.get(i.get("type"), i.get("type", "")), "severity": i.get("severity", ""), "status": i.get("status", "")} for i in incs}

    # Apply status filter post-fetch if needed and enrich items
    result_items = []
    for item in items:
        review = item.get("next_review_at", "")
        inc_info = inc_map.get(item.get("incidencia_id"), {})
        is_closed = inc_info.get("status") in closed_statuses
        is_overdue = not is_closed and review and review < now_str
        computed_status = "completado" if is_closed else ("vencido" if is_overdue else "pendiente")

        if status and computed_status != status:
            continue

        st_info = student_map.get(item.get("student_id"), {})
        item["student_name"] = st_info.get("name", "")
        item["student_grade"] = st_info.get("grade", "")
        item["incidencia_title"] = inc_info.get("title", "")
        item["incidencia_severity"] = inc_info.get("severity", "")
        item["is_overdue"] = is_overdue
        item["computed_status"] = computed_status
        result_items.append(item)

    return {
        "items": result_items,
        "summary": summary,
        "total": total_filtered,
        "page": page,
        "page_size": page_size,
    }


@router.get("/incidencias/{incidencia_id}/seguimientos")
async def list_seguimientos(incidencia_id: str, user=Depends(require_role(COORD_VIEW_ROLES))):
    school_id = user["school_id"]

    # Verify incidencia exists and belongs to school
    inc = await db.coordinacion_incidencias.find_one(
        {"id": incidencia_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1}
    )
    if not inc:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")

    items = await db.coordinacion_seguimientos.find(
        {"incidencia_id": incidencia_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0}
    ).sort("entry_date", -1).to_list(200)

    # Enrich with creator names
    creator_ids = list({s["created_by"] for s in items if s.get("created_by")})
    if creator_ids:
        creators = await db.users.find(
            {"id": {"$in": creator_ids}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(50)
        creator_map = {u["id"]: f"{u.get('name', '')} {u.get('last_name', '')}".strip() for u in creators}
        for item in items:
            item["created_by_name"] = creator_map.get(item.get("created_by"), "")

    return {"items": items, "total": len(items)}


@router.post("/incidencias/{incidencia_id}/seguimientos")
async def create_seguimiento(
    incidencia_id: str,
    data: SeguimientoCreate,
    user=Depends(require_role(["coordinator"]))
):
    school_id = user["school_id"]

    inc = await db.coordinacion_incidencias.find_one(
        {"id": incidencia_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1, "student_id": 1, "status": 1}
    )
    if not inc:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")

    if data.new_status not in INCIDENCIA_STATUSES:
        raise HTTPException(status_code=400, detail=f"Estado invalido: {data.new_status}")
    if data.parent_involvement not in PARENT_INVOLVEMENT_OPTIONS:
        raise HTTPException(status_code=400, detail=f"Tipo de participacion invalido: {data.parent_involvement}")

    now = datetime.now(timezone.utc).isoformat()
    seg_id = str(uuid.uuid4())

    seguimiento = {
        "id": seg_id,
        "school_id": school_id,
        "incidencia_id": incidencia_id,
        "student_id": inc["student_id"],
        "entry_date": now,
        "observation": data.observation,
        "commitment": data.commitment,
        "student_response": data.student_response,
        "parent_involvement": data.parent_involvement,
        "next_steps": data.next_steps,
        "next_review_at": data.next_review_at,
        "new_status": data.new_status,
        "created_at": now,
        "updated_at": now,
        "created_by": user["id"],
        "deleted_at": None,
    }

    await db.coordinacion_seguimientos.insert_one(seguimiento)
    del seguimiento["_id"]

    # Atomically update incidencia status
    await db.coordinacion_incidencias.update_one(
        {"id": incidencia_id},
        {"$set": {"status": data.new_status, "updated_at": now, "updated_by": user["id"]}}
    )

    await log_coordinacion_audit(user["id"], "create", "seguimiento", seg_id, inc["student_id"], school_id)
    logger.info(f"Seguimiento created: {seg_id} for incidencia {incidencia_id}")

    return seguimiento


@router.patch("/seguimientos/{seguimiento_id}")
async def update_seguimiento(seguimiento_id: str, data: SeguimientoCreate, user=Depends(require_role(["coordinator"]))):
    school_id = user["school_id"]
    seg = await db.coordinacion_seguimientos.find_one(
        {"id": seguimiento_id, "school_id": school_id, "deleted_at": None}, {"_id": 0}
    )
    if not seg:
        raise HTTPException(status_code=404, detail="Seguimiento no encontrado")

    now = datetime.now(timezone.utc).isoformat()
    update = {
        "observation": data.observation,
        "commitment": data.commitment,
        "student_response": data.student_response,
        "parent_involvement": data.parent_involvement,
        "next_steps": data.next_steps,
        "next_review_at": data.next_review_at,
        "new_status": data.new_status,
        "updated_at": now,
    }
    await db.coordinacion_seguimientos.update_one({"id": seguimiento_id}, {"$set": update})

    # Update parent incidencia status
    await db.coordinacion_incidencias.update_one(
        {"id": seg["incidencia_id"]},
        {"$set": {"status": data.new_status, "updated_at": now, "updated_by": user["id"]}}
    )

    return {"message": "Seguimiento actualizado"}


# ══════════════════════════════════════════════════════════════════════════════
# DERIVACIONES
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/derivaciones")
async def create_derivacion(data: DerivacionCreate, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]

    # Validate area
    if data.to_area not in DERIVACION_AREAS:
        raise HTTPException(status_code=400, detail=f"Area invalida: {data.to_area}")
    if data.priority not in DERIVACION_PRIORITIES:
        raise HTTPException(status_code=400, detail=f"Prioridad invalida: {data.priority}")

    # Validate incidencia exists and belongs to school
    inc = await db.coordinacion_incidencias.find_one(
        {"id": data.incidencia_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1, "student_id": 1, "title": 1}
    )
    if not inc:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")

    # Try auto-assign if to_user_id not provided
    to_user_id = data.to_user_id
    if not to_user_id and data.to_area in AREA_TO_ROLE:
        target_role = AREA_TO_ROLE[data.to_area]
        candidates = await db.users.find(
            {"school_id": school_id, "role": target_role, "status": {"$ne": "inactive"}},
            {"_id": 0, "id": 1}
        ).to_list(10)
        if len(candidates) == 1:
            to_user_id = candidates[0]["id"]
        # If 0 or multiple candidates, leave unassigned

    now = datetime.now(timezone.utc).isoformat()
    deriv_id = str(uuid.uuid4())

    derivacion = {
        "id": deriv_id,
        "school_id": school_id,
        "incidencia_id": data.incidencia_id,
        "student_id": inc["student_id"],
        "from_user_id": user["id"],
        "to_area": data.to_area,
        "to_user_id": to_user_id,
        "status": "pendiente",
        "priority": data.priority,
        "reason": data.reason,
        "notes": data.notes,
        "resolution_notes": None,
        "status_changed_at": now,
        "seen_by_coordinator": True,
        "created_at": now,
        "updated_at": now,
        "created_by": user["id"],
        "deleted_at": None,
    }

    await db.coordinacion_derivaciones.insert_one(derivacion)
    del derivacion["_id"]

    # Update incidencia status to "derivada"
    await db.coordinacion_incidencias.update_one(
        {"id": data.incidencia_id},
        {"$set": {"status": "derivada", "updated_at": now, "updated_by": user["id"]}}
    )

    await log_coordinacion_audit(user["id"], "create", "derivacion", deriv_id, inc["student_id"], school_id)
    logger.info(f"Derivacion created: {deriv_id} to {data.to_area} for incidencia {data.incidencia_id}")

    return derivacion


@router.get("/derivaciones")
async def list_derivaciones(
    status: Optional[str] = None,
    to_area: Optional[str] = None,
    unassigned: Optional[bool] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user=Depends(require_role(COORD_VIEW_ROLES))
):
    school_id = user["school_id"]
    query = {"school_id": school_id, "deleted_at": None}

    if status:
        query["status"] = status
    if to_area:
        query["to_area"] = to_area
    if unassigned:
        query["to_user_id"] = None

    total = await db.coordinacion_derivaciones.count_documents(query)
    items = await db.coordinacion_derivaciones.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)

    # Enrich with names
    student_ids = list(set(i.get("student_id") for i in items if i.get("student_id")))
    user_ids = list(set(
        [i.get("from_user_id") for i in items if i.get("from_user_id")] +
        [i.get("to_user_id") for i in items if i.get("to_user_id")]
    ))
    all_ids = list(set(student_ids + user_ids))

    name_map = {}
    if all_ids:
        users_data = await db.users.find(
            {"id": {"$in": all_ids}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(100)
        name_map = {u["id"]: f"{u.get('name','')} {u.get('last_name','')}".strip() for u in users_data}

    # Enrich with incidencia titles
    inc_ids = list(set(i.get("incidencia_id") for i in items if i.get("incidencia_id")))
    inc_map = {}
    if inc_ids:
        incs = await db.coordinacion_incidencias.find(
            {"id": {"$in": inc_ids}}, {"_id": 0, "id": 1, "title": 1}
        ).to_list(100)
        inc_map = {i["id"]: i.get("title", "") for i in incs}

    for item in items:
        item["student_name"] = name_map.get(item.get("student_id"), "")
        item["from_user_name"] = name_map.get(item.get("from_user_id"), "")
        item["to_user_name"] = name_map.get(item.get("to_user_id"), "Sin asignar")
        item["incidencia_title"] = inc_map.get(item.get("incidencia_id"), "")

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/derivaciones/notifications")
async def get_derivacion_notifications(user=Depends(require_role(COORD_WRITE_ROLES))):
    """Count derivations with status changes not yet seen by coordinator"""
    school_id = user["school_id"]
    count = await db.coordinacion_derivaciones.count_documents({
        "school_id": school_id,
        "deleted_at": None,
        "from_user_id": user["id"],
        "seen_by_coordinator": False,
    })
    return {"unseen_count": count}


@router.get("/derivaciones/{derivacion_id}")
async def get_derivacion(derivacion_id: str, user=Depends(require_role(COORD_VIEW_ROLES))):
    school_id = user["school_id"]
    deriv = await db.coordinacion_derivaciones.find_one(
        {"id": derivacion_id, "school_id": school_id, "deleted_at": None}, {"_id": 0}
    )
    if not deriv:
        raise HTTPException(status_code=404, detail="Derivacion no encontrada")

    # Enrich with names
    ids_to_resolve = [deriv.get("student_id"), deriv.get("from_user_id"), deriv.get("to_user_id")]
    ids_to_resolve = [i for i in ids_to_resolve if i]
    name_map = {}
    if ids_to_resolve:
        users_data = await db.users.find(
            {"id": {"$in": ids_to_resolve}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(10)
        name_map = {u["id"]: f"{u.get('name','')} {u.get('last_name','')}".strip() for u in users_data}

    deriv["student_name"] = name_map.get(deriv.get("student_id"), "")
    deriv["from_user_name"] = name_map.get(deriv.get("from_user_id"), "")
    deriv["to_user_name"] = name_map.get(deriv.get("to_user_id"), "Sin asignar")

    # Get incidencia title
    inc = await db.coordinacion_incidencias.find_one(
        {"id": deriv.get("incidencia_id")}, {"_id": 0, "title": 1}
    )
    deriv["incidencia_title"] = inc.get("title", "") if inc else ""

    # Mark as seen by coordinator
    if deriv.get("from_user_id") == user["id"] and not deriv.get("seen_by_coordinator"):
        await db.coordinacion_derivaciones.update_one(
            {"id": derivacion_id},
            {"$set": {"seen_by_coordinator": True}}
        )

    return deriv


@router.patch("/derivaciones/{derivacion_id}")
async def update_derivacion(derivacion_id: str, data: DerivacionUpdate, user=Depends(require_role(COORD_VIEW_ROLES))):
    """Update derivacion - coordinators, target area users, admin, director can update"""
    school_id = user["school_id"]
    deriv = await db.coordinacion_derivaciones.find_one(
        {"id": derivacion_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1, "student_id": 1, "from_user_id": 1, "status": 1}
    )
    if not deriv:
        raise HTTPException(status_code=404, detail="Derivacion no encontrada")

    now = datetime.now(timezone.utc).isoformat()
    update = {"updated_at": now, "updated_by": user["id"]}

    if data.status:
        if data.status not in DERIVACION_STATUSES:
            raise HTTPException(status_code=400, detail=f"Status invalido: {data.status}")
        update["status"] = data.status
        update["status_changed_at"] = now
        # If someone other than the original coordinator changes status, mark unseen
        if user["id"] != deriv.get("from_user_id"):
            update["seen_by_coordinator"] = False
    if data.notes is not None:
        update["notes"] = data.notes
    if data.resolution_notes is not None:
        update["resolution_notes"] = data.resolution_notes
    if data.to_user_id is not None:
        # Validate user exists in school
        target_user = await db.users.find_one(
            {"id": data.to_user_id, "school_id": school_id}, {"_id": 0, "id": 1}
        )
        if not target_user:
            raise HTTPException(status_code=404, detail="Usuario destino no encontrado")
        update["to_user_id"] = data.to_user_id

    await db.coordinacion_derivaciones.update_one({"id": derivacion_id}, {"$set": update})
    await log_coordinacion_audit(user["id"], "update", "derivacion", derivacion_id, deriv.get("student_id"), school_id)

    # If derivacion resolved/cancelled, optionally update incidencia
    if data.status in ("resuelta", "cancelada"):
        # Check if there are other active derivations for this incidencia
        inc_id = deriv.get("incidencia_id") if "incidencia_id" not in update else None
        if not inc_id:
            d_full = await db.coordinacion_derivaciones.find_one({"id": derivacion_id}, {"_id": 0, "incidencia_id": 1})
            inc_id = d_full.get("incidencia_id") if d_full else None
        if inc_id:
            active_derivs = await db.coordinacion_derivaciones.count_documents({
                "incidencia_id": inc_id, "school_id": school_id,
                "deleted_at": None, "status": {"$in": ["pendiente", "en_proceso"]},
                "id": {"$ne": derivacion_id}
            })
            if active_derivs == 0 and data.status == "resuelta":
                await db.coordinacion_incidencias.update_one(
                    {"id": inc_id, "status": "derivada"},
                    {"$set": {"status": "resuelta", "updated_at": now}}
                )

    return {"message": "Derivacion actualizada"}


@router.get("/staff/{area}")
async def get_staff_by_area(area: str, user=Depends(require_role(COORD_VIEW_ROLES))):
    """Get staff members that can be assigned to a derivation area"""
    school_id = user["school_id"]
    role = AREA_TO_ROLE.get(area)
    if not role:
        return {"staff": []}
    staff = await db.users.find(
        {"school_id": school_id, "role": role, "status": {"$ne": "inactive"}},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "role": 1}
    ).to_list(50)
    for s in staff:
        s["full_name"] = f"{s.get('name','')} {s.get('last_name','')}".strip()
    return {"staff": staff}


# ══════════════════════════════════════════════════════════════════════════════
# REUNIONES CON PADRES
# ══════════════════════════════════════════════════════════════════════════════

def generate_reunion_token(reunion_id, parent_id, school_id):
    """Generate a short-lived JWT for parent confirmation (7 days)"""
    payload = {
        "reunion_id": reunion_id,
        "parent_id": parent_id,
        "school_id": school_id,
        "purpose": "reunion_confirm",
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


@router.post("/reuniones")
async def create_reunion(data: ReunionCreate, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]

    # Validate student exists in school
    student = await db.users.find_one(
        {"id": data.student_id, "school_id": school_id, "role": "student"},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1}
    )
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")

    # If no parent_ids provided, find linked parents
    parent_ids = data.parent_ids
    if not parent_ids:
        parents = await db.users.find(
            {"school_id": school_id, "role": "parent", "children": data.student_id},
            {"_id": 0, "id": 1}
        ).to_list(10)
        parent_ids = [p["id"] for p in parents]

    # Validate incidencia if provided
    if data.incidencia_id:
        inc = await db.coordinacion_incidencias.find_one(
            {"id": data.incidencia_id, "school_id": school_id, "deleted_at": None},
            {"_id": 0, "id": 1}
        )
        if not inc:
            raise HTTPException(status_code=404, detail="Incidencia no encontrada")

    now = datetime.now(timezone.utc).isoformat()
    reunion_id = str(uuid.uuid4())

    reunion = {
        "id": reunion_id,
        "school_id": school_id,
        "student_id": data.student_id,
        "incidencia_id": data.incidencia_id,
        "scheduled_at": data.scheduled_at,
        "location": data.location,
        "agenda": data.agenda,
        "parent_ids": parent_ids,
        "confirmed_parents": [],
        "status": "programada",
        "notes": data.notes,
        "outcome": None,
        "commitments": None,
        "created_at": now,
        "updated_at": now,
        "created_by": user["id"],
        "deleted_at": None,
    }

    await db.coordinacion_reuniones.insert_one(reunion)
    del reunion["_id"]

    # Generate confirmation tokens for each parent
    confirmation_links = []
    for pid in parent_ids:
        token_str = generate_reunion_token(reunion_id, pid, school_id)
        confirmation_links.append({
            "parent_id": pid,
            "token": token_str,
        })

    # Update incidencia status if linked
    if data.incidencia_id:
        await db.coordinacion_incidencias.update_one(
            {"id": data.incidencia_id},
            {"$set": {"status": "citacion_programada", "updated_at": now}}
        )

    await log_coordinacion_audit(user["id"], "create", "reunion", reunion_id, data.student_id, school_id)
    logger.info(f"Reunion created: {reunion_id} for student {data.student_id}")

    reunion["confirmation_links"] = confirmation_links
    return reunion


@router.post("/reuniones/confirm")
async def confirm_reunion(token_str: str = Query(..., alias="token")):
    """Public endpoint - parent confirms reunion attendance via JWT token"""
    try:
        payload = jwt.decode(token_str, JWT_SECRET, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=400, detail="El enlace de confirmacion ha expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=400, detail="Enlace de confirmacion invalido")

    if payload.get("purpose") != "reunion_confirm":
        raise HTTPException(status_code=400, detail="Token invalido")

    reunion_id = payload["reunion_id"]
    parent_id = payload["parent_id"]
    school_id = payload["school_id"]

    reunion = await db.coordinacion_reuniones.find_one(
        {"id": reunion_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1, "parent_ids": 1, "confirmed_parents": 1, "status": 1}
    )
    if not reunion:
        raise HTTPException(status_code=404, detail="Reunion no encontrada")

    if parent_id not in reunion.get("parent_ids", []):
        raise HTTPException(status_code=403, detail="No estas invitado a esta reunion")

    confirmed = reunion.get("confirmed_parents", [])
    if parent_id in confirmed:
        return {"message": "Ya confirmaste tu asistencia", "already_confirmed": True}

    confirmed.append(parent_id)
    update = {"confirmed_parents": confirmed, "updated_at": datetime.now(timezone.utc).isoformat()}

    # Auto-update status to confirmed if at least one parent confirms
    if reunion["status"] == "programada":
        update["status"] = "confirmada"

    await db.coordinacion_reuniones.update_one({"id": reunion_id}, {"$set": update})

    return {"message": "Asistencia confirmada correctamente", "already_confirmed": False}


@router.get("/reuniones")
async def list_reuniones(
    status: Optional[str] = None,
    student_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user=Depends(require_role(COORD_VIEW_ROLES))
):
    school_id = user["school_id"]
    query = {"school_id": school_id, "deleted_at": None}

    if status:
        query["status"] = status
    if student_id:
        query["student_id"] = student_id

    total = await db.coordinacion_reuniones.count_documents(query)
    items = await db.coordinacion_reuniones.find(
        query, {"_id": 0}
    ).sort("scheduled_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)

    # Enrich with names
    all_user_ids = set()
    for item in items:
        if item.get("student_id"):
            all_user_ids.add(item["student_id"])
        for pid in item.get("parent_ids", []):
            all_user_ids.add(pid)
        if item.get("created_by"):
            all_user_ids.add(item["created_by"])

    name_map = {}
    if all_user_ids:
        users_data = await db.users.find(
            {"id": {"$in": list(all_user_ids)}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(100)
        name_map = {u["id"]: f"{u.get('name','')} {u.get('last_name','')}".strip() for u in users_data}

    for item in items:
        item["student_name"] = name_map.get(item.get("student_id"), "")
        item["created_by_name"] = name_map.get(item.get("created_by"), "")
        item["parent_names"] = [name_map.get(pid, "") for pid in item.get("parent_ids", [])]

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/reuniones/{reunion_id}")
async def get_reunion(reunion_id: str, user=Depends(require_role(COORD_VIEW_ROLES))):
    school_id = user["school_id"]
    reunion = await db.coordinacion_reuniones.find_one(
        {"id": reunion_id, "school_id": school_id, "deleted_at": None}, {"_id": 0}
    )
    if not reunion:
        raise HTTPException(status_code=404, detail="Reunion no encontrada")

    # Enrich with names
    all_ids = [reunion.get("student_id"), reunion.get("created_by")] + reunion.get("parent_ids", [])
    all_ids = [i for i in all_ids if i]
    name_map = {}
    if all_ids:
        users_data = await db.users.find(
            {"id": {"$in": all_ids}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(50)
        name_map = {u["id"]: f"{u.get('name','')} {u.get('last_name','')}".strip() for u in users_data}

    reunion["student_name"] = name_map.get(reunion.get("student_id"), "")
    reunion["created_by_name"] = name_map.get(reunion.get("created_by"), "")
    reunion["parent_names"] = [name_map.get(pid, "") for pid in reunion.get("parent_ids", [])]
    reunion["confirmed_parent_names"] = [name_map.get(pid, "") for pid in reunion.get("confirmed_parents", [])]

    # Generate confirmation links for coordinator to share
    confirmation_links = []
    for pid in reunion.get("parent_ids", []):
        if pid not in reunion.get("confirmed_parents", []):
            token_str = generate_reunion_token(reunion_id, pid, school_id)
            confirmation_links.append({
                "parent_id": pid,
                "parent_name": name_map.get(pid, ""),
                "token": token_str,
            })
    reunion["pending_confirmation_links"] = confirmation_links

    return reunion


@router.patch("/reuniones/{reunion_id}")
async def update_reunion(reunion_id: str, data: ReunionUpdate, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]
    reunion = await db.coordinacion_reuniones.find_one(
        {"id": reunion_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1, "student_id": 1}
    )
    if not reunion:
        raise HTTPException(status_code=404, detail="Reunion no encontrada")

    now = datetime.now(timezone.utc).isoformat()
    update = {"updated_at": now, "updated_by": user["id"]}

    if data.status:
        if data.status not in REUNION_STATUSES:
            raise HTTPException(status_code=400, detail=f"Estado invalido: {data.status}")
        update["status"] = data.status
    if data.scheduled_at is not None:
        update["scheduled_at"] = data.scheduled_at
    if data.location is not None:
        update["location"] = data.location
    if data.agenda is not None:
        update["agenda"] = data.agenda
    if data.notes is not None:
        update["notes"] = data.notes
    if data.outcome is not None:
        update["outcome"] = data.outcome
    if data.commitments is not None:
        update["commitments"] = data.commitments

    await db.coordinacion_reuniones.update_one({"id": reunion_id}, {"$set": update})
    await log_coordinacion_audit(user["id"], "update", "reunion", reunion_id, reunion.get("student_id"), school_id)

    return {"message": "Reunion actualizada"}


@router.get("/parents/{student_id}")
async def get_student_parents(student_id: str, user=Depends(require_role(COORD_VIEW_ROLES))):
    """Get parents linked to a student"""
    school_id = user["school_id"]
    parents = await db.users.find(
        {"school_id": school_id, "role": "parent", "children": student_id},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "email": 1, "phone": 1}
    ).to_list(20)
    for p in parents:
        p["full_name"] = f"{p.get('name','')} {p.get('last_name','')}".strip()
    return {"parents": parents}


# ══════════════════════════════════════════════════════════════════════════════
# FICHA EXTENDIDA DEL ESTUDIANTE
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/estudiante/{student_id}/ficha")
async def get_student_ficha(
    student_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user=Depends(require_role(COORD_VIEW_ROLES))
):
    school_id = user["school_id"]

    # Validate student exists
    student = await db.users.find_one(
        {"id": student_id, "school_id": school_id, "role": "student"},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "grado_id": 1, "seccion_id": 1, "photo_url": 1}
    )
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")

    # Enrich with grade/section names
    grade_name = ""
    section_name = ""
    if student.get("grado_id"):
        g = await db.grades.find_one({"id": student["grado_id"]}, {"_id": 0, "nombre": 1})
        grade_name = g["nombre"] if g else ""
    if student.get("seccion_id"):
        s = await db.sections.find_one({"id": student["seccion_id"]}, {"_id": 0, "nombre": 1})
        section_name = s["nombre"] if s else ""

    # Defense in depth: roles that can see confidential
    can_see_confidential = user["role"] in ["coordinator", "admin", "owner", "director", "psicologo"]

    base_filter = {"school_id": school_id, "student_id": student_id, "deleted_at": None}

    # Fetch all 4 event types
    inc_filter = {**base_filter}
    if not can_see_confidential:
        inc_filter["confidential"] = {"$ne": True}

    incidencias = await db.coordinacion_incidencias.find(
        inc_filter, {"_id": 0}
    ).to_list(500)

    seguimientos = await db.coordinacion_seguimientos.find(
        {**base_filter}, {"_id": 0}
    ).to_list(500)

    deriv_filter = {"school_id": school_id, "student_id": student_id, "deleted_at": None}
    derivaciones = await db.coordinacion_derivaciones.find(
        deriv_filter, {"_id": 0}
    ).to_list(200)

    reuniones = await db.coordinacion_reuniones.find(
        {"school_id": school_id, "student_id": student_id, "deleted_at": None}, {"_id": 0}
    ).to_list(200)

    # Build unified timeline
    timeline = []
    for inc in incidencias:
        timeline.append({
            "event_type": "incidencia",
            "id": inc["id"],
            "occurred_at": inc.get("occurred_at") or inc.get("created_at"),
            "title": inc.get("title", ""),
            "severity": inc.get("severity"),
            "status": inc.get("status"),
            "confidential": inc.get("confidential", False),
        })
    for seg in seguimientos:
        timeline.append({
            "event_type": "seguimiento",
            "id": seg["id"],
            "occurred_at": seg.get("entry_date") or seg.get("created_at"),
            "title": seg.get("observation", "")[:100],
            "incidencia_id": seg.get("incidencia_id"),
            "new_status": seg.get("new_status"),
            "next_review_at": seg.get("next_review_at"),
        })
    for der in derivaciones:
        timeline.append({
            "event_type": "derivacion",
            "id": der["id"],
            "occurred_at": der.get("created_at"),
            "title": f"Derivacion a {DERIVACION_AREA_LABELS.get(der.get('to_area'), der.get('to_area', ''))}",
            "to_area": der.get("to_area"),
            "status": der.get("status"),
            "priority": der.get("priority"),
        })
    for reu in reuniones:
        timeline.append({
            "event_type": "reunion",
            "id": reu["id"],
            "occurred_at": reu.get("scheduled_at") or reu.get("created_at"),
            "title": f"Reunion: {reu.get('agenda', '')[:60]}",
            "status": reu.get("status"),
            "location": reu.get("location"),
        })

    # Sort descending by date
    timeline.sort(key=lambda x: x.get("occurred_at") or "", reverse=True)
    total = len(timeline)

    # Paginate
    start = (page - 1) * page_size
    timeline_page = timeline[start:start + page_size]

    # Summary
    now = datetime.now(timezone.utc)
    thirty_days_ago = (now - timedelta(days=30)).isoformat()
    open_statuses = ["nueva", "en_revision", "en_seguimiento", "citacion_programada", "derivada"]
    incidencias_abiertas = sum(1 for i in incidencias if i.get("status") in open_statuses)
    recent_inc = sum(1 for i in incidencias if (i.get("occurred_at") or i.get("created_at", "")) > thirty_days_ago)
    reincidencia_30d = recent_inc >= 3

    ultimo_evento = timeline[0]["occurred_at"] if timeline else None

    return {
        "student": {
            "id": student["id"],
            "full_name": f"{student.get('name','')} {student.get('last_name','')}".strip(),
            "grade": grade_name,
            "section": section_name,
            "photo_url": student.get("photo_url"),
        },
        "summary": {
            "total_incidencias": len(incidencias),
            "incidencias_abiertas": incidencias_abiertas,
            "reincidencia_30d": reincidencia_30d,
            "ultimo_evento_at": ultimo_evento,
            "total_derivaciones": len(derivaciones),
            "total_reuniones": len(reuniones),
        },
        "timeline": timeline_page,
        "page": page,
        "page_size": page_size,
        "total": total,
    }



# ══════════════════════════════════════════════════════════════════════════════
# CHARLAS GRUPALES
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/charlas")
async def create_charla(body: CharlaCreate, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]
    now = datetime.now(timezone.utc).isoformat()
    charla_id = str(uuid.uuid4())
    doc = {
        "id": charla_id,
        "school_id": school_id,
        "title": body.title,
        "description": body.description,
        "scheduled_at": body.scheduled_at,
        "duration_minutes": body.duration_minutes,
        "location": body.location,
        "target_grades": body.target_grades,
        "target_sections": body.target_sections,
        "topics": body.topics,
        "materials": [],
        "attendance": [],
        "status": "programada",
        "notes": body.notes,
        "created_at": now,
        "updated_at": now,
        "created_by": user["id"],
        "deleted_at": None,
    }
    await db.coordinacion_charlas.insert_one(doc)
    await log_coordinacion_audit(user["id"], "create", "charla", charla_id, None, school_id)
    doc.pop("_id", None)
    return doc

@router.get("/charlas")
async def list_charlas(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    user=Depends(require_role(COORD_VIEW_ROLES))
):
    school_id = user["school_id"]
    query = {"school_id": school_id, "deleted_at": None}
    if status:
        query["status"] = status
    total = await db.coordinacion_charlas.count_documents(query)
    items = await db.coordinacion_charlas.find(
        query, {"_id": 0}
    ).sort("scheduled_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    # Enrich with created_by name
    creator_ids = list(set(i.get("created_by") for i in items if i.get("created_by")))
    name_map = {}
    if creator_ids:
        creators = await db.users.find(
            {"id": {"$in": creator_ids}}, {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(50)
        name_map = {c["id"]: f"{c.get('name','')} {c.get('last_name','')}".strip() for c in creators}
    for i in items:
        i["created_by_name"] = name_map.get(i.get("created_by"), "")
        i["attendance_count"] = len([a for a in i.get("attendance", []) if a.get("present")])
        i["materials_count"] = len(i.get("materials", []))
    return {"items": items, "total": total, "page": page, "page_size": page_size}

@router.get("/charlas/{charla_id}")
async def get_charla(charla_id: str, user=Depends(require_role(COORD_VIEW_ROLES))):
    school_id = user["school_id"]
    charla = await db.coordinacion_charlas.find_one(
        {"id": charla_id, "school_id": school_id, "deleted_at": None}, {"_id": 0}
    )
    if not charla:
        raise HTTPException(status_code=404, detail="Charla no encontrada")
    # Enrich creator name
    creator = await db.users.find_one(
        {"id": charla.get("created_by")}, {"_id": 0, "name": 1, "last_name": 1}
    )
    charla["created_by_name"] = f"{creator.get('name','')} {creator.get('last_name','')}".strip() if creator else ""
    # Enrich target grade/section names
    if charla.get("target_grades"):
        grades = await db.grades.find(
            {"id": {"$in": charla["target_grades"]}}, {"_id": 0, "id": 1, "nombre": 1}
        ).to_list(50)
        charla["target_grade_names"] = {g["id"]: g["nombre"] for g in grades}
    if charla.get("target_sections"):
        sections = await db.sections.find(
            {"id": {"$in": charla["target_sections"]}}, {"_id": 0, "id": 1, "nombre": 1}
        ).to_list(50)
        charla["target_section_names"] = {s["id"]: s["nombre"] for s in sections}
    return charla

@router.patch("/charlas/{charla_id}")
async def update_charla(charla_id: str, body: CharlaUpdate, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]
    existing = await db.coordinacion_charlas.find_one(
        {"id": charla_id, "school_id": school_id, "deleted_at": None}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Charla no encontrada")
    updates = {k: v for k, v in body.dict(exclude_unset=True).items() if v is not None}
    if "status" in updates and updates["status"] not in CHARLA_STATUSES:
        raise HTTPException(status_code=400, detail="Estado invalido")
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.coordinacion_charlas.update_one({"id": charla_id}, {"$set": updates})
    await log_coordinacion_audit(user["id"], "update", "charla", charla_id, None, school_id)
    updated = await db.coordinacion_charlas.find_one({"id": charla_id}, {"_id": 0})
    return updated

@router.delete("/charlas/{charla_id}")
async def delete_charla(charla_id: str, user=Depends(require_role(COORD_DELETE_ROLES))):
    school_id = user["school_id"]
    existing = await db.coordinacion_charlas.find_one(
        {"id": charla_id, "school_id": school_id, "deleted_at": None}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Charla no encontrada")
    now = datetime.now(timezone.utc).isoformat()
    await db.coordinacion_charlas.update_one({"id": charla_id}, {"$set": {"deleted_at": now, "updated_at": now}})
    await log_coordinacion_audit(user["id"], "delete", "charla", charla_id, None, school_id)
    return {"ok": True}

@router.post("/charlas/{charla_id}/materiales")
async def add_charla_material(charla_id: str, body: MaterialAdd, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]
    charla = await db.coordinacion_charlas.find_one(
        {"id": charla_id, "school_id": school_id, "deleted_at": None}
    )
    if not charla:
        raise HTTPException(status_code=404, detail="Charla no encontrada")
    if body.type not in ("image", "pdf", "video", "link"):
        raise HTTPException(status_code=400, detail="Tipo de material invalido")
    material = {
        "id": str(uuid.uuid4()),
        "type": body.type,
        "url": body.url,
        "public_id": body.public_id,
        "name": body.name,
        "size_bytes": body.size_bytes,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": user["id"],
    }
    await db.coordinacion_charlas.update_one(
        {"id": charla_id},
        {"$push": {"materials": material}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return material

@router.delete("/charlas/{charla_id}/materiales/{material_id}")
async def remove_charla_material(charla_id: str, material_id: str, user=Depends(require_role(COORD_WRITE_ROLES))):
    school_id = user["school_id"]
    charla = await db.coordinacion_charlas.find_one(
        {"id": charla_id, "school_id": school_id, "deleted_at": None}
    )
    if not charla:
        raise HTTPException(status_code=404, detail="Charla no encontrada")
    material = next((m for m in charla.get("materials", []) if m["id"] == material_id), None)
    if not material:
        raise HTTPException(status_code=404, detail="Material no encontrado")
    # Destroy from Cloudinary if applicable
    if material.get("public_id") and material.get("type") != "link":
        try:
            resource_type = "video" if material["type"] == "video" else ("raw" if material["type"] == "pdf" else "image")
            cloudinary.uploader.destroy(material["public_id"], resource_type=resource_type)
        except Exception as e:
            logger.error(f"Error destroying Cloudinary asset {material['public_id']}: {e}")
    await db.coordinacion_charlas.update_one(
        {"id": charla_id},
        {"$pull": {"materials": {"id": material_id}}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"ok": True}

@router.get("/charlas/{charla_id}/estudiantes")
async def get_charla_students(charla_id: str, user=Depends(require_role(COORD_VIEW_ROLES))):
    """Get expected students for a charla based on target grades/sections."""
    school_id = user["school_id"]
    charla = await db.coordinacion_charlas.find_one(
        {"id": charla_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "target_grades": 1, "target_sections": 1}
    )
    if not charla:
        raise HTTPException(status_code=404, detail="Charla no encontrada")
    student_query = {
        "school_id": school_id,
        "role": "student",
        "student_status": {"$in": ["enrolled", "active"]},
    }
    if charla.get("target_sections"):
        student_query["seccion_id"] = {"$in": charla["target_sections"]}
    elif charla.get("target_grades"):
        student_query["grado_id"] = {"$in": charla["target_grades"]}
    else:
        return {"students": []}
    students = await db.users.find(
        student_query,
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "grado_id": 1, "seccion_id": 1, "photo_url": 1}
    ).sort([("last_name", 1), ("name", 1)]).to_list(500)
    return {"students": students}

@router.post("/charlas/{charla_id}/asistencia")
async def save_charla_attendance(charla_id: str, body: AsistenciaUpdate, user=Depends(require_role(COORD_WRITE_ROLES))):
    """Model A: Manual attendance after the charla ends."""
    school_id = user["school_id"]
    charla = await db.coordinacion_charlas.find_one(
        {"id": charla_id, "school_id": school_id, "deleted_at": None}
    )
    if not charla:
        raise HTTPException(status_code=404, detail="Charla no encontrada")
    # Build attendance records enriched with student names
    student_ids = [a["student_id"] for a in body.attendance]
    students = await db.users.find(
        {"id": {"$in": student_ids}},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1}
    ).to_list(500)
    name_map = {s["id"]: f"{s.get('name','')} {s.get('last_name','')}".strip() for s in students}
    attendance = [
        {
            "student_id": a["student_id"],
            "student_name": name_map.get(a["student_id"], ""),
            "present": a.get("present", False),
        }
        for a in body.attendance
    ]
    now = datetime.now(timezone.utc).isoformat()
    await db.coordinacion_charlas.update_one(
        {"id": charla_id},
        {"$set": {"attendance": attendance, "updated_at": now}}
    )
    await log_coordinacion_audit(user["id"], "attendance", "charla", charla_id, None, school_id)
    return {"ok": True, "attendance_count": len([a for a in attendance if a["present"]])}


# ══════════════════════════════════════════════════════════════════════════════
# AGENDA INTEGRADA
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/agenda")
async def get_agenda(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(require_role(COORD_VIEW_ROLES))
):
    """Unified calendar: reuniones + derivaciones + next_review_at from seguimientos.
    Designed for event_source extensibility (charlas in Phase 3)."""
    school_id = user["school_id"]

    # Default to current month if no dates provided
    now = datetime.now(timezone.utc)
    if not start_date:
        start_date = now.replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        next_month = now.replace(day=28) + timedelta(days=4)
        end_date = next_month.replace(day=1).strftime("%Y-%m-%d")

    events = []

    # 1. Reuniones
    reuniones = await db.coordinacion_reuniones.find(
        {
            "school_id": school_id,
            "deleted_at": None,
            "scheduled_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"},
        },
        {"_id": 0, "id": 1, "student_id": 1, "scheduled_at": 1, "status": 1, "location": 1, "agenda": 1}
    ).to_list(200)

    for r in reuniones:
        events.append({
            "event_source": "reunion",
            "id": r["id"],
            "date": r.get("scheduled_at", ""),
            "title": f"Reunion: {(r.get('agenda') or '')[:50]}",
            "status": r.get("status"),
            "student_id": r.get("student_id"),
            "location": r.get("location"),
        })

    # 2. Derivaciones (use created_at as reference date)
    derivaciones = await db.coordinacion_derivaciones.find(
        {
            "school_id": school_id,
            "deleted_at": None,
            "status": {"$in": ["pendiente", "en_proceso"]},
            "created_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"},
        },
        {"_id": 0, "id": 1, "student_id": 1, "created_at": 1, "to_area": 1, "status": 1, "priority": 1}
    ).to_list(200)

    for d in derivaciones:
        events.append({
            "event_source": "derivacion",
            "id": d["id"],
            "date": d.get("created_at", ""),
            "title": f"Derivacion a {DERIVACION_AREA_LABELS.get(d.get('to_area'), d.get('to_area', ''))}",
            "status": d.get("status"),
            "student_id": d.get("student_id"),
            "priority": d.get("priority"),
        })

    # 3. Next reviews from seguimientos
    seguimientos = await db.coordinacion_seguimientos.find(
        {
            "school_id": school_id,
            "deleted_at": None,
            "next_review_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"},
        },
        {"_id": 0, "id": 1, "student_id": 1, "incidencia_id": 1, "next_review_at": 1, "observation": 1}
    ).to_list(200)

    for s in seguimientos:
        events.append({
            "event_source": "review",
            "id": s["id"],
            "date": s.get("next_review_at", ""),
            "title": f"Revision: {(s.get('observation') or '')[:50]}",
            "status": "pendiente",
            "student_id": s.get("student_id"),
            "incidencia_id": s.get("incidencia_id"),
        })

    # 4. Charlas grupales
    charlas = await db.coordinacion_charlas.find(
        {
            "school_id": school_id,
            "deleted_at": None,
            "scheduled_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"},
        },
        {"_id": 0, "id": 1, "title": 1, "scheduled_at": 1, "status": 1, "location": 1}
    ).to_list(200)

    for c in charlas:
        events.append({
            "event_source": "charla",
            "id": c["id"],
            "date": c.get("scheduled_at", ""),
            "title": f"Charla: {(c.get('title') or '')[:50]}",
            "status": c.get("status"),
            "location": c.get("location"),
        })

    # Enrich with student names
    student_ids = list(set(e.get("student_id") for e in events if e.get("student_id")))
    name_map = {}
    if student_ids:
        students = await db.users.find(
            {"id": {"$in": student_ids}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(100)
        name_map = {s["id"]: f"{s.get('name','')} {s.get('last_name','')}".strip() for s in students}

    for e in events:
        e["student_name"] = name_map.get(e.get("student_id"), "")

    # Sort by date ascending
    events.sort(key=lambda x: x.get("date") or "")

    return {"events": events, "start_date": start_date, "end_date": end_date, "total": len(events)}


@router.get("/agenda/check-conflict")
async def check_agenda_conflict(
    date: str,
    duration_minutes: int = 60,
    exclude_id: Optional[str] = None,
    user=Depends(require_role(COORD_WRITE_ROLES))
):
    """Check for scheduling conflicts (adapted from psychology_agenda.py)"""
    school_id = user["school_id"]
    try:
        dt = datetime.fromisoformat(date.replace("Z", "+00:00"))
        end_dt = dt + timedelta(minutes=duration_minutes)
    except Exception:
        return {"has_conflict": False}

    query = {
        "school_id": school_id,
        "deleted_at": None,
        "status": {"$in": ["programada", "confirmada"]},
        "scheduled_at": {"$lt": end_dt.isoformat()},
    }
    if exclude_id:
        query["id"] = {"$ne": exclude_id}

    # Check against reuniones
    conflict = await db.coordinacion_reuniones.find_one(query, {"_id": 0, "id": 1, "agenda": 1, "scheduled_at": 1})
    return {
        "has_conflict": conflict is not None,
        "conflict": {"title": conflict.get("agenda", "")[:60], "date": conflict.get("scheduled_at")} if conflict else None
    }


# ══════════════════════════════════════════════════════════════════════════════
# VISTAS PADRE (endpoints dedicados)
# ══════════════════════════════════════════════════════════════════════════════

PARENT_ROLES = ["parent"]
STUDENT_ROLES = ["student"]


@router.get("/parent/students")
async def parent_get_students(user=Depends(require_role(PARENT_ROLES))):
    """Parent: list their linked children"""
    school_id = user["school_id"]
    children_ids = user.get("children", [])
    if not children_ids:
        return {"students": []}

    students = await db.users.find(
        {"id": {"$in": children_ids}, "school_id": school_id, "role": "student"},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "photo_url": 1, "grado_id": 1, "seccion_id": 1}
    ).to_list(20)

    for s in students:
        s["full_name"] = f"{s.get('name','')} {s.get('last_name','')}".strip()

    return {"students": students}


@router.get("/parent/incidencias")
async def parent_get_incidencias(
    student_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user=Depends(require_role(PARENT_ROLES))
):
    """Parent: only incidencias of their children, non-confidential, with notify_parents=true"""
    school_id = user["school_id"]
    children_ids = user.get("children", [])
    if not children_ids:
        return {"items": [], "total": 0, "page": page, "page_size": page_size}

    query = {
        "school_id": school_id,
        "deleted_at": None,
        "student_id": {"$in": children_ids},
        "confidential": {"$ne": True},
        "notify_parents": True,
    }
    if student_id:
        if student_id not in children_ids:
            raise HTTPException(status_code=403, detail="No tienes acceso a este estudiante")
        query["student_id"] = student_id

    total = await db.coordinacion_incidencias.count_documents(query)
    items = await db.coordinacion_incidencias.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)

    # Enrich
    student_ids = list(set(i.get("student_id") for i in items if i.get("student_id")))
    name_map = {}
    if student_ids:
        students_data = await db.users.find(
            {"id": {"$in": student_ids}}, {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(50)
        name_map = {s["id"]: f"{s.get('name','')} {s.get('last_name','')}".strip() for s in students_data}
    for item in items:
        item["student_name"] = name_map.get(item.get("student_id"), "")

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/parent/reuniones")
async def parent_get_reuniones(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user=Depends(require_role(PARENT_ROLES))
):
    """Parent: only reuniones where they are in parent_ids"""
    school_id = user["school_id"]
    parent_id = user["id"]

    query = {
        "school_id": school_id,
        "deleted_at": None,
        "parent_ids": parent_id,
    }

    total = await db.coordinacion_reuniones.count_documents(query)
    items = await db.coordinacion_reuniones.find(
        query, {"_id": 0}
    ).sort("scheduled_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)

    # Enrich
    student_ids = list(set(i.get("student_id") for i in items))
    name_map = {}
    if student_ids:
        students_data = await db.users.find(
            {"id": {"$in": student_ids}}, {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).to_list(50)
        name_map = {s["id"]: f"{s.get('name','')} {s.get('last_name','')}".strip() for s in students_data}
    for item in items:
        item["student_name"] = name_map.get(item.get("student_id"), "")
        item["is_confirmed"] = parent_id in (item.get("confirmed_parents") or [])

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.post("/parent/reuniones/{reunion_id}/confirm")
async def parent_confirm_reunion_intranet(reunion_id: str, user=Depends(require_role(PARENT_ROLES))):
    """Parent intranet confirmation (already logged in, no JWT token needed)"""
    school_id = user["school_id"]
    parent_id = user["id"]

    reunion = await db.coordinacion_reuniones.find_one(
        {"id": reunion_id, "school_id": school_id, "deleted_at": None},
        {"_id": 0, "id": 1, "parent_ids": 1, "confirmed_parents": 1, "status": 1}
    )
    if not reunion:
        raise HTTPException(status_code=404, detail="Reunion no encontrada")
    if parent_id not in reunion.get("parent_ids", []):
        raise HTTPException(status_code=403, detail="No estas invitado a esta reunion")

    confirmed = reunion.get("confirmed_parents", [])
    if parent_id in confirmed:
        return {"message": "Ya confirmaste tu asistencia", "already_confirmed": True}

    confirmed.append(parent_id)
    update = {"confirmed_parents": confirmed, "updated_at": datetime.now(timezone.utc).isoformat()}
    if reunion["status"] == "programada":
        update["status"] = "confirmada"

    await db.coordinacion_reuniones.update_one({"id": reunion_id}, {"$set": update})
    return {"message": "Asistencia confirmada correctamente", "already_confirmed": False}


# ══════════════════════════════════════════════════════════════════════════════
# VISTAS ALUMNO (endpoints dedicados)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/student/compromisos")
async def student_get_compromisos(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user=Depends(require_role(STUDENT_ROLES))
):
    """Student: only non-confidential incidencias where they are the student, showing commitments"""
    school_id = user["school_id"]
    student_id = user["id"]

    # Get incidencias (non-confidential only)
    incidencias = await db.coordinacion_incidencias.find(
        {
            "school_id": school_id,
            "student_id": student_id,
            "deleted_at": None,
            "confidential": {"$ne": True},
        },
        {"_id": 0, "id": 1, "title": 1, "type": 1, "severity": 1, "status": 1, "created_at": 1}
    ).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(page_size)

    total = await db.coordinacion_incidencias.count_documents({
        "school_id": school_id, "student_id": student_id, "deleted_at": None, "confidential": {"$ne": True}
    })

    # Get latest seguimiento for each incidencia (for commitments)
    inc_ids = [i["id"] for i in incidencias]
    if inc_ids:
        segs = await db.coordinacion_seguimientos.find(
            {"incidencia_id": {"$in": inc_ids}, "deleted_at": None},
            {"_id": 0, "incidencia_id": 1, "commitment": 1, "next_review_at": 1, "created_at": 1}
        ).sort("created_at", -1).to_list(200)

        # Group by incidencia, take latest
        seg_map = {}
        for s in segs:
            iid = s["incidencia_id"]
            if iid not in seg_map:
                seg_map[iid] = s

        for inc in incidencias:
            seg = seg_map.get(inc["id"])
            inc["latest_commitment"] = seg.get("commitment") if seg else None
            inc["next_review_at"] = seg.get("next_review_at") if seg else None

    return {"items": incidencias, "total": total, "page": page, "page_size": page_size}



# ══════════════════════════════════════════════════════════════════════════════
# SHARED HELPER: REINCIDENTES PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

async def _get_reincidentes(school_id: str, days: int = 30, threshold: int = 3, limit: int = None):
    """Shared pipeline: students with >= threshold incidencias in last N days."""
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    pipeline = [
        {"$match": {"school_id": school_id, "deleted_at": None, "created_at": {"$gte": cutoff}}},
        {"$group": {"_id": "$student_id", "count": {"$sum": 1}}},
        {"$match": {"count": {"$gte": threshold}}},
        {"$sort": {"count": -1}},
    ]
    if limit:
        pipeline.append({"$limit": limit})
    results = await db.coordinacion_incidencias.aggregate(pipeline).to_list(500)
    student_ids = [r["_id"] for r in results]
    student_map = {}
    if student_ids:
        students = await db.users.find(
            {"id": {"$in": student_ids}},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1, "grado_id": 1, "seccion_id": 1}
        ).to_list(500)
        g_ids = list({s.get("grado_id") for s in students if s.get("grado_id")})
        s_ids = list({s.get("seccion_id") for s in students if s.get("seccion_id")})
        grade_map, section_map = {}, {}
        if g_ids:
            grades = await db.grades.find({"id": {"$in": g_ids}}, {"_id": 0, "id": 1, "nombre": 1}).to_list(50)
            grade_map = {g["id"]: g["nombre"] for g in grades}
        if s_ids:
            secs = await db.sections.find({"id": {"$in": s_ids}}, {"_id": 0, "id": 1, "nombre": 1}).to_list(100)
            section_map = {s["id"]: s["nombre"] for s in secs}
        for st in students:
            gname = grade_map.get(st.get("grado_id"), "")
            sname = section_map.get(st.get("seccion_id"), "")
            student_map[st["id"]] = {
                "full_name": f"{st.get('name','')} {st.get('last_name','')}".strip(),
                "grade": f"{gname} {sname}".strip() if gname else "",
            }
    return [
        {
            "student_id": r["_id"],
            "full_name": student_map.get(r["_id"], {}).get("full_name", ""),
            "grade": student_map.get(r["_id"], {}).get("grade", ""),
            "count": r["count"],
        }
        for r in results
    ]


# ══════════════════════════════════════════════════════════════════════════════
# REPORTES AVANZADOS
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/reportes/incidencias-por-grado")
async def report_incidencias_por_grado(
    periodo: Optional[str] = None,
    severidad: Optional[str] = None,
    tipo: Optional[str] = None,
    user=Depends(require_role(COORD_VIEW_ROLES)),
):
    school_id = user["school_id"]
    match = {"school_id": school_id, "deleted_at": None}
    if periodo:
        match["created_at"] = {"$gte": periodo}
    if severidad:
        match["severity"] = severidad
    if tipo:
        match["type"] = tipo

    pipeline = [
        {"$match": match},
        {"$lookup": {
            "from": "users", "localField": "student_id", "foreignField": "id",
            "pipeline": [{"$project": {"_id": 0, "grado_id": 1, "seccion_id": 1}}],
            "as": "student_info"
        }},
        {"$unwind": {"path": "$student_info", "preserveNullAndEmptyArrays": True}},
        {"$group": {
            "_id": {"grado_id": "$student_info.grado_id", "seccion_id": "$student_info.seccion_id"},
            "count": {"$sum": 1},
            "by_severity": {"$push": "$severity"},
        }},
        {"$sort": {"count": -1}},
    ]
    results = await db.coordinacion_incidencias.aggregate(pipeline).to_list(200)

    # Resolve grade/section names
    g_ids = list({r["_id"].get("grado_id") for r in results if r["_id"].get("grado_id")})
    s_ids = list({r["_id"].get("seccion_id") for r in results if r["_id"].get("seccion_id")})
    grade_map, section_map = {}, {}
    if g_ids:
        grades = await db.grades.find({"id": {"$in": g_ids}}, {"_id": 0, "id": 1, "nombre": 1}).to_list(50)
        grade_map = {g["id"]: g["nombre"] for g in grades}
    if s_ids:
        secs = await db.sections.find({"id": {"$in": s_ids}}, {"_id": 0, "id": 1, "nombre": 1}).to_list(100)
        section_map = {s["id"]: s["nombre"] for s in secs}

    items = []
    total_count = 0
    for r in results:
        gid = r["_id"].get("grado_id", "")
        sid = r["_id"].get("seccion_id", "")
        sev_counts = {}
        for s in r.get("by_severity", []):
            sev_counts[s] = sev_counts.get(s, 0) + 1
        items.append({
            "grado_id": gid, "seccion_id": sid,
            "grado_nombre": grade_map.get(gid, "Sin grado"),
            "seccion_nombre": section_map.get(sid, ""),
            "label": f"{grade_map.get(gid, 'Sin grado')} {section_map.get(sid, '')}".strip(),
            "count": r["count"],
            "by_severity": sev_counts,
        })
        total_count += r["count"]

    return {"items": items, "total": total_count}


@router.get("/reportes/reincidentes")
async def report_reincidentes(
    periodo: int = Query(30, ge=1, le=365),
    umbral: int = Query(3, ge=2, le=20),
    limit: Optional[int] = None,
    user=Depends(require_role(COORD_VIEW_ROLES)),
):
    items = await _get_reincidentes(user["school_id"], days=periodo, threshold=umbral, limit=limit)
    return {"items": items, "total": len(items), "periodo_dias": periodo, "umbral": umbral}


@router.get("/reportes/cobertura-charlas")
async def report_cobertura_charlas(
    periodo: Optional[str] = None,
    user=Depends(require_role(COORD_VIEW_ROLES)),
):
    school_id = user["school_id"]
    match = {"school_id": school_id, "deleted_at": None, "status": "realizada"}
    if periodo:
        match["scheduled_at"] = {"$gte": periodo}

    charlas = await db.coordinacion_charlas.find(match, {"_id": 0}).to_list(500)
    if not charlas:
        return {"items": [], "total_charlas": 0, "total_convocados": 0, "total_asistentes": 0, "cobertura_pct": 0}

    items = []
    total_convocados = 0
    total_asistentes = 0
    for c in charlas:
        att = c.get("attendance", [])
        convocados = len(att)
        asistentes = len([a for a in att if a.get("present")])
        pct = round((asistentes / convocados * 100), 1) if convocados > 0 else 0
        total_convocados += convocados
        total_asistentes += asistentes
        items.append({
            "charla_id": c["id"],
            "title": c.get("title", ""),
            "scheduled_at": c.get("scheduled_at", ""),
            "topics": c.get("topics", []),
            "convocados": convocados,
            "asistentes": asistentes,
            "cobertura_pct": pct,
        })

    global_pct = round((total_asistentes / total_convocados * 100), 1) if total_convocados > 0 else 0
    return {
        "items": items, "total_charlas": len(charlas),
        "total_convocados": total_convocados, "total_asistentes": total_asistentes,
        "cobertura_pct": global_pct,
    }


@router.get("/reportes/efectividad-seguimientos")
async def report_efectividad_seguimientos(
    periodo: Optional[str] = None,
    user=Depends(require_role(COORD_VIEW_ROLES)),
):
    school_id = user["school_id"]
    match = {"school_id": school_id, "deleted_at": None}
    if periodo:
        match["created_at"] = {"$gte": periodo}

    total = await db.coordinacion_incidencias.count_documents(match)
    closed_match = {**match, "status": {"$in": ["resuelta", "cerrada", "archivada"]}}
    cerradas = await db.coordinacion_incidencias.count_documents(closed_match)
    abiertas = total - cerradas
    pct = round((cerradas / total * 100), 1) if total > 0 else 0

    # By status breakdown
    pipeline = [
        {"$match": match},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]
    by_status = await db.coordinacion_incidencias.aggregate(pipeline).to_list(20)
    status_breakdown = {r["_id"]: r["count"] for r in by_status}

    return {
        "total": total, "cerradas": cerradas, "abiertas": abiertas,
        "efectividad_pct": pct, "by_status": status_breakdown,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN XLSX / PDF
# ══════════════════════════════════════════════════════════════════════════════

REPORT_TYPES = ["incidencias-por-grado", "reincidentes", "cobertura-charlas", "efectividad-seguimientos"]

@router.get("/reportes/{report_type}/export")
async def export_report(
    report_type: str,
    format: str = Query("xlsx", regex="^(xlsx|pdf)$"),
    periodo: Optional[str] = None,
    severidad: Optional[str] = None,
    tipo: Optional[str] = None,
    umbral: int = 3,
    user=Depends(require_role(COORD_VIEW_ROLES)),
):
    if report_type not in REPORT_TYPES:
        raise HTTPException(status_code=400, detail=f"Tipo de reporte invalido. Opciones: {REPORT_TYPES}")

    from fastapi.responses import StreamingResponse
    import io

    user_name = f"{user.get('name', '')} {user.get('last_name', '')}".strip()
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")

    # Fetch report data using the same endpoints
    if report_type == "incidencias-por-grado":
        data = await report_incidencias_por_grado(periodo=periodo, severidad=severidad, tipo=tipo, user=user)
        title = "Incidencias por Grado"
        headers_list = ["Grado/Seccion", "Total", "Baja", "Media", "Alta", "Critica"]
        rows = []
        for item in data["items"]:
            rows.append([
                item["label"], item["count"],
                item["by_severity"].get("baja", 0), item["by_severity"].get("media", 0),
                item["by_severity"].get("alta", 0), item["by_severity"].get("critica", 0),
            ])
    elif report_type == "reincidentes":
        periodo_days = int(periodo) if periodo and periodo.isdigit() else 30
        data = await report_reincidentes(periodo=periodo_days, umbral=umbral, user=user)
        title = "Estudiantes Reincidentes"
        headers_list = ["Estudiante", "Grado", "Incidencias (30d)"]
        rows = [[r["full_name"], r["grade"], r["count"]] for r in data["items"]]
    elif report_type == "cobertura-charlas":
        data = await report_cobertura_charlas(periodo=periodo, user=user)
        title = "Cobertura de Charlas"
        headers_list = ["Charla", "Fecha", "Convocados", "Asistentes", "Cobertura %"]
        rows = [[c["title"], c["scheduled_at"][:10] if c["scheduled_at"] else "", c["convocados"], c["asistentes"], f"{c['cobertura_pct']}%"] for c in data["items"]]
    else:  # efectividad-seguimientos
        data = await report_efectividad_seguimientos(periodo=periodo, user=user)
        title = "Efectividad de Seguimientos"
        headers_list = ["Metrica", "Valor"]
        rows = [
            ["Total incidencias", data["total"]],
            ["Cerradas/Resueltas", data["cerradas"]],
            ["Abiertas", data["abiertas"]],
            ["Efectividad", f"{data['efectividad_pct']}%"],
        ]
        for status, count in data.get("by_status", {}).items():
            label = STATUS_LABELS.get(status, status)
            rows.append([f"  Estado: {label}", count])

    filters_text = f"Periodo: {periodo or 'Todo'}"
    if severidad:
        filters_text += f" | Severidad: {severidad}"
    if tipo:
        filters_text += f" | Tipo: {tipo}"

    if format == "xlsx":
        return _generate_xlsx(title, headers_list, rows, filters_text, user_name, now_str)
    else:
        return _generate_pdf(title, headers_list, rows, filters_text, user_name, now_str)


def _generate_xlsx(title, headers, rows, filters_text, user_name, now_str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Metadata rows
    header_font = Font(bold=True, size=14, color="1a365d")
    meta_font = Font(italic=True, size=10, color="4a5568")
    ws.append([title])
    ws["A1"].font = header_font
    ws.append([filters_text])
    ws["A2"].font = meta_font
    ws.append([f"Generado: {now_str} por {user_name}"])
    ws["A3"].font = meta_font
    ws.append([])

    # Column headers
    col_header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    col_header_font = Font(bold=True, color="ffffff", size=10)
    thin_border = Border(
        left=Side(style="thin", color="d1d5db"),
        right=Side(style="thin", color="d1d5db"),
        top=Side(style="thin", color="d1d5db"),
        bottom=Side(style="thin", color="d1d5db"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    alt_fill = PatternFill(start_color="f7fafc", end_color="f7fafc", fill_type="solid")
    for row_idx, row_data in enumerate(rows, 6):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
            if (row_idx - 6) % 2 == 1:
                cell.fill = alt_fill

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = max(len(str(headers[col_idx - 1])), *[len(str(r[col_idx - 1])) if col_idx - 1 < len(r) else 0 for r in rows], 8)
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{title.replace(' ', '_')}_{now_str[:10]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _generate_pdf(title, headers, rows, filters_text, user_name, now_str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from fastapi.responses import StreamingResponse
    import io

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("CustomTitle", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1a365d"))
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#4a5568"), spaceAfter=4)

    elements = [
        Paragraph(title, title_style),
        Paragraph(filters_text, meta_style),
        Paragraph(f"Generado: {now_str} por {user_name}", meta_style),
        Spacer(1, 10*mm),
    ]

    # Table
    table_data = [headers] + [[str(v) for v in row] for row in rows]
    num_cols = len(headers)
    col_widths = [max(60, min(180, 500 // num_cols))] * num_cols

    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)

    # Footer
    elements.append(Spacer(1, 15*mm))
    footer_style = ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#9ca3af"))
    elements.append(Paragraph(f"EduNet - Modulo Coordinacion | {now_str}", footer_style))

    doc.build(elements)
    buf.seek(0)
    filename = f"{title.replace(' ', '_')}_{now_str[:10]}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/dashboard")
async def get_dashboard(user=Depends(require_role(COORD_VIEW_ROLES))):
    school_id = user["school_id"]
    now = datetime.now(timezone.utc)
    today_str = now.strftime("%Y-%m-%d")

    async def get_kpis():
        base = {"school_id": school_id, "deleted_at": None}
        active_statuses = ["nueva", "en_revision", "en_seguimiento", "citacion_programada", "derivada"]
        active = await db.coordinacion_incidencias.count_documents({**base, "status": {"$in": active_statuses}})
        new_today = await db.coordinacion_incidencias.count_documents({
            **base, "created_at": {"$regex": f"^{today_str}"}
        })
        students_in_followup = len(await db.coordinacion_incidencias.distinct(
            "student_id", {**base, "status": "en_seguimiento"}
        ))
        reuniones_pending = await db.coordinacion_reuniones.count_documents({
            "school_id": school_id, "deleted_at": None, "status": {"$in": ["programada", "confirmada"]}
        })
        charlas_upcoming = await db.coordinacion_charlas.count_documents({
            "school_id": school_id, "deleted_at": None, "status": "programada"
        })
        derivaciones_pending = await db.coordinacion_derivaciones.count_documents({
            "school_id": school_id, "deleted_at": None, "status": "pendiente"
        })
        return {
            "incidencias_activas": active,
            "incidencias_nuevas_hoy": new_today,
            "estudiantes_en_seguimiento": students_in_followup,
            "reuniones_pendientes": reuniones_pending,
            "charlas_proximas": charlas_upcoming,
            "derivaciones_pendientes": derivaciones_pending,
        }

    async def get_by_severity():
        base = {"school_id": school_id, "deleted_at": None}
        result = {}
        for sev in INCIDENCIA_SEVERITIES:
            result[sev] = await db.coordinacion_incidencias.count_documents({**base, "severity": sev})
        return result

    async def get_by_grade():
        pipeline = [
            {"$match": {"school_id": school_id, "deleted_at": None}},
            {"$group": {"_id": "$grade_id", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 20}
        ]
        results = await db.coordinacion_incidencias.aggregate(pipeline).to_list(20)
        grade_ids = [r["_id"] for r in results if r["_id"]]
        grade_map = {}
        if grade_ids:
            grades = await db.grades.find({"id": {"$in": grade_ids}}, {"_id": 0, "id": 1, "nombre": 1}).to_list(20)
            grade_map = {g["id"]: g["nombre"] for g in grades}
        return [{"grade_id": r["_id"], "grade_name": grade_map.get(r["_id"], ""), "count": r["count"]} for r in results]

    async def get_reincidentes():
        return await _get_reincidentes(school_id, days=30, threshold=3, limit=10)

    async def get_recent_incidencias():
        items = await db.coordinacion_incidencias.find(
            {"school_id": school_id, "deleted_at": None},
            {"_id": 0, "id": 1, "title": 1, "severity": 1, "status": 1, "student_id": 1, "created_at": 1, "occurred_at": 1, "confidential": 1}
        ).sort("created_at", -1).limit(5).to_list(5)
        student_ids = [i["student_id"] for i in items if i.get("student_id")]
        if student_ids:
            students = await db.users.find(
                {"id": {"$in": student_ids}},
                {"_id": 0, "id": 1, "name": 1, "last_name": 1}
            ).to_list(50)
            sm = {s["id"]: f"{s.get('name','')} {s.get('last_name','')}".strip() for s in students}
            for i in items:
                i["student_name"] = sm.get(i.get("student_id"), "")
        return items

    kpis, by_severity, by_grade, reincidentes, recent = await asyncio.gather(
        get_kpis(), get_by_severity(), get_by_grade(), get_reincidentes(), get_recent_incidencias()
    )

    # Build alerts from reincidentes
    alertas = []
    for r in reincidentes:
        alertas.append({
            "type": "reincidencia",
            "student_id": r["student_id"],
            "message": f"{r['full_name']} tiene {r['count']} incidencias en los ultimos 30 dias"
        })

    return {
        "kpis": kpis,
        "by_severity": by_severity,
        "by_grade": by_grade,
        "reincidentes": reincidentes,
        "alertas": alertas,
        "recent_incidencias": recent,
    }
