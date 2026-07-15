"""
Grades module: Registro Auxiliar + Consolidado de Notas
Handles grade entry, auto-save, period locking, and consolidated views.
"""
from fastapi import APIRouter, HTTPException, Depends, Query, BackgroundTasks
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
from datetime import datetime, timezone
import uuid
import logging
import math

from .core import (
    db, get_current_user, resolve_user_from_token,
    require_role, require_admin, require_staff,
    is_admin_user, has_role, is_demo_user, check_demo_user_block,
    JWT_SECRET, JWT_ALGORITHM, now_iso, generate_id,
    ADMIN_ROLES, STAFF_ROLES, ACADEMIC_STUDENT_FILTER,
)
from services.ranking import compute_ranking
from services.register_sync import COLUMN_FIELD_MAP

import jwt

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")


def _deny_students(user):
    """SECURITY guard for staff-only grade endpoints. Blocks students and
    parents from reading or writing Registro Auxiliar / Consolidado data.
    Staff roles (teacher/coordinator/admin/owner/director/auxiliar) keep their
    existing access; the per-course teacher-assignment checks remain in place."""
    if user.get("role") in ("student", "parent"):
        raise HTTPException(status_code=403, detail="No tienes permiso para acceder a esta información")


async def _fetch_section_students(school_id: str, section_id: str):
    """Fetch enrolled/active students for a SINGLE section, matching on either
    `seccion_id` or `section_id` (schools use both field names). Sorted by last
    name, name. No cross-section merging — each register shows exactly the
    students of its own section, identical for owner and teacher."""
    students = await db.users.find(
        {
            "school_id": school_id,
            "role": "student",
            "student_status": {"$in": ["enrolled", "active"]},
            "$or": [
                {"seccion_id": section_id},
                {"section_id": section_id},
            ],
        },
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "dni": 1},
    ).sort([("last_name", 1), ("name", 1)]).to_list(500)

    seen = set()
    unique = []
    for s in students:
        sid = s.get("id")
        if sid and sid not in seen:
            seen.add(sid)
            unique.append(s)
    return unique


async def _resolve_effective_section_id(school_id: str, subject_id: str, section_id: str,
                                        role: str = None, teacher_id: str = None) -> str:
    """Resolve the section whose ENROLLED students belong to this course.

    Production bug (Eusebio Arróniz, jun-2026): a subject's stored `section_id`
    can be SWAPPED/mismatched relative to the section recorded on the course's
    `academic_assignments` doc (e.g. "Álgebra 4°A" has subject.section_id → B
    while the teacher's assignment → A). The teacher dashboard card and the
    official Usuarios/Estudiantes list use the teacher's ASSIGNMENT section, so
    the TEACHER's Registro Auxiliar must use that same source.

    Scope of the override: TEACHERS ONLY. Owners/admins/coordinators navigate
    the register BY SECTION explicitly (they can view every section of a
    subject), so for them the requested `section_id` is authoritative and must
    NOT be overridden — otherwise all of a subject's sections collapse to one
    assignment section and the rosters appear swapped between A and B.

    Returns the teacher's assignment section_id when applicable, else the
    requested `section_id` unchanged."""
    if role == "teacher" and teacher_id:
        # 1) If the teacher has an assignment for the EXACT requested section,
        #    honor it. This disambiguates the case where the same subject is
        #    assigned to MULTIPLE sections (e.g. Diana teaches "X" in both
        #    3 años · ÚNICA and 4 años · ÚNICA): the section the teacher opened
        #    must win, instead of an arbitrary `find_one` picking the wrong one.
        if section_id:
            exact = await db.academic_assignments.find_one({
                "school_id": school_id, "teacher_id": teacher_id,
                "subject_id": subject_id, "section_id": section_id,
            })
            if exact:
                return section_id
        # 2) No assignment matches the requested section → the subject's stored
        #    section is swapped relative to the teacher's assignment. Fall back
        #    to the assignment's section (original swap-correction behavior).
        a = await db.academic_assignments.find_one({
            "school_id": school_id, "teacher_id": teacher_id,
            "subject_id": subject_id, "status": "activo",
        }) or await db.academic_assignments.find_one({
            "school_id": school_id, "teacher_id": teacher_id, "subject_id": subject_id,
        })
        if a and a.get("section_id"):
            return a["section_id"]

    # Owner/admin/others: honor the explicitly requested section.
    return section_id


async def _assert_teacher_assignment(school_id: str, teacher_id: str, subject_id: str, section_id: str):
    """RBAC guard for teachers on Registro Auxiliar endpoints.

    Accepts the teacher when they have an ACTIVE assignment for this subject in
    the exact section, OR (fallback) ANY assignment for this (teacher, subject).

    The fallback exists because a subject's stored `section_id` can differ from
    the section recorded on the teacher's assignment (duplicate grade/section
    docs in production). The assignment is the source of truth for "this teacher
    teaches this subject", so we grant permission — but students are STILL
    fetched from the requested `section_id` (the same section the owner uses),
    so teacher and owner always see the exact same roster. No section merging.

    Returns the matched assignment dict, or raises 403 when the teacher has no
    assignment at all for this subject."""
    assignment = await db.academic_assignments.find_one({
        "school_id": school_id,
        "teacher_id": teacher_id,
        "subject_id": subject_id,
        "section_id": section_id,
        "status": "activo",
    })
    if assignment:
        return assignment

    # Fallback: any assignment for this (teacher, subject), regardless of section
    # or status. Mirrors the teacher dashboard (`/teacher/courses`), which lists
    # the course whenever such an assignment exists.
    assignment = await db.academic_assignments.find_one({
        "school_id": school_id,
        "teacher_id": teacher_id,
        "subject_id": subject_id,
    })
    if assignment:
        return assignment

    raise HTTPException(status_code=403, detail="No tienes asignacion para este curso")




# ══════════════════════════════════════════════════════════════════════════════
# MODELS
# ══════════════════════════════════════════════════════════════════════════════

class GradeEntry(BaseModel):
    student_id: str
    # Actitudinal sub-fields
    act_co: Optional[float] = None
    act_re: Optional[float] = None
    # Revisión de Fichas sub-fields
    rf_r1: Optional[float] = None
    rf_r2: Optional[float] = None
    rf_r3: Optional[float] = None
    rf_r4: Optional[float] = None
    rf_r5: Optional[float] = None
    # Competencia sub-fields
    comp_c1: Optional[float] = None
    comp_c2: Optional[float] = None
    # Participaciones sub-fields
    part_p1: Optional[float] = None
    part_p2: Optional[float] = None
    part_p3: Optional[float] = None
    part_exp: Optional[float] = None
    part_tg: Optional[float] = None
    part_p: Optional[float] = None
    # Single-column fields
    exam_mensual: Optional[float] = None
    exam_bimestral: Optional[float] = None
    # Phase 5 — free-form map for columns that belong to the school's
    # custom template. Keys are column ids (UUID-style), values are
    # vigesimal notes (0-20) or None to clear the cell. Stored under
    # `student_grades.grades_dynamic.<column_id>` so legacy static
    # fields stay untouched.
    grades_dynamic: Optional[Dict[str, Optional[float]]] = None

GRADE_SUB_FIELDS = [
    "act_co", "act_re",
    "rf_r1", "rf_r2", "rf_r3", "rf_r4", "rf_r5",
    "comp_c1", "comp_c2",
    "part_p1", "part_p2", "part_p3", "part_exp", "part_tg", "part_p",
    "exam_mensual", "exam_bimestral",
]

class GradeSaveRequest(BaseModel):
    subject_id: str
    section_id: str
    period_id: str
    grades: List[GradeEntry]

class EvalConfigUpdate(BaseModel):
    attitude_weight: Optional[float] = 0.10
    worksheets_weight: Optional[float] = 0.25
    competency_weight: Optional[float] = 0.05
    participation_weight: Optional[float] = 0.25
    monthly_exam_weight: Optional[float] = 0.15
    bimestral_exam_weight: Optional[float] = 0.20

class LockRequest(BaseModel):
    subject_id: str
    section_id: str
    period_id: str

# ══════════════════════════════════════════════════════════════════════════════
# HELPER: Calculate final grade
# ══════════════════════════════════════════════════════════════════════════════

def _avg(values):
    """Calculate average of non-None values. Returns None if all are None."""
    nums = [v for v in values if v is not None]
    if not nums:
        return None
    return round(sum(nums) / len(nums), 1)

def _resolve_dynamic_value(sub_id, field_key, grade: dict, grades_dyn: dict):
    """Resolve a single stored value for a subcolumna / columna-final using the
    same lookup priority across the codebase. Returns float or None."""
    val = None
    if sub_id and sub_id in grades_dyn:
        val = grades_dyn.get(sub_id)
    if val is None and field_key and field_key in grades_dyn:
        val = grades_dyn.get(field_key)
    if val is None and field_key:
        if field_key in grade:
            val = grade.get(field_key)
        elif field_key in COLUMN_FIELD_MAP:
            val = grade.get(COLUMN_FIELD_MAP[field_key])
    if val is None and sub_id:
        if sub_id in grade and sub_id not in ("id", "subcolumnas"):
            val = grade.get(sub_id)
        elif sub_id in COLUMN_FIELD_MAP:
            val = grade.get(COLUMN_FIELD_MAP[sub_id])
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _criterio_avg(cri: dict, grade: dict, grades_dyn: dict):
    """Mean of a criterio's input/examen subcolumna values (skips derived
    promedio columns). Returns float or None when there's no data."""
    values = []
    for sub in cri.get("subcolumnas") or []:
        tipo = (sub.get("tipo") or "input").lower()
        if tipo in ("promedio_auto", "promedio", "promedio_manual", "auto"):
            continue
        v = _resolve_dynamic_value(sub.get("id"), sub.get("field_key"), grade, grades_dyn)
        if v is not None:
            values.append(v)
    if not values:
        return None
    return round(sum(values) / len(values), 1)


def _calculate_final_grade_grupo_mode(grade: dict, template: dict):
    """Modo 'grupo': cada grupo pondera el PROMEDIO SIMPLE de sus miembros
    (criterios y/o columnas finales). Final = Σ(promedio_grupo × %grupo)."""
    grupos = template.get("grupos") or []
    if not grupos:
        return None
    grades_dyn = grade.get("grades_dynamic") or {}
    criterios_by_id = {c.get("id"): c for c in (template.get("criterios") or [])}
    finales_by_id = {c.get("id"): c for c in (template.get("columnas_finales") or [])}

    total_weighted = 0.0
    total_weight = 0.0
    for g in grupos:
        pct = g.get("porcentaje")
        try:
            weight = float(pct) / 100.0 if pct is not None else 0.0
        except (TypeError, ValueError):
            weight = 0.0
        if weight <= 0:
            continue

        member_values = []
        for mid in g.get("miembro_ids") or []:
            if mid in criterios_by_id:
                v = _criterio_avg(criterios_by_id[mid], grade, grades_dyn)
            elif mid in finales_by_id:
                col = finales_by_id[mid]
                v = _resolve_dynamic_value(col.get("id"), col.get("field_key"), grade, grades_dyn)
            else:
                v = None
            if v is not None:
                member_values.append(v)
        if not member_values:
            continue
        grupo_avg = sum(member_values) / len(member_values)
        total_weighted += grupo_avg * weight
        total_weight += weight

    if total_weight <= 0:
        return None
    result = total_weighted / total_weight if total_weight < 0.999 else total_weighted
    return round(result, 1)


def calculate_final_grade_from_template(grade: dict, template: dict) -> Optional[float]:
    """Compute final bimestral grade using a CUSTOM (dynamic) Registro Auxiliar template.

    Two weighting modes:
      - 'criterio' (default): each criterio carries its own porcentaje.
      - 'grupo': criterios/columnas finales are bundled into grupos, each grupo
        carries a shared porcentaje; the grupo value is the simple mean of its
        members' values.
    Returns None if there's no data. Normalizes when total weights sum to less
    than 1.0 so partial registers still produce a meaningful grade.
    """
    if not template:
        return None

    if (template.get("modo_ponderacion") == "grupo") and (template.get("grupos")):
        return _calculate_final_grade_grupo_mode(grade, template)

    criterios = template.get("criterios") or []
    if not criterios:
        return None

    grades_dyn = grade.get("grades_dynamic") or {}

    total_weighted = 0.0
    total_weight = 0.0

    for cri in criterios:
        pct = cri.get("porcentaje")
        try:
            weight = float(pct) / 100.0 if pct is not None else 0.0
        except (TypeError, ValueError):
            weight = 0.0
        if weight <= 0:
            continue

        crit_avg = _criterio_avg(cri, grade, grades_dyn)
        if crit_avg is None:
            continue
        total_weighted += crit_avg * weight
        total_weight += weight

    # Columnas finales (ej. EXAMEN MENSUAL/BIMESTRAL, PARC, ORAL): cada una pondera
    # su propio porcentaje con su valor directo. El frontend (calcularPromedioBimestral)
    # las incluye, así que el backend DEBE hacerlo también para que la nota final
    # coincida con el Registro Auxiliar.
    for col in template.get("columnas_finales") or []:
        pct = col.get("porcentaje")
        try:
            weight = float(pct) / 100.0 if pct is not None else 0.0
        except (TypeError, ValueError):
            weight = 0.0
        if weight <= 0:
            continue
        val = _resolve_dynamic_value(col.get("id"), col.get("field_key"), grade, grades_dyn)
        if val is None:
            continue
        total_weighted += val * weight
        total_weight += weight

    if total_weight <= 0:
        return None
    # Normalize partial pesos so a register with only some criterios still
    # produces a meaningful grade (same behaviour as the static algorithm).
    result = total_weighted / total_weight if total_weight < 0.999 else total_weighted
    return round(result, 1)


def calculate_final_grade(grade: dict, config: dict, template: Optional[dict] = None) -> Optional[float]:
    """Calculate final bimestral grade from sub-fields using weighted averages.

    If a CUSTOM (non-system) template is provided, delegate to the dynamic
    template-driven algorithm. Otherwise fall back to the legacy hardcoded
    static-fields algorithm (used by the Plantilla del Sistema).
    """
    if template and not template.get("es_sistema"):
        return calculate_final_grade_from_template(grade, template)

    # Legacy static-fields algorithm (Plantilla del Sistema)
    act_avg = _avg([grade.get("act_co"), grade.get("act_re")])
    rf_avg = _avg([grade.get("rf_r1"), grade.get("rf_r2"), grade.get("rf_r3"), grade.get("rf_r4"), grade.get("rf_r5")])
    comp_avg = _avg([grade.get("comp_c1"), grade.get("comp_c2")])
    part_avg = _avg([grade.get("part_p1"), grade.get("part_p2"), grade.get("part_p3"), grade.get("part_exp"), grade.get("part_tg"), grade.get("part_p")])
    exam_mens = grade.get("exam_mensual")
    exam_bim = grade.get("exam_bimestral")

    weighted = [
        (act_avg, config.get("attitude_weight", 0.10)),
        (rf_avg, config.get("worksheets_weight", 0.25)),
        (comp_avg, config.get("competency_weight", 0.05)),
        (part_avg, config.get("participation_weight", 0.25)),
        (exam_mens, config.get("monthly_exam_weight", 0.15)),
        (exam_bim, config.get("bimestral_exam_weight", 0.20)),
    ]

    total = 0.0
    total_weight = 0.0
    for val, weight in weighted:
        if val is not None:
            total += val * weight
            total_weight += weight
    if total_weight == 0:
        return None
    result = total / total_weight if total_weight < 1.0 else total
    return round(result, 1)


# ══════════════════════════════════════════════════════════════════════════════
# EVALUATION CONFIG
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/grades/config/{subject_id}/{section_id}")
async def get_eval_config(subject_id: str, section_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    _deny_students(user)
    school_id = user.get("school_id")

    config = await db.evaluation_config.find_one(
        {"school_id": school_id, "subject_id": subject_id, "section_id": section_id},
        {"_id": 0}
    )
    if not config:
        config = {
            "attitude_weight": 0.10,
            "worksheets_weight": 0.25,
            "competency_weight": 0.05,
            "participation_weight": 0.25,
            "monthly_exam_weight": 0.15,
            "bimestral_exam_weight": 0.20,
        }
    return config

@router.put("/grades/config/{subject_id}/{section_id}")
async def update_eval_config(subject_id: str, section_id: str, data: EvalConfigUpdate, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    _deny_students(user)
    check_demo_user_block(user)
    school_id = user.get("school_id")

    # Validate weights sum to 1.0
    total = (data.attitude_weight + data.worksheets_weight + data.competency_weight +
             data.participation_weight + data.monthly_exam_weight + data.bimestral_exam_weight)
    if abs(total - 1.0) > 0.01:
        raise HTTPException(status_code=400, detail=f"Los pesos deben sumar 100%. Actualmente suman {total*100:.0f}%")

    config_data = {
        "school_id": school_id,
        "subject_id": subject_id,
        "section_id": section_id,
        **data.model_dump(),
        "updated_at": now_iso(),
        "updated_by": user["id"],
    }

    await db.evaluation_config.update_one(
        {"school_id": school_id, "subject_id": subject_id, "section_id": section_id},
        {"$set": config_data, "$setOnInsert": {"id": generate_id(), "created_at": now_iso()}},
        upsert=True
    )
    return {"message": "Configuracion actualizada"}


# ══════════════════════════════════════════════════════════════════════════════
# GET REGISTER (Registro Auxiliar)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/grades/register/{subject_id}/{section_id}/{period_id}")
async def get_grade_register(subject_id: str, section_id: str, period_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    _deny_students(user)
    school_id = user.get("school_id")
    role = user.get("role")

    # Teachers can only see their own assignments. Permission is relaxed so a
    # subject linked to a different section doc than the teacher's assignment no
    # longer 403s.
    if role == "teacher":
        await _assert_teacher_assignment(school_id, user["id"], subject_id, section_id)

    # Resolve the section whose enrolled students actually belong to this course
    # (from the course's assignment), so the roster matches the dashboard card
    # and Usuarios/Estudiantes even when subject.section_id is swapped/mismatched.
    effective_section_id = await _resolve_effective_section_id(
        school_id, subject_id, section_id, role=role, teacher_id=user.get("id")
    )

    # Get students for the effective section only (no cross-section merging).
    students = await _fetch_section_students(school_id, effective_section_id)

    # Get existing grades
    grades = await db.student_grades.find(
        {"school_id": school_id, "subject_id": subject_id, "section_id": section_id, "period_id": period_id},
        {"_id": 0}
    ).to_list(200)
    grades_map = {g["student_id"]: g for g in grades}

    # Get eval config
    config = await db.evaluation_config.find_one(
        {"school_id": school_id, "subject_id": subject_id, "section_id": section_id},
        {"_id": 0}
    )
    if not config:
        config = {
            "attitude_weight": 0.10,
            "worksheets_weight": 0.25,
            "competency_weight": 0.05,
            "participation_weight": 0.25,
            "monthly_exam_weight": 0.15,
            "bimestral_exam_weight": 0.20,
        }

    # Get register status
    reg_status = await db.grade_register_status.find_one(
        {"school_id": school_id, "subject_id": subject_id, "section_id": section_id, "period_id": period_id},
        {"_id": 0}
    )

    # Load the active template once so the PROM. BIMESTRAL column resolves the
    # final grade with the SAME precedence used by the Consolidado:
    #   final_grade_manual > final_grade > on-the-fly recompute (custom template).
    # This prevents the historical bug where the Consolidado showed a grade
    # (entered via "Notas del Profesor / Manual de Notas" -> final_grade_manual)
    # while the Registro Auxiliar rendered the final column empty.
    from services.register_sync import get_active_template_for_school
    reg_template = await get_active_template_for_school(db, school_id)
    is_custom_template = bool(reg_template and not reg_template.get("es_sistema"))

    def _resolve_final(g: dict) -> Optional[float]:
        manual = g.get("final_grade_manual")
        if manual is not None:
            return manual
        final_val = g.get("final_grade")
        if is_custom_template and (
            g.get("grades_dynamic") or any(g.get(f) is not None for f in GRADE_SUB_FIELDS)
        ):
            try:
                recomputed = calculate_final_grade(g, config, template=reg_template)
                if recomputed is not None:
                    final_val = recomputed
            except Exception as e:
                logger.warning(f"[REGISTER] on-the-fly recompute failed student={g.get('student_id')}: {e}")
        return final_val

    # Build response
    student_grades = []
    for i, student in enumerate(students):
        g = grades_map.get(student["id"], {})
        entry = {
            "number": i + 1,
            "student_id": student["id"],
            "student_name": f"{student.get('last_name', '')} {student.get('name', '')}".strip(),
            # Actitudinal
            "act_co": g.get("act_co"),
            "act_re": g.get("act_re"),
            # Revisión Fichas
            "rf_r1": g.get("rf_r1"),
            "rf_r2": g.get("rf_r2"),
            "rf_r3": g.get("rf_r3"),
            "rf_r4": g.get("rf_r4"),
            "rf_r5": g.get("rf_r5"),
            # Competencia
            "comp_c1": g.get("comp_c1"),
            "comp_c2": g.get("comp_c2"),
            # Participaciones
            "part_p1": g.get("part_p1"),
            "part_p2": g.get("part_p2"),
            "part_p3": g.get("part_p3"),
            "part_exp": g.get("part_exp"),
            "part_tg": g.get("part_tg"),
            "part_p": g.get("part_p"),
            # Exámenes
            "exam_mensual": g.get("exam_mensual"),
            "exam_bimestral": g.get("exam_bimestral"),
            "final_grade": _resolve_final(g),
            # Surface the manual override so the gradebook can flag rows whose
            # final grade was entered via the "Notas del Profesor" portal and
            # therefore legitimately have empty detail cells.
            "final_grade_manual": g.get("final_grade_manual"),
            # Phase 5 — pass through the dynamic subdocument so the
            # gradebook can render custom-template columns read from
            # `grades_dynamic[column_id]`.
            "grades_dynamic": g.get("grades_dynamic") or {},
        }
        student_grades.append(entry)

    # Get subject and period info
    subject = await db.subjects.find_one({"id": subject_id}, {"_id": 0, "name": 1, "code": 1})
    period = await db.academic_periods.find_one({"id": period_id}, {"_id": 0, "nombre": 1})

    return {
        "students": student_grades,
        "config": {k: v for k, v in config.items() if k not in ("school_id", "subject_id", "section_id", "id", "created_at", "updated_at", "updated_by")},
        "status": reg_status.get("status", "open") if reg_status else "open",
        "locked_at": reg_status.get("locked_at") if reg_status else None,
        "locked_by_name": reg_status.get("locked_by_name") if reg_status else None,
        "subject_name": subject.get("name", "") if subject else "",
        "period_name": period.get("nombre", "") if period else "",
        # Single-source-of-truth mapping for the frontend so it can resolve
        # legacy plantillas (Plantilla del Sistema: sub.id="io" / "re" / "t1"
        # …) against the actual top-level fields where the notes live (act_co
        # / act_re / rf_r1 …). Custom modern plantillas with UUID-style ids
        # won't appear here and are read from `grades_dynamic[sub.id]`.
        "legacy_field_map": dict(COLUMN_FIELD_MAP),
    }


# ══════════════════════════════════════════════════════════════════════════════
# SAVE / AUTOSAVE GRADES
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/grades/save")
async def save_grades(data: GradeSaveRequest, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    _deny_students(user)
    check_demo_user_block(user)
    school_id = user.get("school_id")
    role = user.get("role")

    # Check if register is locked
    reg_status = await db.grade_register_status.find_one(
        {"school_id": school_id, "subject_id": data.subject_id, "section_id": data.section_id, "period_id": data.period_id}
    )
    if reg_status and reg_status.get("status") in ("closed", "approved"):
        if role not in ADMIN_ROLES:
            raise HTTPException(status_code=403, detail="El registro esta cerrado. Solo un administrador puede reabrirlo.")

    # Teachers can only save their own assignments (accepts sibling/duplicate
    # sections, consistent with the GET register guard).
    if role == "teacher":
        await _assert_teacher_assignment(school_id, user["id"], data.subject_id, data.section_id)

    # Get eval config
    config = await db.evaluation_config.find_one(
        {"school_id": school_id, "subject_id": data.subject_id, "section_id": data.section_id},
        {"_id": 0}
    )
    if not config:
        config = {
            "attitude_weight": 0.10, "worksheets_weight": 0.25, "competency_weight": 0.05,
            "participation_weight": 0.25, "monthly_exam_weight": 0.15, "bimestral_exam_weight": 0.20,
        }

    # Fetch the school's ACTIVE Registro Auxiliar template once. When it's a
    # CUSTOM template, `calculate_final_grade` needs it to know which dynamic
    # columns to aggregate. We resolve it once to avoid N queries inside the loop.
    from services.register_sync import get_active_template_for_school
    template = await get_active_template_for_school(db, school_id)
    is_custom_template = bool(template and not template.get("es_sistema"))

    saved_count = 0
    for entry in data.grades:
        grade_data = {}
        for field in GRADE_SUB_FIELDS:
            grade_data[field] = getattr(entry, field, None)

        # Validate grades are 0-20
        for field, val in grade_data.items():
            if val is not None and (val < 0 or val > 20):
                raise HTTPException(status_code=400, detail=f"Nota invalida para {field}: {val}. Debe ser entre 0 y 20")

        # Phase 5 — validate & merge dynamic columns into the `$set`
        # payload as dotted keys so only the touched cells are updated
        # (we never replace the whole subdocument, to avoid clobbering
        # other columns written by tasks/exams sync).
        set_payload = {}
        if entry.grades_dynamic:
            for col_id, val in entry.grades_dynamic.items():
                if val is not None and (val < 0 or val > 20):
                    raise HTTPException(
                        status_code=400,
                        detail=f"Nota invalida para columna '{col_id}': {val}. Debe ser entre 0 y 20",
                    )
                set_payload[f"grades_dynamic.{col_id}"] = val

        # Calculate final grade. For custom templates we need to merge the
        # incoming dynamic cells with what's already on disk so the criterio
        # averages are computed from the WHOLE register row, not just the
        # cells the teacher touched in this request.
        if is_custom_template:
            existing = await db.student_grades.find_one(
                {
                    "school_id": school_id,
                    "student_id": entry.student_id,
                    "subject_id": data.subject_id,
                    "section_id": data.section_id,
                    "period_id": data.period_id,
                },
                {"_id": 0},
            ) or {}
            # Legacy-formatted row detection: the doc on disk has NO dynamic
            # data and the incoming entry didn't send any either, yet flat
            # legacy fields are populated. This is the case of bimesters that
            # were graded before the school migrated to a custom template
            # (e.g. 1st bimester at Precursores TJ). For these rows the
            # template-driven algorithm would produce `None` because the
            # plantilla maps to columns that simply don't exist on the doc.
            # Compute final_grade via the legacy algorithm instead (no
            # changes to `calculate_final_grade` itself — just route to its
            # legacy branch by passing `template=None`).
            existing_dynamic = existing.get("grades_dynamic") or {}
            entry_has_dynamic = bool(entry.grades_dynamic) and any(
                v is not None for v in entry.grades_dynamic.values()
            )
            existing_has_dynamic = any(
                v is not None for v in existing_dynamic.values()
            )
            has_any_legacy = any(
                grade_data.get(f) is not None for f in GRADE_SUB_FIELDS
            ) or any(
                existing.get(f) is not None for f in GRADE_SUB_FIELDS
            )
            is_legacy_row = (
                not entry_has_dynamic
                and not existing_has_dynamic
                and has_any_legacy
            )
            if is_legacy_row:
                merged_legacy = {**existing, **grade_data}
                final = calculate_final_grade(merged_legacy, config, template=None)
            else:
                merged_dynamic = dict(existing_dynamic)
                if entry.grades_dynamic:
                    for k, v in entry.grades_dynamic.items():
                        merged_dynamic[k] = v
                merged_grade = {**existing, **grade_data, "grades_dynamic": merged_dynamic}
                final = calculate_final_grade(merged_grade, config, template=template)
        else:
            final = calculate_final_grade(grade_data, config, template=template)
        grade_data["final_grade"] = final
        grade_data["updated_at"] = now_iso()
        grade_data["updated_by"] = user["id"]

        set_payload.update(grade_data)

        # Defensive (Feb 2026): Some students may have multiple `student_grades`
        # docs for the same (school_id, student_id, subject_id, section_id,
        # period_id) tuple — likely caused by historical races between
        # `/save` and `/autosave` running concurrently without a unique index.
        # `update_one` would only touch ONE of those duplicates; the GET
        # endpoint would then read whichever doc the dict-comprehension
        # `{g["student_id"]: g for g in grades}` happened to keep last,
        # producing the silent "save returns 200 but value not persisted"
        # bug reported in production for puntual students.
        #
        # Fix: use `update_many` so ALL duplicate docs get the same `$set`
        # payload. Idempotent — no data destroyed, no schema change.
        # If duplicates are detected, log a warning so they can be cleaned
        # offline at a later time.
        filter_query = {
            "school_id": school_id,
            "student_id": entry.student_id,
            "subject_id": data.subject_id,
            "section_id": data.section_id,
            "period_id": data.period_id,
        }
        try:
            existing_count = await db.student_grades.count_documents(filter_query, limit=2)
            if existing_count > 1:
                logger.warning(
                    f"[GRADES SAVE] duplicate student_grades docs detected (count>=2): "
                    f"student={entry.student_id} subject={data.subject_id} "
                    f"section={data.section_id} period={data.period_id}"
                )
        except Exception:
            # Count is best-effort; never block the save on telemetry.
            pass

        await db.student_grades.update_many(
            filter_query,
            {
                "$set": set_payload,
                "$setOnInsert": {
                    "id": generate_id(),
                    "school_id": school_id,
                    "student_id": entry.student_id,
                    "subject_id": data.subject_id,
                    "section_id": data.section_id,
                    "period_id": data.period_id,
                    "teacher_id": user["id"],
                    "created_at": now_iso(),
                }
            },
            upsert=True
        )
        saved_count += 1

    return {"message": f"{saved_count} notas guardadas", "saved": saved_count}

@router.post("/grades/autosave")
async def autosave_grades(data: GradeSaveRequest, current_user=Depends(get_current_user)):
    """Same as save but with lighter response for auto-save"""
    return await save_grades(data, current_user)


# ══════════════════════════════════════════════════════════════════════════════
# LOCK / UNLOCK PERIOD
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/grades/lock_period")
async def lock_period(data: LockRequest, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    check_demo_user_block(user)
    school_id = user.get("school_id")

    await db.grade_register_status.update_one(
        {"school_id": school_id, "subject_id": data.subject_id, "section_id": data.section_id, "period_id": data.period_id},
        {
            "$set": {
                "status": "closed",
                "locked_at": now_iso(),
                "locked_by": user["id"],
                "locked_by_name": f"{user.get('name','')} {user.get('last_name','')}".strip(),
            },
            "$setOnInsert": {
                "id": generate_id(),
                "school_id": school_id,
                "subject_id": data.subject_id,
                "section_id": data.section_id,
                "period_id": data.period_id,
                "created_at": now_iso(),
            }
        },
        upsert=True
    )
    return {"message": "Registro cerrado exitosamente"}

@router.post("/grades/unlock_period")
async def unlock_period(data: LockRequest, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo un administrador puede reabrir el registro")
    school_id = user.get("school_id")

    await db.grade_register_status.update_one(
        {"school_id": school_id, "subject_id": data.subject_id, "section_id": data.section_id, "period_id": data.period_id},
        {"$set": {"status": "open", "unlocked_at": now_iso(), "unlocked_by": user["id"]}}
    )
    return {"message": "Registro reabierto exitosamente"}


# ══════════════════════════════════════════════════════════════════════════════
# CONSOLIDATED VIEW
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/grades/consolidated/{section_id}/{period_id}")
async def get_consolidated(section_id: str, period_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    _deny_students(user)
    school_id = user.get("school_id")

    # Get section info
    section = await db.sections.find_one({"id": section_id}, {"_id": 0})
    if not section:
        raise HTTPException(status_code=404, detail="Seccion no encontrada")

    grade_id = section.get("grado_id")

    # Get all subjects for this section
    subjects = await db.subjects.find(
        {"school_id": school_id, "section_id": section_id},
        {"_id": 0, "id": 1, "name": 1, "code": 1}
    ).sort("name", 1).to_list(50)

    if not subjects:
        # Fallback: get subjects for this grade
        subjects = await db.subjects.find(
            {"school_id": school_id, "grade_id": grade_id},
            {"_id": 0, "id": 1, "name": 1, "code": 1}
        ).sort("name", 1).to_list(50)

    subject_ids = [s["id"] for s in subjects]

    # Get students in section (try seccion_id first, then section_id)
    student_filter = {
        "school_id": school_id,
        "role": "student",
        "student_status": {"$in": ["enrolled", "active"]},
        "seccion_id": section_id,
    }
    students = await db.users.find(
        student_filter,
        {"_id": 0, "id": 1, "name": 1, "last_name": 1}
    ).sort([("last_name", 1), ("name", 1)]).to_list(200)

    if not students:
        student_filter["section_id"] = student_filter.pop("seccion_id")
        students = await db.users.find(
            student_filter,
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).sort([("last_name", 1), ("name", 1)]).to_list(200)

    # Get all grades for this section/period — include grades_dynamic so we
    # can recompute final_grade on-the-fly for legacy rows persisted before
    # the dynamic-template aggregation fix, and for rows written by
    # task/exam sync (which never touches final_grade).
    all_grades = await db.student_grades.find(
        {"school_id": school_id, "section_id": section_id, "period_id": period_id, "subject_id": {"$in": subject_ids}},
        {"_id": 0}
    ).to_list(5000)

    # If the school's active template is CUSTOM, pre-load it so we can
    # recompute final_grade for any row that's missing it. The system
    # template path keeps using whatever final_grade was persisted by save.
    from services.register_sync import get_active_template_for_school
    template = await get_active_template_for_school(db, school_id)
    is_custom_template = bool(template and not template.get("es_sistema"))

    # Build grades lookup: {student_id: {subject_id: final_grade}}
    # Teacher's manual override (final_grade_manual) takes precedence over the auto-computed
    # final_grade coming from the Registro Auxiliar.
    grades_lookup = {}
    for g in all_grades:
        sid = g["student_id"]
        if sid not in grades_lookup:
            grades_lookup[sid] = {}
        manual = g.get("final_grade_manual")
        if manual is not None:
            grades_lookup[sid][g["subject_id"]] = manual
            continue
        final_val = g.get("final_grade")
        # Para plantillas CUSTOM: recalcular SIEMPRE en vivo. Si el recálculo da
        # None (fila legacy), se conserva el valor almacenado.
        if is_custom_template and (g.get("grades_dynamic") or any(g.get(f) is not None for f in GRADE_SUB_FIELDS)):
            try:
                recomputed = calculate_final_grade(g, {}, template=template)
                if recomputed is not None:
                    final_val = recomputed
            except Exception as e:
                logger.warning(f"[CONSOLIDADO] on-the-fly recompute failed for student={sid} subj={g.get('subject_id')}: {e}")
        grades_lookup[sid][g["subject_id"]] = final_val

    # Build consolidated data
    consolidated = []
    for i, student in enumerate(students):
        student_data = {
            "number": i + 1,
            "student_id": student["id"],
            "student_name": f"{student.get('last_name', '')} {student.get('name', '')}".strip(),
            "grades": {},
            "average": None,
        }
        grades_for_student = grades_lookup.get(student["id"], {})
        valid_grades = []
        for subj in subjects:
            grade_val = grades_for_student.get(subj["id"])
            student_data["grades"][subj["id"]] = grade_val
            if grade_val is not None:
                valid_grades.append(grade_val)

        if valid_grades:
            student_data["average"] = round(sum(valid_grades) / len(valid_grades), 1)

        consolidated.append(student_data)

    # Sort by average descending for ranking
    ranked = sorted([s for s in consolidated if s["average"] is not None], key=lambda x: -x["average"])
    for rank, student in enumerate(ranked):
        student["rank"] = rank + 1
    # Students without average get no rank
    for student in consolidated:
        if "rank" not in student:
            student["rank"] = None

    # Get period and section names
    period = await db.academic_periods.find_one({"id": period_id}, {"_id": 0, "nombre": 1})
    grade = await db.grades.find_one({"id": grade_id}, {"_id": 0, "nombre": 1})

    return {
        "students": consolidated,
        "subjects": [{"id": s["id"], "name": s["name"], "code": s.get("code", "")} for s in subjects],
        "section_name": section.get("nombre", ""),
        "grade_name": grade.get("nombre", "") if grade else "",
        "period_name": period.get("nombre", "") if period else "",
    }


# ══════════════════════════════════════════════════════════════════════════════
# CONSOLIDATED REPORT (Replica fiel del Excel)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/grades/consolidated-report/{section_id}/{period_id}")
async def get_consolidated_report(section_id: str, period_id: str, current_user=Depends(get_current_user)):
    """
    Returns the full consolidated report data matching the Excel format exactly.
    Includes: school info, section/period info, subjects grouped by area,
    student grades, computed columns, and summary statistics.
    """
    user = await resolve_user_from_token(current_user)
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    _deny_students(user)
    school_id = user.get("school_id")

    # Get school info
    school = await db.schools.find_one({"id": school_id}, {"_id": 0, "name": 1, "subdomain": 1})
    school_name = school.get("name", "") if school else ""

    # Get section info
    section = await db.sections.find_one({"id": section_id}, {"_id": 0})
    if not section:
        raise HTTPException(status_code=404, detail="Seccion no encontrada")
    grade_id = section.get("grado_id")

    # Get grade info
    grade_doc = await db.grades.find_one({"id": grade_id}, {"_id": 0, "nombre": 1, "nivel_id": 1})
    grade_name = grade_doc.get("nombre", "") if grade_doc else ""
    level_id = grade_doc.get("nivel_id") if grade_doc else None

    # Get level info
    level_doc = await db.academic_levels.find_one({"id": level_id}, {"_id": 0, "nombre": 1}) if level_id else None
    level_name = level_doc.get("nombre", "") if level_doc else ""

    # Get period info
    period = await db.academic_periods.find_one({"id": period_id}, {"_id": 0, "nombre": 1})
    period_name = period.get("nombre", "") if period else ""

    # Get academic year
    year_doc = await db.academic_years.find_one({"school_id": school_id}, {"_id": 0, "year": 1})
    school_year = year_doc.get("year", datetime.now().year) if year_doc else datetime.now().year

    # Get tutor for this section
    tutor_assignment = await db.academic_assignments.find_one(
        {"school_id": school_id, "section_id": section_id, "role": "tutor", "status": "activo"},
        {"_id": 0, "teacher_id": 1}
    )
    tutor_name = ""
    if tutor_assignment:
        tutor = await db.users.find_one({"id": tutor_assignment["teacher_id"]}, {"_id": 0, "name": 1, "last_name": 1})
        if tutor:
            tutor_name = f"{tutor.get('last_name', '')}, {tutor.get('name', '')}".strip(", ")

    # Get all subjects for this section
    subjects = await db.subjects.find(
        {"school_id": school_id, "section_id": section_id, "status": {"$ne": "inactive"}},
        {"_id": 0, "id": 1, "name": 1, "code": 1, "area_name": 1, "area_order": 1}
    ).sort("name", 1).to_list(100)

    if not subjects:
        subjects = await db.subjects.find(
            {"school_id": school_id, "grade_id": grade_id, "status": {"$ne": "inactive"}},
            {"_id": 0, "id": 1, "name": 1, "code": 1, "area_name": 1, "area_order": 1}
        ).sort("name", 1).to_list(100)

    # Build column structure: group subjects by area
    # Columns: each has {id, name, type: 'area'|'subject', area_name, subjects_in_area}
    columns = []
    area_groups = {}
    standalone = []

    for s in subjects:
        area = s.get("area_name")
        if area:
            if area not in area_groups:
                area_groups[area] = {
                    "area_name": area,
                    "area_order": s.get("area_order", 999),
                    "subjects": []
                }
            area_groups[area]["subjects"].append(s)
        else:
            standalone.append(s)

    # Sort area groups by area_order, then alphabetically
    sorted_areas = sorted(area_groups.values(), key=lambda a: (a["area_order"], a["area_name"]))

    for area_group in sorted_areas:
        # Add area column (computed average)
        columns.append({
            "id": f"area_{area_group['area_name']}",
            "name": area_group["area_name"],
            "type": "area",
            "subject_ids": [s["id"] for s in area_group["subjects"]]
        })
        # Add sub-subject columns
        for s in area_group["subjects"]:
            columns.append({
                "id": s["id"],
                "name": s["name"],
                "type": "subject",
                "area_name": area_group["area_name"]
            })

    # Add standalone subjects
    for s in standalone:
        columns.append({
            "id": s["id"],
            "name": s["name"],
            "type": "subject",
            "area_name": None
        })

    subject_ids = [s["id"] for s in subjects]

    # Get students
    student_filter = {
        "school_id": school_id,
        "role": "student",
        "student_status": {"$in": ["enrolled", "active"]},
        "seccion_id": section_id,
    }
    students = await db.users.find(
        student_filter,
        {"_id": 0, "id": 1, "name": 1, "last_name": 1}
    ).sort([("last_name", 1), ("name", 1)]).to_list(200)

    if not students:
        student_filter["section_id"] = student_filter.pop("seccion_id")
        students = await db.users.find(
            student_filter,
            {"_id": 0, "id": 1, "name": 1, "last_name": 1}
        ).sort([("last_name", 1), ("name", 1)]).to_list(200)

    # Get all grades — include full row so we can recompute final_grade
    # on-the-fly for legacy rows / dynamic-template schools.
    all_grades = await db.student_grades.find(
        {"school_id": school_id, "section_id": section_id, "period_id": period_id, "subject_id": {"$in": subject_ids}},
        {"_id": 0}
    ).to_list(10000)

    # If the active template is CUSTOM, recompute final_grade for any row
    # missing it (legacy rows or rows updated by task/exam sync).
    from services.register_sync import get_active_template_for_school
    template = await get_active_template_for_school(db, school_id)
    is_custom_template = bool(template and not template.get("es_sistema"))

    # Teacher manual override (final_grade_manual) takes precedence over auto-computed final_grade.
    grades_lookup = {}
    real_lookup = {}  # subject_ids con datos REALES (no solo un final_grade viejo)
    for g in all_grades:
        sid = g["student_id"]
        if sid not in grades_lookup:
            grades_lookup[sid] = {}
            real_lookup[sid] = set()
        has_data = bool(g.get("grades_dynamic")) or any(g.get(f) is not None for f in GRADE_SUB_FIELDS)
        manual = g.get("final_grade_manual")
        if manual is not None:
            grades_lookup[sid][g["subject_id"]] = manual
            real_lookup[sid].add(g["subject_id"])
            continue
        final_val = g.get("final_grade")
        # Para plantillas CUSTOM: recalcular SIEMPRE en vivo (el valor almacenado
        # puede ser antiguo/con otra regla de redondeo). Si el recálculo da None
        # (fila legacy sin datos dinámicos), se conserva el valor almacenado.
        if is_custom_template and (g.get("grades_dynamic") or any(g.get(f) is not None for f in GRADE_SUB_FIELDS)):
            try:
                recomputed = calculate_final_grade(g, {}, template=template)
                if recomputed is not None:
                    final_val = recomputed
            except Exception as e:
                logger.warning(f"[CONSOLIDADO-REPORT] on-the-fly recompute failed for student={sid} subj={g.get('subject_id')}: {e}")
        grades_lookup[sid][g["subject_id"]] = final_val
        # "Real" = sistema (siempre) o custom con datos. Una asignatura duplicada/vacía
        # (custom sin datos, solo con final_grade viejo) NO cuenta para el promedio de área.
        if (not is_custom_template) or has_data:
            real_lookup[sid].add(g["subject_id"])

    # Build student rows with computed fields
    # Ranking, puntaje, promedio, tercio y desaprobados provienen del helper
    # compartido `services.ranking.compute_ranking` (Fase 2 — Turno A).
    ranking_map = await compute_ranking(db, school_id, section_id, period_id)

    # Conducta del bimestre (la nota que coloca el tutor). Se guarda en
    # `conduct_grades` por {student_id, period_id}. La leemos por alumno para
    # que la columna CONDUCTA del consolidado deje de salir vacía.
    student_ids = [s["id"] for s in students]
    conduct_docs = await db.conduct_grades.find(
        {"school_id": school_id, "period_id": period_id, "student_id": {"$in": student_ids}},
        {"_id": 0, "student_id": 1, "letra": 1, "score_numeric": 1},
    ).to_list(1000)
    conduct_map = {c["student_id"]: c for c in conduct_docs}

    student_rows = []
    for i, student in enumerate(students):
        student_grades = grades_lookup.get(student["id"], {})
        rinfo = ranking_map.get(student["id"], {
            "puntaje": None, "promedio": None,
            "orden_merito": None, "tercio": None, "cursos_desaprobados": 0,
        })
        row = {
            "number": i + 1,
            "student_id": student["id"],
            "student_name": f"{student.get('last_name', '')} {student.get('name', '')}".strip(),
            "grades": {},
            "conducta": (conduct_map.get(student["id"]) or {}).get("letra"),
            "promedio": rinfo.get("promedio"),
            "puntaje": rinfo.get("puntaje"),
            "n_desaprobados": rinfo.get("cursos_desaprobados", 0),
            "orden_merito": rinfo.get("orden_merito"),
            "tercio": rinfo.get("tercio"),
            "tardanza_injustificada": None,
            "tardanza_justificada": None,
            "falta_injustificada": None,
            "falta_justificada": None,
        }

        # Construir columnas de display (area summary + subject grades)
        # — independiente del ranking, solo para la grilla del consolidado.
        for col in columns:
            if col["type"] == "area":
                real_set = real_lookup.get(student["id"], set())
                sub_grades = [student_grades.get(sid) for sid in col["subject_ids"] if sid in real_set]
                valid = [g for g in sub_grades if g is not None]
                area_avg = round(sum(valid) / len(valid), 0) if valid else None
                row["grades"][col["id"]] = int(area_avg) if area_avg is not None else None
            elif col["type"] == "subject":
                real_set = real_lookup.get(student["id"], set())
                grade_val = student_grades.get(col["id"]) if col["id"] in real_set else None
                if grade_val is not None:
                    grade_val = round(grade_val)
                row["grades"][col["id"]] = int(grade_val) if grade_val is not None else None

        student_rows.append(row)

    # Compute summary statistics per subject column
    summary_stats = {}
    for col in columns:
        col_grades = []
        for row in student_rows:
            val = row["grades"].get(col["id"])
            if val is not None:
                col_grades.append(val)

        aprobados = sum(1 for g in col_grades if g >= 11)
        desaprobados = sum(1 for g in col_grades if g < 11)
        total = len(col_grades)

        summary_stats[col["id"]] = {
            "promedio": round(sum(col_grades) / len(col_grades), 1) if col_grades else None,
            "aprobados": aprobados if total > 0 else None,
            "desaprobados": desaprobados if total > 0 else None,
            "pct_aprobados": round(aprobados / total * 100) if total > 0 else None,
            "pct_desaprobados": round(desaprobados / total * 100) if total > 0 else None,
            "nota_maxima": max(col_grades) if col_grades else None,
            "nota_minima": min(col_grades) if col_grades else None,
        }

    # Build section display name
    section_display = f"{grade_name} {section.get('nombre', '')} {level_name}".strip()

    return {
        "school_name": school_name,
        "system_name": "CUBICOL Intranet",
        "title": f"CONSOLIDADO DE NOTAS - {school_year}",
        "section_display": section_display,
        "period_name": period_name,
        "tutor_name": tutor_name,
        "school_year": school_year,
        "columns": columns,
        "students": student_rows,
        "summary_stats": summary_stats,
        "total_students": len(students),
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT CONSOLIDATED REPORT (Excel - Replica fiel)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/grades/consolidated-report/{section_id}/{period_id}/export/excel")
async def export_consolidated_report_excel(section_id: str, period_id: str, current_user=Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    import io

    data = await get_consolidated_report(section_id, period_id, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    # Styles
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    area_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    columns = data["columns"]
    # Fixed columns: N°(A), APELLIDOS Y NOMBRES(B-C merged conceptually, using B with wide width)
    # Subject columns start at D
    subject_start_col = 4  # D

    # Row 1: School name
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.cell(row=1, column=1, value=data["school_name"].upper()).font = Font(bold=True, size=11)
    ws.cell(row=1, column=12, value="Fecha:").font = Font(bold=True, size=9)
    ws.cell(row=1, column=14, value=datetime.now().strftime("%d/%m/%Y")).font = Font(size=9)

    # Row 2: System name
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws.cell(row=2, column=1, value=data["system_name"]).font = Font(bold=True, size=10)
    ws.cell(row=2, column=12, value="Hora:").font = Font(bold=True, size=9)
    ws.cell(row=2, column=14, value=datetime.now().strftime("%H:%M:%S")).font = Font(size=9)

    # Row 3-4: Title
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=15)
    ws.cell(row=3, column=1, value=data["title"]).font = Font(bold=True, size=12)
    ws.cell(row=3, column=1).alignment = center

    # Row 5: Context
    ws.cell(row=5, column=1, value="Salón:").font = Font(bold=True, size=9)
    ws.cell(row=5, column=3, value=data["section_display"]).font = Font(size=9)
    ws.cell(row=5, column=4, value="Periodo:").font = Font(bold=True, size=9)
    ws.cell(row=5, column=6, value=data["period_name"]).font = Font(size=9)
    ws.cell(row=5, column=9, value="Tutor:").font = Font(bold=True, size=9)
    ws.cell(row=5, column=11, value=data["tutor_name"]).font = Font(size=9)

    # Row 6: Area headers + subject headers
    # First merge A6:C6 for "ASIGNATURAS"
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=3)
    cell = ws.cell(row=6, column=1, value="ASIGNATURAS")
    cell.font = Font(bold=True, size=9)
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

    col_idx = subject_start_col
    for c in columns:
        cell = ws.cell(row=6, column=col_idx, value=c["name"])
        cell.font = Font(bold=True, size=8)
        cell.alignment = center
        cell.border = thin_border
        if c["type"] == "area":
            cell.fill = area_fill
        else:
            cell.fill = header_fill
        # Merge vertically with row 7
        ws.merge_cells(start_row=6, start_column=col_idx, end_row=7, end_column=col_idx)
        col_idx += 1

    # Summary column headers
    summary_headers = ["CONDUCTA", "PROMEDIO", "PUNTAJE", "N° DESAPROBADOS", "ORDEN DE MÉRITO", "TERCIO",
                       "Tardanza Injustificada", "Tardanza Justificada", "Falta Injustificada", "Falta Justificada"]
    for h in summary_headers:
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = Font(bold=True, size=7)
        cell.alignment = center
        cell.border = thin_border
        cell.fill = summary_fill
        ws.merge_cells(start_row=6, start_column=col_idx, end_row=7, end_column=col_idx)
        col_idx += 1

    # Row 7: N° and APELLIDOS Y NOMBRES
    cell = ws.cell(row=7, column=1, value="N°")
    cell.font = Font(bold=True, size=9)
    cell.alignment = center
    cell.border = thin_border
    cell.fill = header_fill

    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=3)
    cell = ws.cell(row=7, column=2, value="APELLIDOS Y NOMBRES")
    cell.font = Font(bold=True, size=9)
    cell.alignment = center
    cell.border = thin_border
    cell.fill = header_fill

    # Student data rows
    data_start_row = 8
    for row_idx, student in enumerate(data["students"]):
        r = data_start_row + row_idx
        ws.cell(row=r, column=1, value=student["number"]).border = thin_border
        ws.cell(row=r, column=1).alignment = center

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2, value=student["student_name"]).border = thin_border
        ws.cell(row=r, column=3).border = thin_border

        col_idx = subject_start_col
        for c in columns:
            val = student["grades"].get(c["id"])
            cell = ws.cell(row=r, column=col_idx, value=val if val is not None else "")
            cell.alignment = center
            cell.border = thin_border
            if val is not None and val < 11:
                cell.font = Font(color="FF0000")
            col_idx += 1

        # Summary columns
        ws.cell(row=r, column=col_idx, value=student.get("conducta", "")).border = thin_border
        ws.cell(row=r, column=col_idx).alignment = center
        col_idx += 1

        cell = ws.cell(row=r, column=col_idx, value=student.get("promedio"))
        cell.border = thin_border
        cell.alignment = center
        cell.font = Font(bold=True)
        col_idx += 1

        ws.cell(row=r, column=col_idx, value=student.get("puntaje")).border = thin_border
        ws.cell(row=r, column=col_idx).alignment = center
        col_idx += 1

        ws.cell(row=r, column=col_idx, value=student.get("n_desaprobados") or "").border = thin_border
        ws.cell(row=r, column=col_idx).alignment = center
        col_idx += 1

        ws.cell(row=r, column=col_idx, value=student.get("orden_merito")).border = thin_border
        ws.cell(row=r, column=col_idx).alignment = center
        col_idx += 1

        ws.cell(row=r, column=col_idx, value=student.get("tercio", "")).border = thin_border
        ws.cell(row=r, column=col_idx).alignment = center
        col_idx += 1

        for field in ["tardanza_injustificada", "tardanza_justificada", "falta_injustificada", "falta_justificada"]:
            ws.cell(row=r, column=col_idx, value=student.get(field, "")).border = thin_border
            ws.cell(row=r, column=col_idx).alignment = center
            col_idx += 1

    # Summary footer rows
    footer_start = data_start_row + len(data["students"]) + 1
    summary_labels = [
        "Promedio del curso: ",
        "N° de alumnos Aprobados: ",
        "N° de alumnos Desaprobados: ",
        "% de alumnos Aprobados: ",
        "% de alumnos Desaprobados: ",
        "Nota Máxima: ",
        "Nota Mínima: ",
    ]
    summary_keys = ["promedio", "aprobados", "desaprobados", "pct_aprobados", "pct_desaprobados", "nota_maxima", "nota_minima"]

    for label_idx, (label, key) in enumerate(zip(summary_labels, summary_keys)):
        r = footer_start + label_idx
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2, value=label).font = Font(bold=True, size=8)
        ws.cell(row=r, column=2).border = thin_border

        col_idx = subject_start_col
        for c in columns:
            stats = data["summary_stats"].get(c["id"], {})
            val = stats.get(key)
            cell = ws.cell(row=r, column=col_idx, value=val if val is not None else "")
            cell.alignment = center
            cell.border = thin_border
            cell.font = Font(size=8)
            col_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 27
    for i in range(subject_start_col, col_idx + 1):
        ws.column_dimensions[get_column_letter(i)].width = 6

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"consolidado_{data['section_display']}_{data['period_name']}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT CONSOLIDATED (Excel) - Legacy
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/grades/consolidated/{section_id}/{period_id}/export/excel")
async def export_consolidated_excel(section_id: str, period_id: str, current_user=Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    data = await get_consolidated(section_id, period_id, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    green_fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(data["subjects"]) + 2)
    title_cell = ws.cell(row=1, column=1, value=f"CONSOLIDADO DE NOTAS - {data['grade_name']} {data['section_name']} - {data['period_name']}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center

    # Headers
    row = 3
    headers = ["N°", "Apellidos y Nombres"] + [s["name"] for s in data["subjects"]] + ["Promedio", "Puesto"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # Data
    for student in data["students"]:
        row += 1
        ws.cell(row=row, column=1, value=student["number"]).border = thin_border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=student["student_name"]).border = thin_border

        for j, subj in enumerate(data["subjects"]):
            grade_val = student["grades"].get(subj["id"])
            cell = ws.cell(row=row, column=3 + j, value=grade_val)
            cell.alignment = center
            cell.border = thin_border
            if grade_val is not None:
                if grade_val < 10:
                    cell.fill = red_fill
                elif grade_val <= 13:
                    cell.fill = yellow_fill
                else:
                    cell.fill = green_fill

        avg_col = 3 + len(data["subjects"])
        avg_cell = ws.cell(row=row, column=avg_col, value=student["average"])
        avg_cell.alignment = center
        avg_cell.border = thin_border
        avg_cell.font = Font(bold=True)
        if student["average"] is not None:
            if student["average"] < 10:
                avg_cell.fill = red_fill
            elif student["average"] <= 13:
                avg_cell.fill = yellow_fill
            else:
                avg_cell.fill = green_fill

        rank_cell = ws.cell(row=row, column=avg_col + 1, value=student["rank"])
        rank_cell.alignment = center
        rank_cell.border = thin_border
        rank_cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    for i in range(len(data["subjects"]) + 2):
        col_letter = chr(67 + i) if i < 24 else chr(65) + chr(65 + i - 24)
        try:
            ws.column_dimensions[col_letter].width = 14
        except Exception:
            pass

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"consolidado_{data['grade_name']}_{data['section_name']}_{data['period_name']}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ══════════════════════════════════════════════════════════════════════════════
# DUPLICATE-DOCS MAINTENANCE (owner-only)
# ──────────────────────────────────────────────────────────────────────────────
# Toolkit to clean up historical duplicates in `student_grades` caused by
# concurrent /save and /autosave inserts before a unique index existed.
#
# Three explicit steps, called manually by the school owner:
#   1) GET  /api/grades/_maintenance/duplicates/scan        (read-only)
#   2) POST /api/grades/_maintenance/duplicates/consolidate (dry_run=true by default)
#   3) POST /api/grades/_maintenance/duplicates/create-index
#
# All operations are SCOPED to the caller's school_id. The consolidation
# strategy merges non-null values across duplicates (preferring the most
# recently updated value) before deleting the obsolete docs. No grade is
# ever lost.
# ══════════════════════════════════════════════════════════════════════════════

_DUP_GROUP_KEY = ["school_id", "student_id", "subject_id", "section_id", "period_id"]


def _doc_score(doc: dict) -> tuple:
    """Sort key — most recently updated wins, then most non-null grade fields."""
    upd = doc.get("updated_at") or doc.get("created_at") or ""
    non_null = sum(
        1 for f in GRADE_SUB_FIELDS if doc.get(f) is not None
    ) + len([k for k, v in (doc.get("grades_dynamic") or {}).items() if v is not None])
    return (upd, non_null)


def _merge_docs(docs: list) -> dict:
    """Merge a list of duplicate docs into one canonical doc.

    Strategy: start from the most recently updated doc, then fill any null
    static field with a non-null value from the other docs. For grades_dynamic,
    union all keys preferring values from the most recent doc.
    Never overwrites an existing non-null value with another non-null value
    unless the donor doc is strictly newer than the keeper.
    """
    if not docs:
        return {}
    # Sort newest first
    ordered = sorted(docs, key=_doc_score, reverse=True)
    keeper = dict(ordered[0])
    keeper_grades_dyn = dict(keeper.get("grades_dynamic") or {})

    for donor in ordered[1:]:
        for f in GRADE_SUB_FIELDS:
            if keeper.get(f) is None and donor.get(f) is not None:
                keeper[f] = donor[f]
        donor_gd = donor.get("grades_dynamic") or {}
        for k, v in donor_gd.items():
            if v is not None and keeper_grades_dyn.get(k) is None:
                keeper_grades_dyn[k] = v
    keeper["grades_dynamic"] = keeper_grades_dyn
    return keeper


async def _scan_duplicates_for_school(school_id: str) -> dict:
    """Returns counts + sample groups of duplicate `student_grades` docs."""
    pipeline = [
        {"$match": {"school_id": school_id}},
        {"$group": {
            "_id": {k: f"${k}" for k in _DUP_GROUP_KEY},
            "count": {"$sum": 1},
            "doc_ids": {"$push": "$id"},
        }},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"count": -1}},
    ]
    groups = await db.student_grades.aggregate(pipeline).to_list(10000)
    total_dup_docs = sum(g["count"] for g in groups)
    extra_docs = sum(g["count"] - 1 for g in groups)  # how many would be deleted
    sample = [
        {
            "student_id": g["_id"].get("student_id"),
            "subject_id": g["_id"].get("subject_id"),
            "section_id": g["_id"].get("section_id"),
            "period_id": g["_id"].get("period_id"),
            "count": g["count"],
        }
        for g in groups[:25]
    ]
    return {
        "duplicate_groups": len(groups),
        "total_duplicate_docs": total_dup_docs,
        "docs_that_would_be_deleted": extra_docs,
        "sample_groups": sample,
    }


@router.get("/grades/_maintenance/duplicates/scan")
async def scan_duplicate_grades(current_user=Depends(get_current_user)):
    """Read-only — counts duplicate `student_grades` docs for the caller's school."""
    user = await resolve_user_from_token(current_user)
    if not user or user.get("role") not in {"owner", "director"}:
        raise HTTPException(403, "Sólo el owner/director puede ejecutar esta operación")
    school_id = user.get("school_id")
    if not school_id:
        raise HTTPException(400, "Tu usuario no tiene colegio asignado")
    report = await _scan_duplicates_for_school(school_id)
    logger.info(f"[GRADES MAINT] scan school={school_id} groups={report['duplicate_groups']} docs_to_delete={report['docs_that_would_be_deleted']}")
    return report


class ConsolidateRequest(BaseModel):
    dry_run: bool = True
    confirm_token: Optional[str] = None  # required when dry_run=false


@router.post("/grades/_maintenance/recompute-finals")
async def recompute_final_grades(
    dry_run: bool = True,
    period_id: Optional[str] = None,
    current_user=Depends(get_current_user),
):
    """Recalcula y PERSISTE `final_grade` de todas las filas (plantilla custom, sin
    override manual) usando la fórmula vigente, para que TODAS las vistas (incluidas
    exportaciones que leen el valor guardado) muestren la nota correcta.

    dry_run=true (por defecto): solo reporta cuántas cambiarían, sin escribir.
    dry_run=false: aplica los cambios.
    """
    user = await resolve_user_from_token(current_user)
    is_owner = user and (user.get("role") in {"owner", "director"} or user.get("is_owner") is True)
    is_support = user and (user.get("is_support_session") is True or user.get("is_super_admin") is True)
    if not (is_owner or is_support):
        raise HTTPException(403, "Sólo el owner/soporte puede ejecutar esta operación")
    school_id = user.get("school_id")
    if not school_id:
        raise HTTPException(400, "Tu usuario no tiene colegio asignado")

    from services.register_sync import get_active_template_for_school
    template = await get_active_template_for_school(db, school_id)
    if not template or template.get("es_sistema"):
        return {"ok": True, "message": "El colegio usa la Plantilla del Sistema; no requiere recálculo.",
                "checked": 0, "changed": 0, "dry_run": dry_run}

    q = {"school_id": school_id}
    if period_id:
        q["period_id"] = period_id

    checked = 0
    changed = 0
    samples = []
    ops = []
    cursor = db.student_grades.find(q, {"_id": 0})
    async for g in cursor:
        checked += 1
        if g.get("final_grade_manual") is not None:
            continue
        has_data = g.get("grades_dynamic") or any(g.get(f) is not None for f in GRADE_SUB_FIELDS)
        if not has_data:
            continue
        try:
            recomputed = calculate_final_grade(g, {}, template=template)
        except Exception:
            continue
        if recomputed is None:
            continue
        old = g.get("final_grade")
        if old is None or abs(float(old) - float(recomputed)) > 0.001:
            changed += 1
            if len(samples) < 20:
                samples.append({
                    "student_id": g.get("student_id"),
                    "subject_id": g.get("subject_id"),
                    "period_id": g.get("period_id"),
                    "old": old, "new": recomputed,
                    "old_display": (round(old) if isinstance(old, (int, float)) else None),
                    "new_display": round(recomputed),
                })
            if not dry_run:
                ops.append({"student_id": g.get("student_id"), "subject_id": g.get("subject_id"),
                            "period_id": g.get("period_id"), "final_grade": recomputed})

    if not dry_run and ops:
        for op in ops:
            await db.student_grades.update_one(
                {"school_id": school_id, "student_id": op["student_id"],
                 "subject_id": op["subject_id"], "period_id": op["period_id"]},
                {"$set": {"final_grade": op["final_grade"]}},
            )
        logger.info(f"[RECOMPUTE-FINALS] school={school_id} changed={changed} applied")

    # Resolve names for the sample rows (only up to 20)
    stu_ids = list({s["student_id"] for s in samples if s.get("student_id")})
    subj_ids = list({s["subject_id"] for s in samples if s.get("subject_id")})
    per_ids = list({s["period_id"] for s in samples if s.get("period_id")})
    stu_map = {u["id"]: f"{u.get('last_name','')} {u.get('name','')}".strip()
               for u in await db.users.find({"id": {"$in": stu_ids}}, {"_id": 0, "id": 1, "name": 1, "last_name": 1}).to_list(50)}
    subj_map = {s["id"]: s.get("name") for s in await db.subjects.find({"id": {"$in": subj_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(200)}
    per_map = {p["id"]: p.get("nombre") for p in await db.academic_periods.find({"id": {"$in": per_ids}}, {"_id": 0, "id": 1, "nombre": 1}).to_list(50)}
    for s in samples:
        s["student_name"] = stu_map.get(s.get("student_id"), s.get("student_id"))
        s["subject_name"] = subj_map.get(s.get("subject_id"), s.get("subject_id"))
        s["period_name"] = per_map.get(s.get("period_id"), "")

    return {"ok": True, "dry_run": dry_run, "checked": checked, "changed": changed,
            "applied": (0 if dry_run else changed), "samples": samples}



@router.post("/grades/_maintenance/duplicates/consolidate")
async def consolidate_duplicate_grades(payload: ConsolidateRequest, current_user=Depends(get_current_user)):
    """Merge duplicates into one canonical doc per group, then delete the rest.

    Default behaviour: dry_run=true → reports what WOULD happen, changes nothing.
    To actually run: dry_run=false AND confirm_token="CONSOLIDATE_<school_id>".
    """
    user = await resolve_user_from_token(current_user)
    if not user or user.get("role") not in {"owner", "director"}:
        raise HTTPException(403, "Sólo el owner/director puede ejecutar esta operación")
    school_id = user.get("school_id")
    if not school_id:
        raise HTTPException(400, "Tu usuario no tiene colegio asignado")

    pipeline = [
        {"$match": {"school_id": school_id}},
        {"$group": {
            "_id": {k: f"${k}" for k in _DUP_GROUP_KEY},
            "count": {"$sum": 1},
        }},
        {"$match": {"count": {"$gt": 1}}},
    ]
    groups = await db.student_grades.aggregate(pipeline).to_list(10000)

    expected_token = f"CONSOLIDATE_{school_id}"
    will_execute = (not payload.dry_run) and payload.confirm_token == expected_token

    if (not payload.dry_run) and not will_execute:
        raise HTTPException(
            400,
            f"Para ejecutar la consolidación pasá confirm_token='{expected_token}' (dry_run=false)"
        )

    summary = {
        "mode": "execute" if will_execute else "dry_run",
        "duplicate_groups_found": len(groups),
        "groups_processed": 0,
        "docs_kept": 0,
        "docs_deleted": 0,
        "groups_with_value_merges": 0,
        "errors": [],
    }

    for g in groups:
        flt = g["_id"]
        try:
            docs = await db.student_grades.find(flt, {"_id": 0}).to_list(20)
            if len(docs) < 2:
                continue
            merged = _merge_docs(docs)
            # Detect merges that actually pulled values from non-keeper docs
            ordered = sorted(docs, key=_doc_score, reverse=True)
            keeper_original = ordered[0]
            value_merge = False
            for f in GRADE_SUB_FIELDS:
                if (keeper_original.get(f) is None) and (merged.get(f) is not None):
                    value_merge = True
                    break
            if not value_merge:
                ko_gd = keeper_original.get("grades_dynamic") or {}
                m_gd = merged.get("grades_dynamic") or {}
                for k, v in m_gd.items():
                    if ko_gd.get(k) is None and v is not None:
                        value_merge = True
                        break
            if value_merge:
                summary["groups_with_value_merges"] += 1

            summary["groups_processed"] += 1
            summary["docs_kept"] += 1
            summary["docs_deleted"] += len(docs) - 1

            if will_execute:
                keeper_id = keeper_original.get("id")
                # Apply merged values to the keeper doc
                update_doc = {f: merged.get(f) for f in GRADE_SUB_FIELDS}
                update_doc["grades_dynamic"] = merged.get("grades_dynamic") or {}
                update_doc["final_grade"] = merged.get("final_grade")
                update_doc["updated_at"] = now_iso()
                update_doc["updated_by"] = user["id"]
                await db.student_grades.update_one({"id": keeper_id}, {"$set": update_doc})
                # Delete the obsolete duplicates
                other_ids = [d.get("id") for d in docs if d.get("id") and d.get("id") != keeper_id]
                if other_ids:
                    await db.student_grades.delete_many({"id": {"$in": other_ids}})
        except Exception as e:
            logger.exception(f"[GRADES MAINT] error procesando grupo {flt}: {e}")
            summary["errors"].append({"group": flt, "error": str(e)[:200]})

    logger.info(f"[GRADES MAINT] consolidate school={school_id} mode={summary['mode']} groups={summary['groups_processed']} deleted={summary['docs_deleted']}")
    if will_execute and not summary["errors"]:
        # Verify: count remaining duplicate groups after the operation
        verify = await _scan_duplicates_for_school(school_id)
        summary["after_remaining_duplicate_groups"] = verify["duplicate_groups"]
    return summary


@router.post("/grades/_maintenance/duplicates/create-index")
async def create_unique_grades_index(current_user=Depends(get_current_user)):
    """Creates the unique index that prevents future duplicates.

    Refuses to create the index if duplicate groups still exist (Mongo would
    fail anyway). Run /consolidate first, then this.
    """
    user = await resolve_user_from_token(current_user)
    if not user or user.get("role") not in {"owner", "director"}:
        raise HTTPException(403, "Sólo el owner/director puede ejecutar esta operación")
    school_id = user.get("school_id")
    if not school_id:
        raise HTTPException(400, "Tu usuario no tiene colegio asignado")

    # Safety: verify no duplicates remain (en este colegio o globalmente)
    global_check = await db.student_grades.aggregate([
        {"$group": {"_id": {k: f"${k}" for k in _DUP_GROUP_KEY}, "count": {"$sum": 1}}},
        {"$match": {"count": {"$gt": 1}}},
        {"$limit": 1},
    ]).to_list(1)
    if global_check:
        raise HTTPException(
            409,
            "Aún quedan duplicados en student_grades (posiblemente de otro colegio). "
            "El índice único no puede crearse hasta que se consoliden todos."
        )

    index_name = "uniq_grades_school_student_subject_section_period"
    try:
        existing = await db.student_grades.index_information()
        if index_name in existing:
            return {"created": False, "reason": "already_exists", "index_name": index_name}
        await db.student_grades.create_index(
            [
                ("school_id", 1),
                ("student_id", 1),
                ("subject_id", 1),
                ("section_id", 1),
                ("period_id", 1),
            ],
            unique=True,
            name=index_name,
            background=True,
        )
        logger.info(f"[GRADES MAINT] unique index created: {index_name} (requested by school={school_id})")
        return {"created": True, "index_name": index_name}
    except Exception as e:
        logger.exception(f"[GRADES MAINT] create-index falló: {e}")
        raise HTTPException(500, f"No se pudo crear el índice: {str(e)[:200]}")



@router.get("/grades/raw-dump/{subject_id}/{section_id}")
async def dump_raw_student_grades(
    subject_id: str,
    section_id: str,
    period_id: Optional[str] = Query(None),
    current_user = Depends(get_current_user),
):
    """Diagnostic endpoint: dumps the raw student_grades rows in MongoDB for
    a given subject + section (optionally filtered by period). Lets admins
    verify whether grades actually exist in the database vs. just not being
    rendered by the frontend (template column-id mismatch, period selector
    bug, etc.).
    """
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user) and not has_role(user, ["teacher", "coordinator", "auxiliar"]):
        raise HTTPException(status_code=403, detail="Solo administradores y staff pueden ejecutar este diagnóstico")

    school_id = user["school_id"]
    q = {"school_id": school_id, "subject_id": subject_id, "section_id": section_id}
    if period_id:
        q["period_id"] = period_id

    docs = await db.student_grades.find(q, {"_id": 0}).to_list(2000)

    student_ids = list({d.get("student_id") for d in docs if d.get("student_id")})
    students = await db.users.find(
        {"id": {"$in": student_ids}},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1},
    ).to_list(2000) if student_ids else []
    name_by_id = {s["id"]: f"{s.get('name','')} {s.get('last_name','')}".strip() for s in students}

    periods = await db.academic_periods.find(
        {"school_id": school_id}, {"_id": 0, "id": 1, "name": 1, "is_active": 1}
    ).to_list(50)
    name_by_period = {p["id"]: p.get("name") for p in periods}

    per_period_counts: Dict[str, dict] = {}
    dynamic_keys_seen: set = set()
    rows = []
    for d in docs:
        pid = d.get("period_id")
        per_period_counts.setdefault(pid, {
            "period_name": name_by_period.get(pid),
            "rows": 0, "with_dynamic_data": 0, "with_static_data": 0, "with_final_grade": 0,
        })
        per_period_counts[pid]["rows"] += 1

        gd = d.get("grades_dynamic") or {}
        gd = {k: v for k, v in gd.items() if v is not None}
        if gd:
            per_period_counts[pid]["with_dynamic_data"] += 1
            dynamic_keys_seen.update(gd.keys())

        static_values = {k: v for k, v in d.items() if k.startswith(("act_", "rf_", "comp_", "part_", "exam_")) and v is not None}
        if static_values:
            per_period_counts[pid]["with_static_data"] += 1
        if d.get("final_grade") is not None or d.get("final_grade_manual") is not None:
            per_period_counts[pid]["with_final_grade"] += 1

        rows.append({
            "student_id": d.get("student_id"),
            "student_name": name_by_id.get(d.get("student_id"), "?"),
            "period_id": pid,
            "period_name": name_by_period.get(pid),
            "final_grade": d.get("final_grade"),
            "final_grade_manual": d.get("final_grade_manual"),
            "grades_dynamic_count": len(gd),
            "grades_dynamic": gd,
            "static_values": static_values,
            "updated_at": d.get("updated_at"),
        })

    rows.sort(key=lambda r: ((r.get("period_name") or ""), (r.get("student_name") or "")))

    from services.register_sync import get_active_template_for_school
    template = await get_active_template_for_school(db, school_id)
    template_info = None
    if template:
        cols = []
        for cri in template.get("criterios") or []:
            for sub in cri.get("subcolumnas") or []:
                tipo = (sub.get("tipo") or "input").lower()
                if tipo not in ("promedio_auto", "promedio", "promedio_manual", "auto"):
                    cols.append({
                        "id": sub.get("id"),
                        "field_key": sub.get("field_key"),
                        "label": sub.get("label"),
                        "criterio": cri.get("nombre"),
                    })
        template_info = {
            "id": template.get("id"),
            "nombre": template.get("nombre"),
            "es_sistema": template.get("es_sistema"),
            "input_columns": cols,
        }

    # Highlight keys that are stored on disk but no longer match any column id
    # in the currently active template (= orphan keys after template edits).
    active_ids = set()
    if template_info:
        for c in template_info["input_columns"]:
            if c.get("id"):
                active_ids.add(c["id"])
            if c.get("field_key"):
                active_ids.add(c["field_key"])
    orphan_keys = sorted(k for k in dynamic_keys_seen if k not in active_ids) if active_ids else []

    return {
        "school_id": school_id,
        "subject_id": subject_id,
        "section_id": section_id,
        "period_filter": period_id,
        "total_rows": len(docs),
        "per_period_counts": per_period_counts,
        "dynamic_keys_seen": sorted(dynamic_keys_seen),
        "orphan_dynamic_keys": orphan_keys,
        "template": template_info,
        "rows": rows,
    }
