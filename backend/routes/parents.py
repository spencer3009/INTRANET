"""
Parent bulk import endpoints.
Handles template generation, file import, credentials download, and pending management.
"""
from fastapi import APIRouter, HTTPException, Depends, Form, UploadFile, File, Request
from datetime import datetime, timezone
import uuid
import re
import io
import csv
import string
import random
import unicodedata
import logging

from .core import (
    db, get_current_user, resolve_user_from_token, is_admin_user,
    hash_password, now_iso, generate_id,
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from starlette.responses import StreamingResponse

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api")

PARENT_HEADERS = ["Nombre", "Apellido", "DNI", "Correo", "Celular", "Cumpleanos", "Genero", "Direccion", "Usuario", "Contrasena", "Observaciones"]
REQUIRED_HEADERS = {"name", "last_name", "dni"}

COL_MAP = {
    "nombre": "name", "name": "name",
    "apellido": "last_name", "apellidos": "last_name", "last_name": "last_name",
    "dni": "dni", "documento": "dni",
    "correo": "email", "email": "email",
    "celular": "phone", "telefono": "phone", "phone": "phone",
    "cumpleanos": "birthday", "cumpleaños": "birthday", "birthday": "birthday", "fecha_nacimiento": "birthday",
    "genero": "gender", "género": "gender", "gender": "gender", "sexo": "gender",
    "direccion": "address", "dirección": "address", "address": "address",
    "usuario": "username", "username": "username",
    "contrasena": "password", "contraseña": "password", "password": "password",
    "observaciones": "observations", "notas": "observations", "notes": "observations",
}


def normalize_key(k):
    k = k.lower().strip()
    k = k.replace("*", "").strip()
    k = unicodedata.normalize("NFD", k)
    k = "".join(c for c in k if unicodedata.category(c) != "Mn")
    return COL_MAP.get(k, k)


def sanitize_field(value):
    if value is None:
        return ""
    return str(value).strip()


def sanitize_dni(value):
    v = sanitize_field(value)
    return re.sub(r"[^0-9]", "", v)


def sanitize_phone(value):
    v = sanitize_field(value)
    return re.sub(r"[^0-9]", "", v)


def sanitize_email(value):
    return sanitize_field(value).lower()


def generate_password(length=10):
    chars = string.ascii_letters + string.digits
    while True:
        pwd = "".join(random.choices(chars, k=length))
        has_upper = any(c.isupper() for c in pwd)
        has_lower = any(c.islower() for c in pwd)
        has_digit = any(c.isdigit() for c in pwd)
        if has_upper and has_lower and has_digit:
            return pwd


async def generate_username(name, last_name, dni, school_id):
    base = (name + last_name).lower()
    base = unicodedata.normalize("NFD", base)
    base = "".join(c for c in base if unicodedata.category(c) != "Mn")
    base = re.sub(r"[^a-z]", "", base)
    if not base:
        base = "padre"

    username = base
    if await db.users.find_one({"username": username, "school_id": school_id}):
        username = base + dni[-4:] if len(dni) >= 4 else base + dni
    if await db.users.find_one({"username": username, "school_id": school_id}):
        username = name[0].lower() + last_name.lower() + dni
        username = unicodedata.normalize("NFD", username)
        username = re.sub(r"[^a-z0-9]", "", "".join(c for c in username if unicodedata.category(c) != "Mn"))
    suffix = 1
    orig = username
    while await db.users.find_one({"username": username, "school_id": school_id}):
        username = f"{orig}{suffix}"
        suffix += 1
    return username


def parse_birthday(raw):
    if not raw:
        return None
    s = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    return None


def normalize_gender(raw):
    g = raw.lower().strip() if raw else ""
    if g in ("masculino", "male", "m", "hombre"):
        return "Masculino"
    if g in ("femenino", "female", "f", "mujer"):
        return "Femenino"
    if g in ("otro", "other", "o"):
        return "Otro"
    return None


def read_rows_from_file(content, ext):
    rows = []
    if ext == "csv":
        text = content.decode("utf-8-sig")
        for delimiter in [",", ";", "\t"]:
            try:
                reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
                test_rows = list(reader)
                if test_rows and len(test_rows[0]) > 1:
                    rows = [{k.strip(): (v.strip() if v else "") for k, v in r.items()} for r in test_rows]
                    break
            except Exception:
                continue
    elif ext == "xlsx":
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(all_rows):
            if row and any(str(c or "").strip().lower().replace("*", "").strip() in ("nombre", "name") for c in row):
                header_idx = i
                break
        if header_idx is None:
            wb.close()
            raise ValueError("ERR_WRONG_HEADERS")
        headers_raw = [str(c or "").strip() for c in all_rows[header_idx]]
        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            d = {}
            for j, h in enumerate(headers_raw):
                d[h] = str(row[j]).strip() if j < len(row) and row[j] is not None else ""
            rows.append(d)
        wb.close()
    elif ext == "xls":
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        header_idx = None
        for i in range(min(15, sheet.nrows)):
            vals = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
            if any(v.lower().replace("*", "").strip() in ("nombre", "name") for v in vals):
                header_idx = i
                break
        if header_idx is None:
            raise ValueError("ERR_WRONG_HEADERS")
        headers_raw = [str(sheet.cell_value(header_idx, j)).strip() for j in range(sheet.ncols)]
        for i in range(header_idx + 1, sheet.nrows):
            vals = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
            if all(v == "" for v in vals):
                continue
            d = {}
            for j, h in enumerate(headers_raw):
                d[h] = vals[j] if j < len(vals) else ""
            rows.append(d)
    return rows


# ═══════════════════════════════════════════════════════════════
#  GET /api/parents/template — Generate Excel template
# ═══════════════════════════════════════════════════════════════
@router.get("/parents/template")
async def generate_parent_template(current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden generar plantillas")

    school_id = user["school_id"]
    school = await db.settings.find_one({"school_id": school_id}, {"_id": 0, "system_name": 1})
    school_name = (school.get("system_name", "") if school else "").replace(" ", "_") or "colegio"

    wb = Workbook()

    # ── Sheet 1: Padres ──
    ws = wb.active
    ws.title = "Padres"

    blue_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    white_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells("A1:K1")
    ws["A1"] = "Plantilla de Importacion Masiva de Padres/Apoderados"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1565C0")

    # Instruction row
    ws.merge_cells("A2:K2")
    ws["A2"] = "Complete los datos en las filas debajo de los encabezados. Las columnas marcadas con (*) son obligatorias."
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")

    # Auto-gen note
    ws.merge_cells("A3:K3")
    ws["A3"] = "Nota: El usuario y contrasena se generaran automaticamente si no se proporcionan. La vinculacion con estudiantes se realiza desde el modulo de Estudiantes."
    ws["A3"].font = Font(name="Arial", italic=True, size=9, color="1565C0")

    # Headers row 5
    headers_display = ["Nombre *", "Apellido *", "DNI *", "Correo", "Celular", "Cumpleanos", "Genero", "Direccion", "Usuario", "Contrasena", "Observaciones"]
    for col, hdr in enumerate(headers_display, 1):
        cell = ws.cell(row=5, column=col, value=hdr)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    col_widths = [20, 20, 15, 28, 15, 15, 14, 35, 20, 18, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=5, column=i).column_letter].width = w

    ws.freeze_panes = "A6"

    # Example row
    example = ["Juan", "Perez", "45678912", "juan@correo.com", "987654321", "15/03/1985", "Masculino", "Av. Principal 123", "", "", "Apoderado legal"]
    example_font = Font(name="Arial", italic=True, size=10, color="999999")
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=6, column=col, value=val)
        cell.font = example_font
        cell.border = thin_border

    # Empty rows
    for row in range(7, 510):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            if row < 12:
                cell.border = thin_border

    # Lock header rows
    for row in range(1, 6):
        for col in range(1, 12):
            ws.cell(row=row, column=col).protection = Protection(locked=True)
    for row in range(6, 510):
        for col in range(1, 12):
            ws.cell(row=row, column=col).protection = Protection(locked=False)

    from openpyxl.worksheet.protection import SheetProtection
    ws.protection = SheetProtection(
        sheet=True, objects=True, scenarios=True,
        formatCells=False, formatColumns=False, formatRows=False,
        insertRows=True, deleteRows=True,
        selectLockedCells=True, selectUnlockedCells=False
    )

    # ── Sheet 2: Instrucciones ──
    ins = wb.create_sheet("Instrucciones")
    ins_title_font = Font(name="Arial", bold=True, size=14, color="1565C0")
    ins_header_font = Font(name="Arial", bold=True, size=11, color="1565C0")
    ins_text_font = Font(name="Arial", size=10, color="333333")
    ins_note_font = Font(name="Arial", italic=True, size=10, color="E65100")

    ins.column_dimensions["A"].width = 20
    ins.column_dimensions["B"].width = 12
    ins.column_dimensions["C"].width = 50

    ins.merge_cells("A1:C1")
    ins["A1"] = "Instrucciones para la Plantilla de Padres/Apoderados"
    ins["A1"].font = ins_title_font

    ins["A3"] = "Columna"
    ins["B3"] = "Obligatoria"
    ins["C3"] = "Descripcion / Formato"
    for c in ["A3", "B3", "C3"]:
        ins[c].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ins[c].fill = blue_fill

    instructions = [
        ("Nombre", "SI", "Nombre del padre/apoderado. Minimo 2 caracteres, maximo 100."),
        ("Apellido", "SI", "Apellido del padre/apoderado. Minimo 2 caracteres, maximo 100."),
        ("DNI", "SI", "Documento de identidad. Exactamente 8 digitos numericos. Sin guiones ni puntos."),
        ("Correo", "NO", "Correo electronico. Formato: ejemplo@dominio.com"),
        ("Celular", "NO", "Numero de celular. 9 digitos sin codigo de pais."),
        ("Cumpleanos", "NO", "Fecha de nacimiento. Formato: dd/mm/aaaa. Ejemplo: 15/03/1985"),
        ("Genero", "NO", "Opciones: Masculino, Femenino, Otro."),
        ("Direccion", "NO", "Direccion completa. Maximo 200 caracteres."),
        ("Usuario", "NO", "Nombre de usuario para acceder al sistema. Se genera automaticamente si se deja vacio."),
        ("Contrasena", "NO", "Contrasena para acceder al sistema. Se genera automaticamente si se deja vacia. Minimo 6 caracteres."),
        ("Observaciones", "NO", "Notas adicionales. Maximo 500 caracteres."),
    ]
    for i, (col_name, req, desc) in enumerate(instructions, 4):
        ins.cell(row=i, column=1, value=col_name).font = ins_text_font
        cell_req = ins.cell(row=i, column=2, value=req)
        cell_req.font = Font(name="Arial", bold=True, size=10, color="C62828" if req == "SI" else "2E7D32")
        ins.cell(row=i, column=3, value=desc).font = ins_text_font

    note_row = 4 + len(instructions) + 1
    ins.merge_cells(f"A{note_row}:C{note_row}")
    ins[f"A{note_row}"] = "IMPORTANTE: El usuario y contrasena se generaran automaticamente si no se proporcionan."
    ins[f"A{note_row}"].font = ins_note_font

    note_row2 = note_row + 1
    ins.merge_cells(f"A{note_row2}:C{note_row2}")
    ins[f"A{note_row2}"] = "IMPORTANTE: La vinculacion con estudiantes se realiza desde el modulo de Estudiantes, NO desde esta plantilla."
    ins[f"A{note_row2}"].font = ins_note_font

    note_row3 = note_row2 + 1
    ins.merge_cells(f"A{note_row3}:C{note_row3}")
    ins[f"A{note_row3}"] = "Si un DNI ya existe en el sistema, los datos opcionales se actualizaran automaticamente (no se crea un duplicado)."
    ins[f"A{note_row3}"].font = Font(name="Arial", italic=True, size=10, color="1565C0")

    # ── Hidden metadata ──
    meta = wb.create_sheet("edunet_metadata")
    meta.sheet_state = "hidden"
    meta.cell(row=1, column=1, value="school_id")
    meta.cell(row=1, column=2, value=school_id)
    meta.cell(row=2, column=1, value="type")
    meta.cell(row=2, column=2, value="parents")
    meta.cell(row=3, column=1, value="fecha_generacion")
    meta.cell(row=3, column=2, value=datetime.now(timezone.utc).isoformat())

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = re.sub(r"[^a-zA-Z0-9_]", "", school_name.lower().replace(" ", "_")) or "colegio"
    filename = f"plantilla_padres_{safe_name}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ═══════════════════════════════════════════════════════════════
#  POST /api/parents/import — Bulk import parents
# ═══════════════════════════════════════════════════════════════
@router.post("/parents/import")
async def import_parents(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user)
):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden importar padres")

    school_id = user["school_id"]
    now = datetime.now(timezone.utc).isoformat()
    batch_id = f"BATCH-PAR-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"

    # Validate file extension
    ext = file.filename.lower().rsplit(".", 1)[-1] if file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="ERR_FILE_FORMAT: Formato no soportado. Use .xlsx, .xls o .csv")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo excede el limite de 5MB")

    # Parse rows
    try:
        rows = read_rows_from_file(content, ext)
    except ValueError as e:
        if "ERR_WRONG_HEADERS" in str(e):
            raise HTTPException(status_code=400, detail="ERR_WRONG_HEADERS: Los encabezados del archivo no coinciden con la plantilla. Descargue la plantilla actualizada.")
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="ERR_EMPTY_FILE: El archivo no contiene datos")

    # Validate headers
    first_row_keys = {normalize_key(k) for k in rows[0].keys() if k.strip()}
    if not REQUIRED_HEADERS.issubset(first_row_keys):
        missing = REQUIRED_HEADERS - first_row_keys
        raise HTTPException(
            status_code=400,
            detail=f"ERR_WRONG_HEADERS: Faltan columnas obligatorias: {', '.join(missing)}. Descargue la plantilla actualizada."
        )

    created_count = 0
    updated_count = 0
    error_count = 0
    warning_count = 0
    credentials = []
    pending_ids = []
    seen_dnis = {}
    start_time = datetime.now(timezone.utc)

    for idx, raw_row in enumerate(rows):
        row = {normalize_key(k): v for k, v in raw_row.items() if k.strip()}
        row_num = idx + 1
        errors = []

        # Sanitize
        name = sanitize_field(row.get("name", ""))
        last_name = sanitize_field(row.get("last_name", ""))
        dni = sanitize_dni(row.get("dni", ""))
        email = sanitize_email(row.get("email", ""))
        phone = sanitize_phone(row.get("phone", ""))
        address = sanitize_field(row.get("address", ""))
        birthday_raw = sanitize_field(row.get("birthday", ""))
        gender_raw = sanitize_field(row.get("gender", ""))
        username_raw = sanitize_field(row.get("username", ""))
        password_raw = sanitize_field(row.get("password", ""))
        observations = sanitize_field(row.get("observations", ""))

        # Skip example row
        if name.lower() == "juan" and last_name.lower() == "perez" and dni == "45678912":
            continue

        # ── Validate required fields ──
        if not name or len(name) < 2:
            errors.append("ERR_EMPTY_FIELDS: Nombre vacio o muy corto")
        elif len(name) > 100:
            errors.append("ERR_EMPTY_FIELDS: Nombre excede 100 caracteres")

        if not last_name or len(last_name) < 2:
            errors.append("ERR_EMPTY_FIELDS: Apellido vacio o muy corto")
        elif len(last_name) > 100:
            errors.append("ERR_EMPTY_FIELDS: Apellido excede 100 caracteres")

        if not dni:
            errors.append("ERR_EMPTY_FIELDS: DNI vacio")
        elif not re.match(r"^\d{8}$", dni):
            errors.append("ERR_INVALID_DNI: DNI debe tener 8 digitos numericos")

        # ── Validate optional fields ──
        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            errors.append("ERR_INVALID_EMAIL: Formato de correo invalido")
        if phone and not re.match(r"^\d{9}$", phone):
            errors.append("ERR_INVALID_PHONE: Celular debe tener 9 digitos")
        if len(address) > 200:
            errors.append("Direccion excede 200 caracteres")
        if len(observations) > 500:
            errors.append("Observaciones excede 500 caracteres")
        if password_raw and len(password_raw) < 6:
            warning_count += 1

        # Parse birthday and gender
        birthday = parse_birthday(birthday_raw)
        gender = normalize_gender(gender_raw)
        if gender_raw and not gender:
            errors.append("Genero no valido (use Masculino, Femenino u Otro)")

        # Check duplicate in same file
        if dni and dni in seen_dnis:
            pass  # second occurrence = will update the first-created
        if dni:
            seen_dnis[dni] = row_num

        # ── If errors, send to Pending ──
        if errors:
            pending_doc = {
                "id": str(uuid.uuid4()),
                "school_id": school_id,
                "batch_id": batch_id,
                "type": "parent",
                "row_number": row_num,
                "name": name,
                "last_name": last_name,
                "dni": dni,
                "email": email or None,
                "phone": phone or None,
                "address": address or None,
                "birthday": birthday,
                "gender": gender,
                "username": username_raw or None,
                "observations": observations or None,
                "errors": errors,
                "status": "pending",
                "created_at": now,
            }
            await db.import_pending.insert_one(pending_doc)
            pending_doc.pop("_id", None)
            pending_ids.append(pending_doc["id"])
            error_count += 1
            continue

        # ── Check if parent already exists (auto-merge) ──
        existing = await db.users.find_one(
            {"dni": dni, "school_id": school_id, "role": "parent"},
            {"_id": 0}
        )

        if existing:
            update_fields = {}
            if phone and phone != (existing.get("phone") or ""):
                update_fields["phone"] = phone
            if email and email != (existing.get("email") or ""):
                update_fields["email"] = email
            if address and address != (existing.get("address") or ""):
                update_fields["address"] = address
            if birthday and birthday != (existing.get("birthday") or ""):
                update_fields["birthday"] = birthday
            if gender and gender != (existing.get("gender") or ""):
                update_fields["gender"] = gender
            if observations and observations != (existing.get("observations") or ""):
                update_fields["observations"] = observations

            if update_fields:
                update_fields["updated_at"] = now
                await db.users.update_one({"id": existing["id"]}, {"$set": update_fields})

            updated_count += 1
            continue

        # ── Generate username ──
        if username_raw:
            username = username_raw
            if await db.users.find_one({"username": username, "school_id": school_id}):
                pending_doc = {
                    "id": str(uuid.uuid4()),
                    "school_id": school_id,
                    "batch_id": batch_id,
                    "type": "parent",
                    "row_number": row_num,
                    "name": name, "last_name": last_name, "dni": dni,
                    "email": email or None, "phone": phone or None,
                    "address": address or None, "birthday": birthday, "gender": gender,
                    "username": username_raw, "observations": observations or None,
                    "errors": ["ERR_USERNAME_CONFLICT: Usuario ya tomado"],
                    "status": "pending",
                    "created_at": now,
                }
                await db.import_pending.insert_one(pending_doc)
                pending_doc.pop("_id", None)
                pending_ids.append(pending_doc["id"])
                error_count += 1
                continue
        else:
            username = await generate_username(name, last_name, dni, school_id)

        # ── Generate password ──
        plain_password = password_raw if password_raw else generate_password()

        # ── Create parent ──
        parent_doc = {
            "id": str(uuid.uuid4()),
            "name": name,
            "last_name": last_name,
            "dni": dni,
            "email": email or None,
            "phone": phone or None,
            "address": address or None,
            "birthday": birthday,
            "gender": gender,
            "username": username,
            "password": hash_password(plain_password),
            "password_display": plain_password,
            "role": "parent",
            "account_type": "Padre/Apoderado",
            "school_id": school_id,
            "status": "active",
            "observations": observations or None,
            "email_verified": True,
            "import_status": "imported",
            "import_batch_id": batch_id,
            "created_by": "import_bulk",
            "created_at": now,
            "updated_at": now,
        }

        await db.users.insert_one(parent_doc)
        parent_doc.pop("_id", None)
        created_count += 1

        credentials.append({
            "name": name,
            "last_name": last_name,
            "dni": dni,
            "username": username,
            "password": plain_password,
        })

    elapsed = (datetime.now(timezone.utc) - start_time).total_seconds()

    # Store credentials for later download
    if credentials:
        await db.import_credentials.insert_one({
            "batch_id": batch_id,
            "school_id": school_id,
            "type": "parent",
            "credentials": credentials,
            "created_at": now,
        })

    logger.info(f"Parent import batch={batch_id}: {created_count} created, {updated_count} updated, {error_count} errors by user {user['id']}")

    return {
        "success": True,
        "batch_id": batch_id,
        "summary": {
            "total_rows": len(rows),
            "created": created_count,
            "updated": updated_count,
            "errors": error_count,
            "warnings": warning_count,
            "processing_time_ms": round(elapsed * 1000),
        },
        "credentials_available": len(credentials) > 0,
        "pending_ids": pending_ids,
    }


# ═══════════════════════════════════════════════════════════════
#  GET /api/parents/import/{batch_id}/credentials — Download CSV
# ═══════════════════════════════════════════════════════════════
@router.get("/parents/import/{batch_id}/credentials")
async def download_parent_credentials(batch_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")

    doc = await db.import_credentials.find_one(
        {"batch_id": batch_id, "school_id": user["school_id"], "type": "parent"},
        {"_id": 0}
    )
    if not doc or not doc.get("credentials"):
        raise HTTPException(status_code=404, detail="No se encontraron credenciales para este lote")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Nombre", "Apellido", "DNI", "Usuario", "Contrasena"])
    for c in doc["credentials"]:
        writer.writerow([c["name"], c["last_name"], c["dni"], c["username"], c["password"]])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    buffer = io.BytesIO(csv_bytes)

    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="credenciales_padres_{batch_id}.csv"'}
    )


# ═══════════════════════════════════════════════════════════════
#  GET /api/parents/import/{batch_id}/summary
# ═══════════════════════════════════════════════════════════════
@router.get("/parents/import/{batch_id}/summary")
async def get_import_summary(batch_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")

    pending = await db.import_pending.find(
        {"batch_id": batch_id, "school_id": user["school_id"]},
        {"_id": 0}
    ).to_list(500)

    cred_doc = await db.import_credentials.find_one(
        {"batch_id": batch_id, "school_id": user["school_id"]},
        {"_id": 0, "credentials": 0}
    )

    return {
        "batch_id": batch_id,
        "pending": pending,
        "credentials_available": cred_doc is not None,
    }


# ═══════════════════════════════════════════════════════════════
#  PENDING MANAGEMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@router.get("/parents/pending")
async def get_pending_parents(current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")

    pending = await db.import_pending.find(
        {"school_id": user["school_id"], "type": "parent", "status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return pending


@router.put("/parents/pending/{pending_id}")
async def edit_pending_parent(pending_id: str, request: Request, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    body = await request.json()
    allowed = {"name", "last_name", "dni", "email", "phone", "address", "birthday", "gender", "username", "observations"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if not updates:
        raise HTTPException(status_code=400, detail="Sin campos para actualizar")

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.import_pending.update_one(
        {"id": pending_id, "school_id": user["school_id"], "type": "parent"},
        {"$set": updates}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Pendiente no encontrado")
    return {"message": "Pendiente actualizado"}


@router.post("/parents/pending/{pending_id}/activate")
async def activate_pending_parent(pending_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    school_id = user["school_id"]
    now = datetime.now(timezone.utc).isoformat()

    pending = await db.import_pending.find_one(
        {"id": pending_id, "school_id": school_id, "type": "parent", "status": "pending"},
        {"_id": 0}
    )
    if not pending:
        raise HTTPException(status_code=404, detail="Pendiente no encontrado")

    name = sanitize_field(pending.get("name", ""))
    last_name = sanitize_field(pending.get("last_name", ""))
    dni = sanitize_dni(pending.get("dni", ""))

    if not name or not last_name or not dni or not re.match(r"^\d{8}$", dni):
        raise HTTPException(status_code=400, detail="Datos incompletos. Edite los campos obligatorios primero.")

    existing = await db.users.find_one({"dni": dni, "school_id": school_id, "role": "parent"}, {"_id": 0, "id": 1})
    if existing:
        raise HTTPException(status_code=400, detail=f"DNI {dni} ya existe como padre en el sistema")

    username = pending.get("username") or await generate_username(name, last_name, dni, school_id)
    plain_password = generate_password()

    parent_doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "last_name": last_name,
        "dni": dni,
        "email": pending.get("email") or None,
        "phone": pending.get("phone") or None,
        "address": pending.get("address") or None,
        "birthday": pending.get("birthday"),
        "gender": pending.get("gender"),
        "username": username,
        "password": hash_password(plain_password),
        "password_display": plain_password,
        "role": "parent",
        "account_type": "Padre/Apoderado",
        "school_id": school_id,
        "status": "active",
        "observations": pending.get("observations") or None,
        "email_verified": True,
        "import_status": "imported",
        "import_batch_id": pending.get("batch_id"),
        "created_by": "import_bulk",
        "created_at": now,
        "updated_at": now,
    }

    await db.users.insert_one(parent_doc)
    parent_doc.pop("_id", None)

    await db.import_pending.update_one(
        {"id": pending_id},
        {"$set": {"status": "activated", "updated_at": now}}
    )

    return {
        "message": "Padre creado exitosamente",
        "username": username,
        "password": plain_password,
    }


@router.delete("/parents/pending/{pending_id}")
async def delete_pending_parent(pending_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    result = await db.import_pending.delete_one(
        {"id": pending_id, "school_id": user["school_id"], "type": "parent"}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pendiente no encontrado")
    return {"message": "Pendiente eliminado"}
