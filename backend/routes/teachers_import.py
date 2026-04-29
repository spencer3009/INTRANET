"""
Teacher bulk import endpoints.
Replicates the parent bulk-import pattern but tailored to the `teacher` role:
- username = password = DNI (auto-generated, not on the template)
- generates a QR via services.qr_service.generate_user_qr (same as individual teacher creation)
- supports template download, bulk import, credentials CSV, and pending management.
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Request
from datetime import datetime, timezone
import uuid
import re
import io
import csv
import unicodedata
import logging

from .core import (
    db, get_current_user, resolve_user_from_token, is_admin_user,
    hash_password,
)

from services.qr_service import generate_user_qr

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.protection import SheetProtection
from starlette.responses import StreamingResponse

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api")

REQUIRED_HEADERS = {"name", "dni"}

COL_MAP = {
    "nombre": "name", "name": "name",
    "apellido": "last_name", "apellidos": "last_name", "last_name": "last_name",
    "dni": "dni", "documento": "dni",
    "correo": "email", "email": "email",
    "celular": "phone", "telefono": "phone", "phone": "phone",
    "cumpleanos": "birthday", "cumpleaños": "birthday", "birthday": "birthday", "fecha_nacimiento": "birthday",
    "genero": "gender", "género": "gender", "gender": "gender", "sexo": "gender",
    "direccion": "address", "dirección": "address", "address": "address",
    "observaciones": "observations", "notas": "observations", "notes": "observations",
}


def normalize_key(k):
    k = k.lower().strip()
    k = k.replace("*", "").strip()
    k = unicodedata.normalize("NFD", k)
    k = "".join(c for c in k if unicodedata.category(c) != "Mn")
    return COL_MAP.get(k, k)


def sanitize_field(value):
    if value is None:
        return ""
    return str(value).strip()


def sanitize_dni(value):
    v = sanitize_field(value)
    return re.sub(r"[^0-9]", "", v)


def sanitize_phone(value):
    v = sanitize_field(value)
    return re.sub(r"[^0-9]", "", v)


def sanitize_email(value):
    return sanitize_field(value).lower()


def parse_birthday(raw):
    if not raw:
        return None
    s = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    return None


def normalize_gender(raw):
    """Teacher template uses only M / F (per spec). Still accept wider set."""
    g = raw.lower().strip() if raw else ""
    if g in ("m", "masculino", "male", "hombre"):
        return "Masculino"
    if g in ("f", "femenino", "female", "mujer"):
        return "Femenino"
    if g in ("o", "otro", "other"):
        return "Otro"
    return None


def read_rows_from_file(content, ext):
    rows = []
    if ext == "csv":
        text = content.decode("utf-8-sig")
        for delimiter in [",", ";", "\t"]:
            try:
                reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
                test_rows = list(reader)
                if test_rows and len(test_rows[0]) > 1:
                    rows = [{k.strip(): (v.strip() if v else "") for k, v in r.items()} for r in test_rows]
                    break
            except Exception:
                continue
    elif ext == "xlsx":
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(all_rows):
            if row and any(str(c or "").strip().lower().replace("*", "").strip() in ("nombre", "name") for c in row):
                header_idx = i
                break
        if header_idx is None:
            wb.close()
            raise ValueError("ERR_WRONG_HEADERS")
        headers_raw = [str(c or "").strip() for c in all_rows[header_idx]]
        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            d = {}
            for j, h in enumerate(headers_raw):
                d[h] = str(row[j]).strip() if j < len(row) and row[j] is not None else ""
            rows.append(d)
        wb.close()
    elif ext == "xls":
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        header_idx = None
        for i in range(min(15, sheet.nrows)):
            vals = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
            if any(v.lower().replace("*", "").strip() in ("nombre", "name") for v in vals):
                header_idx = i
                break
        if header_idx is None:
            raise ValueError("ERR_WRONG_HEADERS")
        headers_raw = [str(sheet.cell_value(header_idx, j)).strip() for j in range(sheet.ncols)]
        for i in range(header_idx + 1, sheet.nrows):
            vals = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
            if all(v == "" for v in vals):
                continue
            d = {}
            for j, h in enumerate(headers_raw):
                d[h] = vals[j] if j < len(vals) else ""
            rows.append(d)
    return rows


# ═══════════════════════════════════════════════════════════════
#  GET /api/teachers/template — Generate Excel template
# ═══════════════════════════════════════════════════════════════
@router.get("/teachers/template")
async def generate_teacher_template(current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden generar plantillas")

    school_id = user["school_id"]
    school = await db.settings.find_one({"school_id": school_id}, {"_id": 0, "system_name": 1})
    school_name = (school.get("system_name", "") if school else "").replace(" ", "_") or "colegio"

    wb = Workbook()

    # ── Sheet 1: Profesores ──
    ws = wb.active
    ws.title = "Profesores"

    green_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    white_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = "Plantilla de Importacion Masiva de Profesores"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="047857")

    # Instruction row
    ws.merge_cells("A2:I2")
    ws["A2"] = "Complete los datos en las filas debajo de los encabezados. Las columnas marcadas con (*) son obligatorias."
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")

    # Auto-gen note
    ws.merge_cells("A3:I3")
    ws["A3"] = "Nota: El usuario y contrasena se generaran automaticamente a partir del DNI. Se generara un QR para cada profesor."
    ws["A3"].font = Font(name="Arial", italic=True, size=9, color="047857")

    headers_display = [
        "Nombre *", "Apellido", "DNI *", "Correo", "Celular",
        "Cumpleanos", "Genero", "Direccion", "Observaciones",
    ]
    for col, hdr in enumerate(headers_display, 1):
        cell = ws.cell(row=5, column=col, value=hdr)
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    col_widths = [22, 22, 15, 30, 15, 15, 10, 35, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=5, column=i).column_letter].width = w

    ws.freeze_panes = "A6"

    # Example row
    example = [
        "Maria", "Lopez Torres", "45678912", "maria.lopez@correo.com", "987654321",
        "15/03/1985", "F", "Av. Principal 123", "Profesora de matematica",
    ]
    example_font = Font(name="Arial", italic=True, size=10, color="999999")
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=6, column=col, value=val)
        cell.font = example_font
        cell.border = thin_border

    # Empty rows (borders for the first few to guide the user)
    for row in range(7, 510):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if row < 12:
                cell.border = thin_border

    # Lock header rows
    for row in range(1, 6):
        for col in range(1, 10):
            ws.cell(row=row, column=col).protection = Protection(locked=True)
    for row in range(6, 510):
        for col in range(1, 10):
            ws.cell(row=row, column=col).protection = Protection(locked=False)

    ws.protection = SheetProtection(
        sheet=True, objects=True, scenarios=True,
        formatCells=False, formatColumns=False, formatRows=False,
        insertRows=True, deleteRows=True,
        selectLockedCells=True, selectUnlockedCells=False
    )

    # ── Sheet 2: Instrucciones ──
    ins = wb.create_sheet("Instrucciones")
    ins_title_font = Font(name="Arial", bold=True, size=14, color="047857")
    ins_text_font = Font(name="Arial", size=10, color="333333")
    ins_note_font = Font(name="Arial", italic=True, size=10, color="E65100")

    ins.column_dimensions["A"].width = 20
    ins.column_dimensions["B"].width = 12
    ins.column_dimensions["C"].width = 55

    ins.merge_cells("A1:C1")
    ins["A1"] = "Instrucciones para la Plantilla de Profesores"
    ins["A1"].font = ins_title_font

    ins["A3"] = "Columna"
    ins["B3"] = "Obligatoria"
    ins["C3"] = "Descripcion / Formato"
    for c in ["A3", "B3", "C3"]:
        ins[c].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ins[c].fill = green_fill

    instructions = [
        ("Nombre", "SI", "Nombre del profesor. Minimo 2 caracteres, maximo 100."),
        ("Apellido", "NO", "Apellido del profesor. Maximo 100 caracteres."),
        ("DNI", "SI", "Documento de identidad. Exactamente 8 digitos numericos. Sin guiones ni puntos."),
        ("Correo", "NO", "Correo electronico. Formato: ejemplo@dominio.com"),
        ("Celular", "NO", "Numero de celular. 9 digitos sin codigo de pais."),
        ("Cumpleanos", "NO", "Fecha de nacimiento. Formato: dd/mm/aaaa. Ejemplo: 15/03/1985"),
        ("Genero", "NO", "Opciones: M (Masculino) / F (Femenino)."),
        ("Direccion", "NO", "Direccion completa. Maximo 200 caracteres."),
        ("Observaciones", "NO", "Notas adicionales. Maximo 500 caracteres."),
    ]
    for i, (col_name, req, desc) in enumerate(instructions, 4):
        ins.cell(row=i, column=1, value=col_name).font = ins_text_font
        cell_req = ins.cell(row=i, column=2, value=req)
        cell_req.font = Font(name="Arial", bold=True, size=10, color="C62828" if req == "SI" else "2E7D32")
        ins.cell(row=i, column=3, value=desc).font = ins_text_font

    note_row = 4 + len(instructions) + 1
    ins.merge_cells(f"A{note_row}:C{note_row}")
    ins[f"A{note_row}"] = "IMPORTANTE: El usuario y la contrasena se generan automaticamente a partir del DNI del profesor."
    ins[f"A{note_row}"].font = ins_note_font

    note_row2 = note_row + 1
    ins.merge_cells(f"A{note_row2}:C{note_row2}")
    ins[f"A{note_row2}"] = "IMPORTANTE: La asignacion de cursos y aulas se realiza desde el modulo de Horarios, NO desde esta plantilla."
    ins[f"A{note_row2}"].font = ins_note_font

    note_row3 = note_row2 + 1
    ins.merge_cells(f"A{note_row3}:C{note_row3}")
    ins[f"A{note_row3}"] = "Si un DNI ya existe en el sistema como profesor, los datos opcionales se actualizaran automaticamente (no se crea un duplicado)."
    ins[f"A{note_row3}"].font = Font(name="Arial", italic=True, size=10, color="047857")

    # ── Hidden metadata ──
    meta = wb.create_sheet("edunet_metadata")
    meta.sheet_state = "hidden"
    meta.cell(row=1, column=1, value="school_id")
    meta.cell(row=1, column=2, value=school_id)
    meta.cell(row=2, column=1, value="type")
    meta.cell(row=2, column=2, value="teachers")
    meta.cell(row=3, column=1, value="fecha_generacion")
    meta.cell(row=3, column=2, value=datetime.now(timezone.utc).isoformat())

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = re.sub(r"[^a-zA-Z0-9_]", "", school_name.lower().replace(" ", "_")) or "colegio"
    filename = f"plantilla_profesores_{safe_name}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ═══════════════════════════════════════════════════════════════
#  POST /api/teachers/import — Bulk import teachers
# ═══════════════════════════════════════════════════════════════
@router.post("/teachers/import")
async def import_teachers(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user)
):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden importar profesores")

    school_id = user["school_id"]
    now = datetime.now(timezone.utc).isoformat()
    batch_id = f"BATCH-TEA-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"

    ext = file.filename.lower().rsplit(".", 1)[-1] if file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="ERR_FILE_FORMAT: Formato no soportado. Use .xlsx, .xls o .csv")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo excede el limite de 5MB")

    try:
        rows = read_rows_from_file(content, ext)
    except ValueError as e:
        if "ERR_WRONG_HEADERS" in str(e):
            raise HTTPException(status_code=400, detail="El archivo no tiene el formato correcto. Descarga la plantilla e inténtalo de nuevo.")
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="ERR_EMPTY_FILE: El archivo no contiene datos")

    first_row_keys = {normalize_key(k) for k in rows[0].keys() if k.strip()}
    if not REQUIRED_HEADERS.issubset(first_row_keys):
        raise HTTPException(
            status_code=400,
            detail="El archivo no tiene el formato correcto. Descarga la plantilla e inténtalo de nuevo."
        )

    created_count = 0
    updated_count = 0
    error_count = 0
    credentials = []
    pending_ids = []
    seen_dnis = {}
    start_time = datetime.now(timezone.utc)

    for idx, raw_row in enumerate(rows):
        row = {normalize_key(k): v for k, v in raw_row.items() if k.strip()}
        row_num = idx + 1
        errors = []

        # Sanitize
        name = sanitize_field(row.get("name", ""))
        last_name = sanitize_field(row.get("last_name", ""))
        dni = sanitize_dni(row.get("dni", ""))
        email = sanitize_email(row.get("email", ""))
        phone = sanitize_phone(row.get("phone", ""))
        address = sanitize_field(row.get("address", ""))
        birthday_raw = sanitize_field(row.get("birthday", ""))
        gender_raw = sanitize_field(row.get("gender", ""))
        observations = sanitize_field(row.get("observations", ""))

        # Skip example row
        if name.lower() == "maria" and last_name.lower() == "lopez torres" and dni == "45678912":
            continue

        # ── Validate required fields ──
        if not name or len(name) < 2:
            errors.append("ERR_EMPTY_FIELDS: Nombre vacio o muy corto")
        elif len(name) > 100:
            errors.append("ERR_EMPTY_FIELDS: Nombre excede 100 caracteres")

        if last_name and len(last_name) > 100:
            errors.append("Apellido excede 100 caracteres")

        if not dni:
            errors.append("ERR_EMPTY_FIELDS: DNI vacio")
        elif not re.match(r"^\d{8}$", dni):
            errors.append("ERR_INVALID_DNI: DNI debe tener 8 digitos numericos")

        # ── Validate optional fields ──
        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            errors.append("ERR_INVALID_EMAIL: Formato de correo invalido")
        if phone and not re.match(r"^\d{9}$", phone):
            errors.append("ERR_INVALID_PHONE: Celular debe tener 9 digitos")
        if len(address) > 200:
            errors.append("Direccion excede 200 caracteres")
        if len(observations) > 500:
            errors.append("Observaciones excede 500 caracteres")

        birthday = parse_birthday(birthday_raw)
        gender = normalize_gender(gender_raw)
        if gender_raw and not gender:
            errors.append("Genero no valido (use M o F)")

        if dni:
            seen_dnis[dni] = row_num

        # ── If errors, send to Pending ──
        if errors:
            pending_doc = {
                "id": str(uuid.uuid4()),
                "school_id": school_id,
                "batch_id": batch_id,
                "type": "teacher",
                "row_number": row_num,
                "name": name,
                "last_name": last_name,
                "dni": dni,
                "email": email or None,
                "phone": phone or None,
                "address": address or None,
                "birthday": birthday,
                "gender": gender,
                "observations": observations or None,
                "errors": errors,
                "status": "pending",
                "created_at": now,
            }
            await db.import_pending.insert_one(pending_doc)
            pending_doc.pop("_id", None)
            pending_ids.append(pending_doc["id"])
            error_count += 1
            continue

        # ── Check if teacher already exists in this school (auto-merge) ──
        existing = await db.users.find_one(
            {"dni": dni, "school_id": school_id, "role": "teacher"},
            {"_id": 0}
        )

        if existing:
            update_fields = {}
            if last_name and last_name != (existing.get("last_name") or ""):
                update_fields["last_name"] = last_name
            if phone and phone != (existing.get("phone") or ""):
                update_fields["phone"] = phone
            if email and email != (existing.get("email") or ""):
                update_fields["email"] = email
            if address and address != (existing.get("address") or ""):
                update_fields["address"] = address
            if birthday and birthday != (existing.get("birthday") or ""):
                update_fields["birthday"] = birthday
            if gender and gender != (existing.get("gender") or ""):
                update_fields["gender"] = gender
            if observations and observations != (existing.get("observations") or ""):
                update_fields["observations"] = observations

            # Backfill QR if missing
            if not existing.get("qr_token"):
                try:
                    qr_id, qr_token = await generate_user_qr(db)
                    update_fields["qr_id"] = qr_id
                    update_fields["qr_token"] = qr_token
                    update_fields["qr_version"] = 2
                except Exception as e:
                    logger.warning(f"[TEACHER-IMPORT] QR backfill failed for {existing.get('id')}: {e}")

            if update_fields:
                update_fields["updated_at"] = now
                await db.users.update_one({"id": existing["id"]}, {"$set": update_fields})

            updated_count += 1
            continue

        # ── Check username conflict (username = DNI) ──
        username = dni
        if await db.users.find_one({"username": username, "school_id": school_id}):
            pending_doc = {
                "id": str(uuid.uuid4()),
                "school_id": school_id,
                "batch_id": batch_id,
                "type": "teacher",
                "row_number": row_num,
                "name": name, "last_name": last_name, "dni": dni,
                "email": email or None, "phone": phone or None,
                "address": address or None, "birthday": birthday, "gender": gender,
                "observations": observations or None,
                "errors": [f"ERR_USERNAME_CONFLICT: El DNI {dni} ya está registrado como usuario en este colegio"],
                "status": "pending",
                "created_at": now,
            }
            await db.import_pending.insert_one(pending_doc)
            pending_doc.pop("_id", None)
            pending_ids.append(pending_doc["id"])
            error_count += 1
            continue

        # ── Password = DNI (same as spec) ──
        plain_password = dni

        # ── Generate short QR (same as individual creation) ──
        try:
            qr_id, qr_token = await generate_user_qr(db)
        except Exception as e:
            logger.warning(f"[TEACHER-IMPORT] QR gen failed row {row_num}: {e}")
            qr_id, qr_token = None, None

        teacher_doc = {
            "id": str(uuid.uuid4()),
            "name": name,
            "last_name": last_name or None,
            "dni": dni,
            "email": email or None,
            "phone": phone or None,
            "address": address or None,
            "birthday": birthday,
            "gender": gender,
            "username": username,
            "password": hash_password(plain_password),
            "password_display": plain_password,
            "role": "teacher",
            "account_type": "Profesor",
            "school_id": school_id,
            "status": "active",
            "observations": observations or None,
            "email_verified": True,
            "import_status": "imported",
            "import_batch_id": batch_id,
            "created_by": "import_bulk",
            "created_at": now,
            "updated_at": now,
        }
        if qr_id and qr_token:
            teacher_doc["qr_id"] = qr_id
            teacher_doc["qr_token"] = qr_token
            teacher_doc["qr_version"] = 2

        await db.users.insert_one(teacher_doc)
        teacher_doc.pop("_id", None)
        created_count += 1

        credentials.append({
            "name": name,
            "last_name": last_name,
            "dni": dni,
            "username": username,
            "password": plain_password,
        })

    elapsed = (datetime.now(timezone.utc) - start_time).total_seconds()

    if credentials:
        await db.import_credentials.insert_one({
            "batch_id": batch_id,
            "school_id": school_id,
            "type": "teacher",
            "credentials": credentials,
            "created_at": now,
        })

    logger.info(f"Teacher import batch={batch_id}: {created_count} created, {updated_count} updated, {error_count} errors by user {user['id']}")

    return {
        "success": True,
        "batch_id": batch_id,
        "summary": {
            "total_rows": len(rows),
            "created": created_count,
            "updated": updated_count,
            "errors": error_count,
            "warnings": 0,
            "processing_time_ms": round(elapsed * 1000),
        },
        "credentials_available": len(credentials) > 0,
        "pending_ids": pending_ids,
    }


# ═══════════════════════════════════════════════════════════════
#  GET /api/teachers/import/{batch_id}/credentials — Download CSV
# ═══════════════════════════════════════════════════════════════
@router.get("/teachers/import/{batch_id}/credentials")
async def download_teacher_credentials(batch_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")

    doc = await db.import_credentials.find_one(
        {"batch_id": batch_id, "school_id": user["school_id"], "type": "teacher"},
        {"_id": 0}
    )
    if not doc or not doc.get("credentials"):
        raise HTTPException(status_code=404, detail="No se encontraron credenciales para este lote")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Nombre", "Apellido", "DNI", "Usuario", "Contrasena"])
    for c in doc["credentials"]:
        writer.writerow([c["name"], c.get("last_name", ""), c["dni"], c["username"], c["password"]])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    buffer = io.BytesIO(csv_bytes)

    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="credenciales_profesores_{batch_id}.csv"'}
    )


# ═══════════════════════════════════════════════════════════════
#  PENDING MANAGEMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════════
@router.get("/teachers/pending")
async def get_pending_teachers(current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")

    pending = await db.import_pending.find(
        {"school_id": user["school_id"], "type": "teacher", "status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return pending


@router.put("/teachers/pending/{pending_id}")
async def edit_pending_teacher(pending_id: str, request: Request, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    school_id = user["school_id"]
    body = await request.json()
    allowed = {"name", "last_name", "dni", "email", "phone", "address", "birthday", "gender", "observations"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if not updates:
        raise HTTPException(status_code=400, detail="Sin campos para actualizar")

    # Sanitize
    if "dni" in updates:
        updates["dni"] = sanitize_dni(updates["dni"])
    if "email" in updates:
        updates["email"] = sanitize_email(updates["email"])
    if "phone" in updates:
        updates["phone"] = sanitize_phone(updates["phone"])

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()

    # Re-validate — clear errors if now valid
    new_errors = []
    name_val = updates.get("name", "")
    if "name" in updates and (not name_val or len(name_val) < 2):
        new_errors.append("ERR_EMPTY_FIELDS: Nombre vacio o muy corto")

    dni_val = updates.get("dni", "")
    if "dni" in updates:
        if not dni_val:
            new_errors.append("ERR_EMPTY_FIELDS: DNI vacio")
        elif not re.match(r"^\d{8}$", dni_val):
            new_errors.append("ERR_INVALID_DNI: DNI debe tener 8 digitos numericos")

    email_val = updates.get("email", "")
    if email_val and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email_val):
        new_errors.append("ERR_INVALID_EMAIL: Formato de correo invalido")

    phone_val = updates.get("phone", "")
    if phone_val and not re.match(r"^\d{9}$", phone_val):
        new_errors.append("ERR_INVALID_PHONE: Celular debe tener 9 digitos")

    updates["errors"] = new_errors

    result = await db.import_pending.update_one(
        {"id": pending_id, "school_id": school_id, "type": "teacher"},
        {"$set": updates}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pendiente no encontrado")
    return {"message": "Pendiente actualizado", "errors": new_errors}


@router.post("/teachers/pending/{pending_id}/activate")
async def activate_pending_teacher(pending_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    school_id = user["school_id"]
    now = datetime.now(timezone.utc).isoformat()

    pending = await db.import_pending.find_one(
        {"id": pending_id, "school_id": school_id, "type": "teacher", "status": "pending"},
        {"_id": 0}
    )
    if not pending:
        raise HTTPException(status_code=404, detail="Pendiente no encontrado")

    name = sanitize_field(pending.get("name", ""))
    last_name = sanitize_field(pending.get("last_name", ""))
    dni = sanitize_dni(pending.get("dni", ""))

    if not name or not dni:
        raise HTTPException(status_code=400, detail="Datos incompletos. Edita los campos obligatorios primero.")
    if not re.match(r"^\d{8}$", dni):
        raise HTTPException(status_code=400, detail="El DNI debe tener 8 digitos numericos")

    existing = await db.users.find_one({"dni": dni, "school_id": school_id, "role": "teacher"}, {"_id": 0, "id": 1})
    if existing:
        raise HTTPException(status_code=400, detail=f"DNI {dni} ya existe como profesor en el sistema")
    if await db.users.find_one({"username": dni, "school_id": school_id}, {"_id": 0, "id": 1}):
        raise HTTPException(status_code=400, detail=f"El DNI {dni} ya está usado como usuario en este colegio")

    username = dni
    plain_password = dni

    try:
        qr_id, qr_token = await generate_user_qr(db)
    except Exception as e:
        logger.warning(f"[TEACHER-ACTIVATE] QR gen failed for {pending_id}: {e}")
        qr_id, qr_token = None, None

    teacher_doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "last_name": last_name or None,
        "dni": dni,
        "email": pending.get("email") or None,
        "phone": pending.get("phone") or None,
        "address": pending.get("address") or None,
        "birthday": pending.get("birthday"),
        "gender": pending.get("gender"),
        "username": username,
        "password": hash_password(plain_password),
        "password_display": plain_password,
        "role": "teacher",
        "account_type": "Profesor",
        "school_id": school_id,
        "status": "active",
        "observations": pending.get("observations") or None,
        "email_verified": True,
        "import_status": "imported",
        "import_batch_id": pending.get("batch_id"),
        "created_by": "import_bulk",
        "created_at": now,
        "updated_at": now,
    }
    if qr_id and qr_token:
        teacher_doc["qr_id"] = qr_id
        teacher_doc["qr_token"] = qr_token
        teacher_doc["qr_version"] = 2

    await db.users.insert_one(teacher_doc)
    teacher_doc.pop("_id", None)

    await db.import_pending.update_one(
        {"id": pending_id},
        {"$set": {"status": "activated", "updated_at": now}}
    )

    return {
        "message": "Profesor creado exitosamente",
        "username": username,
        "password": plain_password,
    }


@router.delete("/teachers/pending/{pending_id}")
async def delete_pending_teacher(pending_id: str, current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    result = await db.import_pending.delete_one(
        {"id": pending_id, "school_id": user["school_id"], "type": "teacher"}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pendiente no encontrado")
    return {"message": "Pendiente eliminado"}
