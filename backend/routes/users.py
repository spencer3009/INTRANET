"""
User CRUD, student import
Extracted from server.py during modularization.
"""
from fastapi import APIRouter, HTTPException, Depends, Query, Body, Form, UploadFile, File, BackgroundTasks, Request
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Literal
from datetime import datetime, timezone, timedelta
from enum import Enum
import uuid
import re
import logging

from .core import (
    db, get_current_user, resolve_user_from_token, is_admin_user,
    require_role, require_admin, require_staff, require_section_access,
    is_demo_user, check_demo_user_block, require_not_demo, is_real_owner,
    is_system_user, check_system_user_block, is_protected_user,
    has_role, is_student, is_parent, is_staff,
    can_access_section, get_user_permissions,
    hash_password, verify_password, create_token,
    get_academic_filter, invalidate_student_cache, STUDENT_VISIBLE_FILTER,
    JWT_SECRET, JWT_ALGORITHM, now_iso, generate_id,
    ADMIN_ROLES, STAFF_ROLES, ROLE_HIERARCHY,
    ACADEMIC_STUDENT_FILTER, ACADEMIC_STUDENT_FILTER_WITH_PENDING,
    create_system_support_user,
)

from services.qr_service import generate_user_qr
import io
import csv
import unicodedata
import cloudinary
import cloudinary.uploader
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")


@router.get("/students/{student_id}/constancia-matricula")
async def download_constancia_matricula(student_id: str, current_user=Depends(get_current_user)):
    """Genera y descarga la Constancia de Matrícula (PDF) de un alumno. Solo admin/director/owner."""
    user = await resolve_user_from_token(current_user)
    if not user or user.get("role") not in ("owner", "admin", "director"):
        raise HTTPException(status_code=403, detail="No tienes permisos para generar constancias")
    school_id = user.get("school_id")

    student = await db.users.find_one(
        {"id": student_id, "school_id": school_id, "role": "student"}, {"_id": 0}
    )
    if not student:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    return await serve_student_constancia(student, school_id)


async def serve_student_constancia(student: dict, school_id: str):
    """Return the ACTIVE constancia as a PDF Response: custom (Drive) if enabled, else generated default."""
    from fastapi.responses import Response
    from services.constancia_pdf_generator import generate_constancia_pdf

    student_id = student.get("id")
    custom = student.get("constancia_custom")
    if student.get("constancia_use_custom") and custom and custom.get("drive_file_id"):
        from routes.report_cards_pdf import _stream_drive_pdf
        return await _stream_drive_pdf(
            school_id, custom["drive_file_id"], custom.get("file_name") or "Constancia_Matricula.pdf"
        )

    school = await db.schools.find_one({"id": school_id}, {"_id": 0}) or {}
    year, periodo_del, periodo_al = await _constancia_period(school_id)
    codigo_modular = school.get("codigo_modular") or ""
    ruc = (school.get("libreta_stamp_config") or {}).get("ruc") or ""

    detail = await _constancia_student_detail(student, school_id)
    pdf_bytes = generate_constancia_pdf(
        school=school, student=student,
        level_name=detail["level_name"], grade_name=detail["grade_name"],
        section_name=detail["section_name"], year=year,
        turno_name=detail["turno_name"], apoderado_name=detail["apoderado_name"],
        codigo_modular=codigo_modular, ruc=ruc,
        periodo_del=periodo_del, periodo_al=periodo_al,
    )
    safe_name = re.sub(r"[^A-Za-z0-9]+", "_", f"{student.get('name','')}_{student.get('last_name','')}").strip("_")
    filename = f"Constancia_Matricula_{safe_name or student_id}.pdf"
    return Response(
        content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _image_to_pdf(image_bytes: bytes) -> bytes:
    """Convert an image (jpg/png/webp) into a single-page PDF, fit & centered on the page.
    Uses the image's own orientation (portrait/landscape) to avoid distortion."""
    from PIL import Image
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as _canvas
    from reportlab.lib.utils import ImageReader

    img = Image.open(io.BytesIO(image_bytes))
    if img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    iw, ih = img.size
    page_w, page_h = landscape(A4) if iw >= ih else A4
    margin = 1.0 * cm
    max_w, max_h = page_w - 2 * margin, page_h - 2 * margin
    ratio = min(max_w / iw, max_h / ih)
    draw_w, draw_h = iw * ratio, ih * ratio
    x = (page_w - draw_w) / 2
    y = (page_h - draw_h) / 2

    img_buf = io.BytesIO()
    img.save(img_buf, format="JPEG", quality=90)
    img_buf.seek(0)

    out = io.BytesIO()
    c = _canvas.Canvas(out, pagesize=(page_w, page_h))
    c.drawImage(ImageReader(img_buf), x, y, width=draw_w, height=draw_h, preserveAspectRatio=True)
    c.showPage()
    c.save()
    out.seek(0)
    return out.read()



async def _ensure_constancias_folder(service, materials_folder_id: str) -> str:
    """Find or create the 'Constancias' subfolder inside the school's materials folder."""
    q = (
        f"name='Constancias' and '{materials_folder_id}' in parents and "
        f"mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    res = service.files().list(q=q, fields="files(id)").execute()
    items = res.get("files", [])
    if items:
        return items[0]["id"]
    meta = {"name": "Constancias", "mimeType": "application/vnd.google-apps.folder", "parents": [materials_folder_id]}
    f = service.files().create(body=meta, fields="id").execute()
    return f.get("id")


@router.post("/students/{student_id}/constancia-custom")
async def upload_constancia_custom(student_id: str, file: UploadFile = File(...),
                                   current_user=Depends(get_current_user)):
    """Sube una Constancia de Matrícula personalizada (PDF) al Drive del colegio. Solo admin/director/owner."""
    from googleapiclient.http import MediaIoBaseUpload
    from routes.exams import get_drive_service

    user = await resolve_user_from_token(current_user)
    if not user or user.get("role") not in ("owner", "admin", "director"):
        raise HTTPException(status_code=403, detail="No tienes permisos para subir constancias")
    school_id = user.get("school_id")

    student = await db.users.find_one(
        {"id": student_id, "school_id": school_id, "role": "student"}, {"_id": 0}
    )
    if not student:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    ctype = (file.content_type or "").lower()
    fname_lower = (file.filename or "").lower()
    is_pdf = ctype == "application/pdf" or fname_lower.endswith(".pdf")
    is_image = ctype in ("image/jpeg", "image/jpg", "image/png", "image/webp") or \
        fname_lower.endswith((".jpg", ".jpeg", ".png", ".webp"))
    if not (is_pdf or is_image or ctype == "application/octet-stream"):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF, JPG, JPEG, PNG o WEBP")
    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo supera el límite de 10 MB")

    # Si es imagen, convertir a PDF (una imagen = una página, ajustada a la hoja, centrada).
    if is_image and not is_pdf:
        try:
            content = _image_to_pdf(content)
        except Exception as e:
            logger.exception("Error convirtiendo imagen a PDF")
            raise HTTPException(status_code=400, detail="No se pudo procesar la imagen. Verifica que sea un JPG, PNG o WEBP válido.")

    school = await db.schools.find_one({"id": school_id}, {"_id": 0}) or {}
    if not school.get("google_drive_connected"):
        raise HTTPException(status_code=409, detail="Google Drive no está conectado. Ve a Ajustes → Integraciones para conectarlo.")
    materials_folder_id = school.get("google_drive_materials_folder_id")
    if not materials_folder_id:
        raise HTTPException(status_code=409, detail="Google Drive no tiene carpeta de materiales configurada.")

    try:
        service = await get_drive_service(school_id)
        folder_id = await _ensure_constancias_folder(service, materials_folder_id)
        safe_label = f"{student.get('last_name','')}_{student.get('name','')}".strip().replace(" ", "_") or student_id
        media = MediaIoBaseUpload(io.BytesIO(content), mimetype="application/pdf", resumable=True)
        drive_file = service.files().create(
            body={"name": f"Constancia_{safe_label}.pdf", "parents": [folder_id]},
            media_body=media, fields="id, name",
        ).execute()
        drive_file_id = drive_file.get("id")
    except HTTPException:
        raise
    except Exception as e:
        err = str(e).lower()
        if "invalid_grant" in err or "token has been expired" in err or "token has been revoked" in err:
            await db.schools.update_one({"id": school_id}, {"$set": {"google_drive_connected": False, "google_drive_token_invalidated_at": now_iso()}})
            raise HTTPException(status_code=409, detail="La conexión con Google Drive expiró o fue revocada. Reconéctala en Ajustes → Integraciones → Google Drive.")
        logger.exception("Failed to upload constancia to Drive")
        raise HTTPException(status_code=502, detail=f"Error subiendo a Google Drive: {e}")

    base_name = re.sub(r"\.(pdf|jpg|jpeg|png|webp)$", "", file.filename or "", flags=re.IGNORECASE) or f"Constancia_{student_id}"
    constancia_custom = {
        "drive_file_id": drive_file_id,
        "file_name": f"{base_name}.pdf",
        "file_size": len(content),
        "storage_type": "google_drive",
        "uploaded_by": user["id"],
        "uploaded_at": now_iso(),
    }
    await db.users.update_one({"id": student_id}, {"$set": {"constancia_custom": constancia_custom, "constancia_use_custom": True}})
    return {"ok": True, "constancia_custom": constancia_custom, "constancia_use_custom": True}


class ConstanciaModeBody(BaseModel):
    use_custom: bool


@router.put("/students/{student_id}/constancia-mode")
async def set_constancia_mode(student_id: str, body: ConstanciaModeBody, current_user=Depends(get_current_user)):
    """Activa/desactiva la constancia personalizada. use_custom=False → usa la constancia por defecto generada."""
    user = await resolve_user_from_token(current_user)
    if not user or user.get("role") not in ("owner", "admin", "director"):
        raise HTTPException(status_code=403, detail="No tienes permisos")
    school_id = user.get("school_id")
    student = await db.users.find_one({"id": student_id, "school_id": school_id, "role": "student"}, {"_id": 0, "id": 1, "constancia_custom": 1})
    if not student:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    if body.use_custom and not (student.get("constancia_custom") or {}).get("drive_file_id"):
        raise HTTPException(status_code=400, detail="Primero sube el PDF de la constancia personalizada")
    await db.users.update_one({"id": student_id}, {"$set": {"constancia_use_custom": body.use_custom}})
    return {"ok": True, "constancia_use_custom": body.use_custom}


@router.delete("/students/{student_id}/constancia-custom")
async def delete_constancia_custom(student_id: str, current_user=Depends(get_current_user)):
    """Quita la constancia personalizada (vuelve a usar la predeterminada). Solo admin/director/owner."""
    user = await resolve_user_from_token(current_user)
    if not user or user.get("role") not in ("owner", "admin", "director"):
        raise HTTPException(status_code=403, detail="No tienes permisos")
    school_id = user.get("school_id")
    student = await db.users.find_one({"id": student_id, "school_id": school_id, "role": "student"}, {"_id": 0, "id": 1, "constancia_custom": 1})
    if not student:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    custom = student.get("constancia_custom")
    if custom and custom.get("drive_file_id"):
        try:
            from routes.exams import get_drive_service
            service = await get_drive_service(school_id)
            service.files().delete(fileId=custom["drive_file_id"]).execute()
        except Exception:
            logger.warning(f"No se pudo borrar el archivo de Drive para {student_id} (se deja huérfano)")
    await db.users.update_one({"id": student_id}, {"$unset": {"constancia_custom": "", "constancia_use_custom": ""}})
    return {"ok": True}



async def _constancia_period(school_id: str):
    """Return (year, periodo_del dd/mm/yyyy, periodo_al dd/mm/yyyy) from academic_periods."""
    ay = await db.academic_years.find_one({"school_id": school_id}, {"_id": 0, "id": 1, "year": 1}, sort=[("year", -1)])
    year = str(ay.get("year")) if ay and ay.get("year") else str(datetime.now(timezone.utc).year)
    q = {"school_id": school_id}
    if ay and ay.get("id"):
        q["academic_year_id"] = ay["id"]
    periods = await db.academic_periods.find(q, {"_id": 0, "fecha_inicio": 1, "fecha_fin": 1}).to_list(50)
    starts = sorted([p.get("fecha_inicio") for p in periods if p.get("fecha_inicio")])
    ends = sorted([p.get("fecha_fin") for p in periods if p.get("fecha_fin")])

    def _fmt(iso):
        try:
            return datetime.strptime(iso[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return ""
    periodo_del = _fmt(starts[0]) if starts else f"01/03/{year}"
    periodo_al = _fmt(ends[-1]) if ends else f"31/12/{year}"
    return year, periodo_del, periodo_al


async def _constancia_student_detail(student: dict, school_id: str):
    level = await db.academic_levels.find_one({"id": student.get("nivel_id")}, {"_id": 0, "nombre": 1})
    grade = await db.grades.find_one({"id": student.get("grado_id")}, {"_id": 0, "nombre": 1})
    section = await db.sections.find_one({"id": student.get("seccion_id")}, {"_id": 0, "nombre": 1})
    turno_name = ""
    if student.get("turno_id"):
        shift = await db.shifts.find_one({"id": student.get("turno_id")}, {"_id": 0, "nombre": 1})
        turno_name = (shift or {}).get("nombre", "") or ""
    apoderado_name = ""
    if student.get("padre_id"):
        parent = await db.users.find_one({"id": student.get("padre_id")}, {"_id": 0, "name": 1, "last_name": 1})
        if parent:
            apoderado_name = f"{parent.get('last_name', '') or ''}, {parent.get('name', '') or ''}".strip(", ")
    return {
        "level_name": (level or {}).get("nombre", "") or "",
        "grade_name": (grade or {}).get("nombre", "") or "",
        "section_name": (section or {}).get("nombre", "") or "",
        "turno_name": turno_name,
        "apoderado_name": apoderado_name,
    }


@router.get("/sections/{section_id}/constancias-matricula")
async def download_constancias_section(section_id: str, current_user=Depends(get_current_user)):
    """Genera un solo PDF con la Constancia de Matrícula de todos los alumnos activos de la sección."""
    from fastapi.responses import Response
    from services.constancia_pdf_generator import generate_constancias_batch_pdf

    user = await resolve_user_from_token(current_user)
    if not user or user.get("role") not in ("owner", "admin", "director"):
        raise HTTPException(status_code=403, detail="No tienes permisos para generar constancias")
    school_id = user.get("school_id")

    students = await db.users.find(
        {"school_id": school_id, "role": "student", "seccion_id": section_id},
        {"_id": 0},
    ).to_list(1000)
    students = [s for s in students if s.get("student_status") != "withdrawn" and s.get("is_active") is not False]
    if not students:
        raise HTTPException(status_code=404, detail="La sección no tiene alumnos activos")
    students.sort(key=lambda s: (s.get("last_name") or "").upper())

    school = await db.schools.find_one({"id": school_id}, {"_id": 0}) or {}
    year, periodo_del, periodo_al = await _constancia_period(school_id)
    codigo_modular = school.get("codigo_modular") or ""
    ruc = (school.get("libreta_stamp_config") or {}).get("ruc") or ""

    items = []
    for st in students:
        d = await _constancia_student_detail(st, school_id)
        items.append({"student": st, **d})

    pdf_bytes = generate_constancias_batch_pdf(
        items=items, school=school, year=year,
        codigo_modular=codigo_modular, ruc=ruc,
        periodo_del=periodo_del, periodo_al=periodo_al,
    )
    section = await db.sections.find_one({"id": section_id}, {"_id": 0, "nombre": 1})
    sec_name = re.sub(r"[^A-Za-z0-9]+", "_", (section or {}).get("nombre", "") or "seccion").strip("_")
    filename = f"Constancias_Matricula_{sec_name}.pdf"
    return Response(
        content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# USERS MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/users")
async def get_tenant_users(current_user = Depends(get_current_user)):
    """
    Get all users for the current tenant.
    Only admins/directors/owners/super_admins can view users.
    System users (Admin Técnico) are only visible to owners and system_admins.
    """
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    
    # Check role - owners, super_admins, directors and admins can view users
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver usuarios")
    
    school_id = user["school_id"]
    
    # Build query - system users only visible to owners and system_admins
    query = {"school_id": school_id, "student_status": {"$ne": "deleted"}}
    
    # Only owners and system_admins can see system users
    can_see_system_users = (
        user.get("is_owner") == True or 
        user.get("role") == "system_admin" or
        user.get("is_super_admin") == True
    )
    
    if not can_see_system_users:
        query["is_system_user"] = {"$ne": True}
    
    # Get all users for this school.
    # NOTE: do NOT cap with a small length here — the frontend loads the full
    # set once and splits it client-side into tabs (students / parents / teachers).
    # A small cap (previously 1000) with no sort silently truncated the most
    # recently created users (e.g. freshly imported teachers landed past the cap
    # in schools with 1000+ students/parents), making them "disappear" from the
    # Teachers tab even though they exist in the DB.
    users_cursor = db.users.find(
        query,
        {"_id": 0, "password": 0, "verification_code": 0}
    )
    users = await users_cursor.to_list(length=None)
    
    return users

@router.get("/users/{user_id}")
async def get_user_by_id(user_id: str, current_user = Depends(get_current_user)):
    """Get a specific user by ID"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    
    # Check role - owners, super_admins, directors and admins can view users
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver usuarios")
    
    target_user = await db.users.find_one(
        {"id": user_id, "school_id": user["school_id"]},
        {"_id": 0, "password": 0, "verification_code": 0}
    )
    
    if not target_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    return target_user

class CreateUserRequest(BaseModel):
    """Request to create a new user"""
    username: Optional[str] = None  # Auto-generated for `personal_mantenimiento`
    password: Optional[str] = None  # Auto-generated for `personal_mantenimiento`
    name: str
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    birthday: Optional[str] = None
    gender: Optional[str] = None
    address: Optional[str] = None
    role: str = "teacher"
    photo_url: Optional[str] = None
    # Demo user flag - only owner can create demo users
    is_demo_user: Optional[bool] = False
    # Academic fields for students
    nivel_id: Optional[str] = None
    grado_id: Optional[str] = None
    seccion_id: Optional[str] = None
    turno_id: Optional[str] = None
    padre_id: Optional[str] = None  # Link student to parent
    # Student complementary info
    condiciones_medicas: Optional[str] = None
    alergias: Optional[str] = None
    doctor_nombre: Optional[str] = None
    doctor_telefono: Optional[str] = None
    persona_autorizada: Optional[str] = None
    persona_autorizada_telefono: Optional[str] = None
    notas: Optional[str] = None
    # Parent-specific fields
    dni: Optional[str] = None
    ocupacion: Optional[str] = None
    lugar_trabajo: Optional[str] = None
    telefono_trabajo: Optional[str] = None
    # Maintenance-specific fields
    maintenance_role: Optional[str] = None
    maintenance_role_custom: Optional[str] = None

@router.get("/users/check-username/{username}")
async def check_username(username: str, current_user = Depends(get_current_user)):
    """Check if username is available"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    
    existing = await db.users.find_one({
        "username": username.lower(),
        "school_id": user["school_id"]
    })
    
    return {
        "available": existing is None,
        "username": username
    }

@router.get("/users/students/search")
async def search_students(q: str = "", current_user = Depends(get_current_user)):
    """Search students by name for autocomplete."""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    
    if not q or len(q) < 2:
        return []
    
    import re
    regex = re.compile(re.escape(q), re.IGNORECASE)
    students = await db.users.find(
        {
            "school_id": user["school_id"],
            "role": "student",
            "$or": [
                {"name": {"$regex": regex}},
                {"last_name": {"$regex": regex}},
            ]
        },
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "grado_id": 1, "seccion_id": 1}
    ).limit(10).to_list(10)
    
    # Enrich with grade/section names
    for s in students:
        grade = await db.grades.find_one({"id": s.get("grado_id")}, {"_id": 0, "nombre": 1})
        section = await db.sections.find_one({"id": s.get("seccion_id")}, {"_id": 0, "nombre": 1})
        s["grade_name"] = grade.get("nombre", "") if grade else ""
        s["section_name"] = section.get("nombre", "") if section else ""
    
    return students


@router.patch("/students/{student_id}/toggle-disable")
async def toggle_student_disable(student_id: str, current_user = Depends(get_current_user)):
    """Alterna is_disabled de un alumno (desactivación temporal / retiro).

    - Al DESACTIVAR: registra disabled_at y RESETEA credenciales con valores
      aleatorios (username 8 alfanum., password 12). Solo se guarda el hash
      bcrypt; se borra plain_password/password_display. Devuelve las nuevas
      credenciales en texto plano UNA sola vez (no se almacenan en claro).
    - Al REACTIVAR: solo limpia disabled_at. No toca credenciales.
    Solo admin / owner / director.
    """
    import secrets, string
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden desactivar alumnos")

    school_id = user["school_id"]
    student = await db.users.find_one({"id": student_id, "school_id": school_id, "role": "student"})
    if not student:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    currently_disabled = student.get("is_disabled") is True

    if not currently_disabled:
        # DESACTIVAR + reset de credenciales
        alnum = string.ascii_letters + string.digits
        lower_num = string.ascii_lowercase + string.digits
        new_username = "".join(secrets.choice(lower_num) for _ in range(8))
        for _ in range(6):
            if not await db.users.find_one({"school_id": school_id, "username": new_username}, {"_id": 1}):
                break
            new_username = "".join(secrets.choice(lower_num) for _ in range(8))
        new_password = "".join(secrets.choice(alnum) for _ in range(12))
        await db.users.update_one(
            {"id": student_id},
            {
                "$set": {
                    "is_disabled": True,
                    "disabled_at": now_iso(),
                    "username": new_username,
                    "password": hash_password(new_password),
                    "updated_at": now_iso(),
                },
                "$unset": {"plain_password": "", "password_display": ""},
            },
        )
        invalidate_student_cache(student_id)
        return {
            "is_disabled": True,
            "message": "Alumno desactivado. Ya no aparecerá en el sistema y sus credenciales fueron reseteadas.",
            "credentials": {"username": new_username, "password": new_password},
        }
    else:
        # REACTIVAR (no toca credenciales)
        await db.users.update_one(
            {"id": student_id},
            {"$set": {"is_disabled": False, "updated_at": now_iso()}, "$unset": {"disabled_at": ""}},
        )
        invalidate_student_cache(student_id)
        return {"is_disabled": False, "message": "Alumno reactivado correctamente."}


@router.get("/students/disabled/search")
async def search_disabled_students(q: str = "", current_user = Depends(get_current_user)):
    """Autocompletador de Ajustes: busca SOLO entre alumnos desactivados (retirados)."""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    query = {"school_id": user["school_id"], "role": "student", "is_disabled": True}
    if q and len(q.strip()) >= 1:
        regex = re.compile(re.escape(q.strip()), re.IGNORECASE)
        query["$or"] = [
            {"name": {"$regex": regex}},
            {"last_name": {"$regex": regex}},
            {"dni": {"$regex": regex}},
        ]
    students = await db.users.find(
        query,
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "dni": 1, "grado_id": 1,
         "seccion_id": 1, "nivel_id": 1, "disabled_at": 1, "username": 1, "photo_url": 1},
    ).limit(20).to_list(20)
    for s in students:
        grade = await db.grades.find_one({"id": s.get("grado_id")}, {"_id": 0, "nombre": 1})
        section = await db.sections.find_one({"id": s.get("seccion_id")}, {"_id": 0, "nombre": 1})
        s["grade_name"] = grade.get("nombre", "") if grade else ""
        s["section_name"] = section.get("nombre", "") if section else ""
    return students


@router.get("/parents/filter-by-children")
async def parents_filter_by_children(nivel_id: str = "", grado_id: str = "", seccion_id: str = "", current_user = Depends(get_current_user)):
    """Devuelve los IDs de padres/apoderados que tienen AL MENOS un hijo que
    cumple los filtros (nivel/grado/sección). Liviano: solo ids + conteo."""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    school_id = user["school_id"]
    sfilter = {"role": "student", "school_id": school_id, "is_active": {"$ne": False}}
    if nivel_id:
        sfilter["nivel_id"] = nivel_id
    if grado_id:
        sfilter["grado_id"] = grado_id
    if seccion_id:
        sfilter["seccion_id"] = seccion_id

    students = await db.users.find(sfilter, {"_id": 0, "id": 1, "padre_id": 1, "parent_id": 1}).to_list(None)
    matching_ids = [s["id"] for s in students]
    parent_ids = set()
    for s in students:
        for pid in (s.get("padre_id"), s.get("parent_id")):
            if pid:
                parent_ids.add(pid)
    if matching_ids:
        linked = await db.users.find(
            {"role": "parent", "school_id": school_id,
             "$or": [{"student_ids": {"$in": matching_ids}}, {"children_ids": {"$in": matching_ids}}]},
            {"_id": 0, "id": 1},
        ).to_list(None)
        for p in linked:
            parent_ids.add(p["id"])

    return {"parent_ids": list(parent_ids), "count": len(parent_ids)}


@router.post("/users")
async def create_user(data: CreateUserRequest, current_user = Depends(get_current_user)):
    """
    Create a new user for the current tenant.
    Only admins/owners can create users.
    Demo users can only be created by the real owner.
    """
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    
    # Block demo users from creating users
    check_demo_user_block(user)
    
    # Check role - only owner or admin can create users
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden crear usuarios")
    
    # Only real owner can create demo users
    if data.is_demo_user:
        if not is_real_owner(user):
            raise HTTPException(status_code=403, detail="Solo el propietario puede crear usuarios demo")
    
    school_id = user["school_id"]

    # Auto-generate credentials for personal_mantenimiento (no login required)
    if data.role == "personal_mantenimiento":
        # Validate maintenance sub-role
        if not data.maintenance_role or not data.maintenance_role.strip():
            raise HTTPException(status_code=400, detail="Debes seleccionar un rol de mantenimiento")
        allowed_roles = {"limpieza", "vigilancia", "guardianía", "porteria", "otro"}
        if data.maintenance_role not in allowed_roles:
            raise HTTPException(status_code=400, detail="Rol de mantenimiento no válido")
        if data.maintenance_role == "otro" and (not data.maintenance_role_custom or not data.maintenance_role_custom.strip()):
            raise HTTPException(status_code=400, detail="Debes especificar el rol")
        if not data.username or not data.username.strip():
            base = (data.dni or str(uuid.uuid4())[:8]).strip()
            data.username = f"mant_{base}"
        if not data.password or not data.password.strip():
            data.password = str(uuid.uuid4())[:12]
    else:
        # Other roles still require username + password
        if not data.username or not data.password:
            raise HTTPException(status_code=400, detail="Usuario y contraseña son obligatorios")

    # Check if username already exists in this school
    existing = await db.users.find_one({
        "username": data.username.lower(),
        "school_id": school_id
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="El usuario ya existe")
    
    # Email duplicates allowed (siblings may share parent's email)
    
    # Create user
    new_user = {
        "id": str(uuid.uuid4()),
        "username": data.username.lower(),
        "password": hash_password(data.password),
        "plain_password": data.password,
        "password_display": data.password,
        "name": data.name,
        "last_name": data.last_name,
        "email": data.email.lower() if data.email else None,
        "phone": data.phone,
        "birthday": data.birthday,
        "gender": data.gender,
        "address": data.address,
        "role": data.role,
        "photo_url": data.photo_url,
        "school_id": school_id,
        "email_verified": True,  # Created by admin, no verification needed
        "is_demo_user": data.is_demo_user or False,  # Demo user flag
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Add academic fields for students (support both 'student' and 'estudiante' roles)
    if data.role in ["student", "estudiante"]:
        new_user["nivel_id"] = data.nivel_id
        new_user["grado_id"] = data.grado_id
        new_user["seccion_id"] = data.seccion_id
        new_user["turno_id"] = data.turno_id
        # Check activation mode from financial settings
        fin_settings = await db.school_financial_settings.find_one(
            {"school_id": school_id}, {"_id": 0, "activacion_modo": 1}
        )
        activation_mode = (fin_settings or {}).get("activacion_modo", "on_create")
        if activation_mode == "on_create":
            new_user["student_status"] = "active"
        else:
            new_user["student_status"] = "pending"
        if data.padre_id:
            new_user["padre_id"] = data.padre_id
        # Complementary info
        if data.condiciones_medicas:
            new_user["condiciones_medicas"] = data.condiciones_medicas
        if data.alergias:
            new_user["alergias"] = data.alergias
        if data.doctor_nombre:
            new_user["doctor_nombre"] = data.doctor_nombre
        if data.doctor_telefono:
            new_user["doctor_telefono"] = f"+51{data.doctor_telefono}" if data.doctor_telefono and not data.doctor_telefono.startswith("+") else data.doctor_telefono
        if data.persona_autorizada:
            new_user["persona_autorizada"] = data.persona_autorizada
        if data.persona_autorizada_telefono:
            new_user["persona_autorizada_telefono"] = f"+51{data.persona_autorizada_telefono}" if data.persona_autorizada_telefono and not data.persona_autorizada_telefono.startswith("+") else data.persona_autorizada_telefono
        if data.notas:
            new_user["notas"] = data.notas
        
        # Generate short QR (centralized service)
        qr_id, qr_token = await generate_user_qr(db)
        new_user["qr_id"] = qr_id
        new_user["qr_token"] = qr_token
        new_user["qr_version"] = 2
    
    # Add DNI for all roles
    if data.dni:
        new_user["dni"] = data.dni

    # Add parent-specific fields
    if data.role == "parent":
        new_user["ocupacion"] = data.ocupacion
        new_user["lugar_trabajo"] = data.lugar_trabajo
        new_user["telefono_trabajo"] = data.telefono_trabajo
    
    # Generate short QR for teachers (centralized service)
    if data.role == "teacher":
        qr_id, qr_token = await generate_user_qr(db)
        new_user["qr_id"] = qr_id
        new_user["qr_token"] = qr_token
        new_user["qr_version"] = 2

    # Generate short QR for personal_mantenimiento (used for attendance)
    if data.role == "personal_mantenimiento":
        qr_id, qr_token = await generate_user_qr(db)
        new_user["qr_id"] = qr_id
        new_user["qr_token"] = qr_token
        new_user["qr_version"] = 2
        # Persist maintenance sub-role
        new_user["maintenance_role"] = data.maintenance_role
        new_user["maintenance_role_custom"] = data.maintenance_role_custom if data.maintenance_role == "otro" else None

    # Generate short QR for auxiliary staff + psicólogos (so they can scan-in
    # their own attendance from Asistencia → Personal Administrativo, same as
    # students and teachers). Applies to all `auxiliar_*` sub-roles + psicologo.
    AUX_ROLES_WITH_QR = {"auxiliar", "auxiliar_asistencia", "auxiliar_alimentacion", "auxiliar_movilidad", "auxiliar_topico", "psicologo"}
    if data.role in AUX_ROLES_WITH_QR:
        qr_id, qr_token = await generate_user_qr(db)
        new_user["qr_id"] = qr_id
        new_user["qr_token"] = qr_token
        new_user["qr_version"] = 2
    
    await db.users.insert_one(new_user)
    
    # Remove sensitive fields before returning
    del new_user["password"]
    if "_id" in new_user:
        del new_user["_id"]
    
    logger.info(f"User created: {data.username} with role {data.role} in school {school_id}")
    
    return {
        "message": "Usuario creado correctamente",
        "user": new_user
    }

class UpdateUserRequest(BaseModel):
    """Request to update an existing user"""
    name: Optional[str] = None
    last_name: Optional[str] = None
    username: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    birthday: Optional[str] = None
    gender: Optional[str] = None
    address: Optional[str] = None
    role: Optional[str] = None
    photo_url: Optional[str] = None
    password: Optional[str] = None  # For password changes
    # Academic fields for students
    nivel_id: Optional[str] = None
    grado_id: Optional[str] = None
    seccion_id: Optional[str] = None
    turno_id: Optional[str] = None
    padre_id: Optional[str] = None
    parent_id: Optional[str] = None  # Alias for padre_id (frontend compatibility)
    # Student complementary info
    condiciones_medicas: Optional[str] = None
    alergias: Optional[str] = None
    doctor_nombre: Optional[str] = None
    doctor_telefono: Optional[str] = None
    persona_autorizada: Optional[str] = None
    persona_autorizada_telefono: Optional[str] = None
    notas: Optional[str] = None
    # Parent-specific fields
    dni: Optional[str] = None
    ocupacion: Optional[str] = None
    lugar_trabajo: Optional[str] = None
    telefono_trabajo: Optional[str] = None
    # Maintenance-specific fields
    maintenance_role: Optional[str] = None
    maintenance_role_custom: Optional[str] = None

@router.put("/users/{user_id}")
async def update_user(user_id: str, data: UpdateUserRequest, current_user = Depends(get_current_user)):
    """Update an existing user"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden editar usuarios")
    
    # Find target user
    target = await db.users.find_one({"id": user_id, "school_id": user["school_id"]})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # SYSTEM USERS CANNOT BE EDITED AT ALL
    if target.get("is_system_user"):
        raise HTTPException(status_code=403, detail=SYSTEM_USER_BLOCKED_MESSAGE)
    
    # Cannot change role of protected users (owner)
    if (target.get("is_protected") or target.get("is_owner")) and data.role and data.role != target.get("role"):
        raise HTTPException(status_code=400, detail="No se puede cambiar el rol del propietario de la intranet")
    
    # Build update data
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if data.name is not None:
        update_data["name"] = data.name
    if data.last_name is not None:
        update_data["last_name"] = data.last_name
    if data.username is not None:
        # Check if username is already used by another user in the same school
        existing_username = await db.users.find_one({
            "username": data.username.lower(),
            "school_id": user["school_id"],
            "id": {"$ne": user_id}
        })
        if existing_username:
            raise HTTPException(status_code=400, detail="Este nombre de usuario ya está en uso")
        update_data["username"] = data.username.lower()
    if data.email is not None:
        update_data["email"] = data.email.lower()
    if data.phone is not None:
        update_data["phone"] = data.phone
    if data.birthday is not None:
        update_data["birthday"] = data.birthday
    if data.gender is not None:
        update_data["gender"] = data.gender
    if data.address is not None:
        update_data["address"] = data.address
    if data.role is not None:
        update_data["role"] = data.role
    if data.photo_url is not None:
        update_data["photo_url"] = data.photo_url
    # Academic fields
    if data.nivel_id is not None:
        update_data["nivel_id"] = data.nivel_id
    if data.grado_id is not None:
        update_data["grado_id"] = data.grado_id
    if data.seccion_id is not None:
        update_data["seccion_id"] = data.seccion_id
    if data.turno_id is not None:
        update_data["turno_id"] = data.turno_id
    if data.padre_id is not None:
        update_data["padre_id"] = data.padre_id
    # Student medical/contact info
    if data.condiciones_medicas is not None:
        update_data["condiciones_medicas"] = data.condiciones_medicas
    if data.alergias is not None:
        update_data["alergias"] = data.alergias
    if data.doctor_nombre is not None:
        update_data["doctor_nombre"] = data.doctor_nombre
    if data.doctor_telefono is not None:
        update_data["doctor_telefono"] = data.doctor_telefono
    if data.persona_autorizada is not None:
        update_data["persona_autorizada"] = data.persona_autorizada
    if data.persona_autorizada_telefono is not None:
        update_data["persona_autorizada_telefono"] = data.persona_autorizada_telefono
    if data.notas is not None:
        update_data["notas"] = data.notas
    # Parent fields
    if data.dni is not None:
        update_data["dni"] = data.dni
    if data.ocupacion is not None:
        update_data["ocupacion"] = data.ocupacion
    if data.lugar_trabajo is not None:
        update_data["lugar_trabajo"] = data.lugar_trabajo
    if data.telefono_trabajo is not None:
        update_data["telefono_trabajo"] = data.telefono_trabajo
    
    # Handle password change
    if data.password is not None and data.password.strip():
        update_data["password"] = hash_password(data.password)
        update_data["plain_password"] = data.password
        update_data["password_display"] = data.password
        logger.info(f"Password changed for user {user_id}")
    
    # Handle parent_id (frontend sends parent_id, backend uses padre_id)
    if data.parent_id is not None:
        update_data["padre_id"] = data.parent_id if data.parent_id else None
        update_data["parent_id"] = data.parent_id if data.parent_id else None

    # Handle maintenance sub-role (for personal_mantenimiento)
    effective_role = data.role or target.get("role")
    if effective_role == "personal_mantenimiento":
        # If maintenance_role is explicitly provided, validate it
        if data.maintenance_role is not None:
            if not data.maintenance_role or not data.maintenance_role.strip():
                raise HTTPException(status_code=400, detail="Debes seleccionar un rol de mantenimiento")
            allowed_roles = {"limpieza", "vigilancia", "guardianía", "porteria", "otro"}
            if data.maintenance_role not in allowed_roles:
                raise HTTPException(status_code=400, detail="Rol de mantenimiento no válido")
            if data.maintenance_role == "otro":
                if not data.maintenance_role_custom or not data.maintenance_role_custom.strip():
                    raise HTTPException(status_code=400, detail="Debes especificar el rol")
                update_data["maintenance_role"] = data.maintenance_role
                update_data["maintenance_role_custom"] = data.maintenance_role_custom
            else:
                update_data["maintenance_role"] = data.maintenance_role
                update_data["maintenance_role_custom"] = None
    
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    # Return updated user
    updated_user = await db.users.find_one(
        {"id": user_id},
        {"_id": 0, "password": 0, "verification_code": 0}
    )
    
    logger.info(f"User {user_id} updated by {user['id']}")
    
    return {"message": "Usuario actualizado correctamente", "user": updated_user}


class TeacherActiveRequest(BaseModel):
    active: bool


def _generate_temp_password(length: int = 8) -> str:
    """Readable temporary password (avoids ambiguous chars 0/O/1/l/I)."""
    import secrets
    alphabet = "ABCDEFGHJKMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(length))


@router.patch("/users/teachers/{teacher_id}/active")
async def set_teacher_active(teacher_id: str, data: TeacherActiveRequest,
                             current_user = Depends(get_current_user)):
    """Activate / deactivate a TEACHER account from the system.

    - Deactivate (active=false): sets status="inactivo" and RESETS the password
      to an unguessable random value so the teacher can no longer log in
      (login is also blocked by the status check in auth.login). Idempotent.
    - Activate (active=true): sets status="activo" and generates a NEW temporary
      password, returned in plaintext ONCE so the admin can hand it to the
      teacher. The teacher then logs in with it.
    Owner/admin only. Restricted to role == "teacher"."""
    import secrets
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden cambiar el estado de un profesor")

    target = await db.users.find_one({"id": teacher_id, "school_id": user["school_id"]})
    if not target:
        raise HTTPException(status_code=404, detail="Profesor no encontrado")
    if target.get("role") != "teacher":
        raise HTTPException(status_code=400, detail="Solo se puede activar/desactivar a profesores")
    if target.get("is_system_user"):
        raise HTTPException(status_code=403, detail=SYSTEM_USER_BLOCKED_MESSAGE)
    if target.get("is_protected") or target.get("is_owner"):
        raise HTTPException(status_code=400, detail="No se puede desactivar al propietario")

    now = datetime.now(timezone.utc).isoformat()
    temp_password = None

    if data.active:
        # Reactivate: issue a fresh temporary password.
        temp_password = _generate_temp_password()
        update_data = {
            "status": "activo",
            "password": hash_password(temp_password),
            "must_change_password": True,
            "deactivated_at": None,
            "reactivated_at": now,
            "updated_at": now,
        }
    else:
        # Deactivate: scramble the password so the current one stops working.
        update_data = {
            "status": "inactivo",
            "password": hash_password(secrets.token_urlsafe(32)),
            "deactivated_at": now,
            "updated_at": now,
        }

    await db.users.update_one({"id": teacher_id}, {"$set": update_data})
    logger.info(f"Teacher {teacher_id} {'activated' if data.active else 'deactivated'} by {user['id']}")

    return {
        "message": "Profesor activado" if data.active else "Profesor desactivado",
        "status": update_data["status"],
        "temp_password": temp_password,  # null on deactivate
    }


class TopicoContactPermissionRequest(BaseModel):
    can_view: bool


@router.patch("/users/topico/{user_id}/contact-permission")
async def set_topico_contact_permission(user_id: str, data: TopicoContactPermissionRequest,
                                        current_user = Depends(get_current_user)):
    """Owner/admin only: toggle the "Información Paciente" permission for an
    auxiliar_topico (nurse) user. When True, that nurse can view a student's
    contact data (phone + parents' name/phone) in the Tópico portal. Persisted
    per-user as `can_view_patient_contact`."""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden cambiar este permiso")

    target = await db.users.find_one({"id": user_id, "school_id": user["school_id"]})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if target.get("role") != "auxiliar_topico":
        raise HTTPException(status_code=400, detail="Este permiso solo aplica a usuarios de Tópico")

    await db.users.update_one(
        {"id": user_id},
        {"$set": {"can_view_patient_contact": bool(data.can_view),
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    logger.info(f"Topico {user_id} contact permission set to {data.can_view} by {user['id']}")
    return {"message": "Permiso actualizado", "can_view_patient_contact": bool(data.can_view)}


@router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user = Depends(get_current_user)):
    """Delete a user and all their related data (cascade delete)"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar usuarios")
    
    # Cannot delete yourself
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")
    
    # Find target user
    target = await db.users.find_one({"id": user_id, "school_id": user["school_id"]})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # SYSTEM USERS CANNOT BE DELETED
    if target.get("is_system_user"):
        raise HTTPException(status_code=403, detail=SYSTEM_USER_BLOCKED_MESSAGE)
    
    # PROTECTED USERS CANNOT BE DELETED (owner, super_admin)
    if target.get("is_protected") or target.get("is_owner") or target.get("is_super_admin"):
        raise HTTPException(status_code=400, detail="Este usuario es el propietario de la intranet y no puede ser eliminado")
    
    school_id = user["school_id"]
    target_role = target.get("role", "")
    
    # ═══════════════════════════════════════════════════════════════════════════
    # CASCADE DELETE - Clean up all related data based on user role
    # ═══════════════════════════════════════════════════════════════════════════
    
    try:
        # STUDENT-SPECIFIC CLEANUP
        if target_role == "student":
            # Delete attendance records
            await db.attendance.delete_many({"student_id": user_id, "school_id": school_id})
            await db.attendances.delete_many({"student_id": user_id, "school_id": school_id})
            
            # Delete grades
            await db.grades.delete_many({"student_id": user_id, "school_id": school_id})
            
            # Delete exam attempts and submissions
            await db.exam_attempts.delete_many({"student_id": user_id, "school_id": school_id})
            
            # Remove from exam submissions (embedded in online_exams)
            await db.online_exams.update_many(
                {"school_id": school_id},
                {"$pull": {"submissions": {"student_id": user_id}}}
            )
            
            # Remove from task submissions (embedded in academic_assignments)
            await db.academic_assignments.update_many(
                {"school_id": school_id},
                {"$pull": {"submissions": {"student_id": user_id}}}
            )
            
            # Delete discipline reports
            await db.discipline_reports.delete_many({"student_id": user_id, "school_id": school_id})
            
            # Delete survey answers
            await db.survey_answers.delete_many({"user_id": user_id, "school_id": school_id})
            
            # Remove from enrolled_students in subjects/courses
            await db.subjects.update_many(
                {"school_id": school_id},
                {"$pull": {"enrolled_students": user_id}}
            )
            
            # Remove likes from course posts
            await db.course_posts.update_many(
                {"school_id": school_id},
                {"$pull": {"likes": user_id}}
            )
            
            # Remove student from parent's children list
            await db.users.update_many(
                {"school_id": school_id, "children": user_id},
                {"$pull": {"children": user_id}}
            )
            
            logger.info(f"Cleaned up student data for user {user_id}")
        
        # TEACHER-SPECIFIC CLEANUP
        elif target_role == "teacher":
            # Remove teacher from subjects
            await db.subjects.update_many(
                {"school_id": school_id, "teacher_id": user_id},
                {"$set": {"teacher_id": None}}
            )
            
            # Remove as secondary teacher
            await db.subjects.update_many(
                {"school_id": school_id},
                {"$pull": {"secondary_teachers": user_id}}
            )
            
            logger.info(f"Cleaned up teacher data for user {user_id}")
        
        # PARENT-SPECIFIC CLEANUP
        elif target_role == "parent":
            # Remove parent reference from children (if any)
            await db.users.update_many(
                {"school_id": school_id, "parent_id": user_id},
                {"$unset": {"parent_id": ""}}
            )
            
            logger.info(f"Cleaned up parent data for user {user_id}")
        
        # COMMON CLEANUP FOR ALL USERS
        # Delete messages sent by user
        await db.messages.delete_many({"sender_id": user_id, "school_id": school_id})
        
        # Delete internal mail
        await db.internal_mail.delete_many({
            "$or": [
                {"sender_id": user_id},
                {"recipient_ids": user_id}
            ],
            "school_id": school_id
        })
        
        # Remove from recipients in internal mail
        await db.internal_mail.update_many(
            {"school_id": school_id},
            {"$pull": {"recipient_ids": user_id, "read_by": user_id, "deleted_by": user_id}}
        )
        
        # Delete academic thread messages
        await db.academic_threads.update_many(
            {"school_id": school_id},
            {"$pull": {"messages": {"sender_id": user_id}}}
        )
        
    except Exception as e:
        logger.error(f"Error during cascade delete for user {user_id}: {e}")
        # Continue with user deletion even if some cleanup fails
    
    # Delete photo from Cloudinary if exists
    if target.get("photo_url"):
        try:
            photo_url = target["photo_url"]
            if "cloudinary.com" in photo_url:
                parts = photo_url.split("/upload/")
                if len(parts) > 1:
                    path_with_ext = parts[1]
                    if path_with_ext.startswith("v"):
                        path_with_ext = "/".join(path_with_ext.split("/")[1:])
                    public_id = path_with_ext.rsplit(".", 1)[0]
                    cloudinary.uploader.destroy(public_id)
                    logger.info(f"Deleted Cloudinary image: {public_id}")
        except Exception as e:
            logger.error(f"Error deleting Cloudinary image: {e}")
    
    # Finally, delete the user
    await db.users.delete_one({"id": user_id})
    
    logger.info(f"User {user_id} ({target_role}) deleted with cascade cleanup by {user['id']}")
    
    return {"message": "Usuario eliminado correctamente junto con todos sus datos relacionados"}

# ══════════════════════════════════════════════════════════════════════════════
# STUDENT IMPORT - Excel/CSV mass import
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/students/import/template")
async def generate_student_template(
    nivel_id: str = "",
    grado_id: str = "",
    seccion_id: str = "",
    turno_id: str = "",
    current_user = Depends(get_current_user)
):
    """Generate Excel template for student import"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden generar plantillas")

    school_id = user["school_id"]

    nivel_name = ""
    grado_name = ""
    seccion_name = ""
    turno_name = ""

    if nivel_id:
        nivel = await db.academic_levels.find_one({"id": nivel_id, "school_id": school_id}, {"_id": 0})
        nivel_name = nivel.get("nombre", nivel.get("name", "")) if nivel else ""
    if grado_id:
        grado = await db.grades.find_one({"id": grado_id, "school_id": school_id}, {"_id": 0})
        grado_name = grado.get("nombre", grado.get("name", "")) if grado else ""
    if seccion_id:
        seccion = await db.sections.find_one({"id": seccion_id, "school_id": school_id}, {"_id": 0})
        seccion_name = seccion.get("nombre", seccion.get("name", "")) if seccion else ""
    if turno_id:
        turno = await db.shifts.find_one({"id": turno_id, "school_id": school_id}, {"_id": 0})
        turno_name = turno.get("nombre", turno.get("name", "")) if turno else ""

    # Get active academic year
    active_year = await db.academic_years.find_one({"school_id": school_id, "is_active": True}, {"_id": 0})
    anio_escolar = str(active_year.get("year", "")) if active_year else str(datetime.now(timezone.utc).year)

    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    info_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    info_font = Font(name="Arial", italic=True, size=10, color="2E7D32")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "Plantilla de Importacion de Estudiantes"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1B5E20")

    # Row 2-3: Academic filter headers + values (locked/protected)
    filter_header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    filter_header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    filter_labels = ["Nivel", "Grado", "Seccion", "Turno"]
    for col, label in enumerate(filter_labels, 1):
        cell = ws.cell(row=2, column=col, value=label)
        cell.fill = filter_header_fill
        cell.font = filter_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.protection = Protection(locked=True)

    filter_value_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    filter_value_font = Font(name="Arial", bold=True, size=11, color="1B5E20")
    filter_values = [nivel_name or "---", grado_name or "---", seccion_name or "---", turno_name or "---"]
    for col, val in enumerate(filter_values, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.fill = filter_value_fill
        cell.font = filter_value_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.protection = Protection(locked=True)

    # Row 5: Instructions
    instruction_font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.merge_cells("A5:I5")
    ws["A5"] = "Instrucciones: Complete los datos de los estudiantes en las filas inferiores y luego vuelva a subir este archivo en el sistema para importarlos automaticamente."
    ws["A5"].font = instruction_font

    # Row 6: Auto-generated credentials note
    note_font = Font(name="Arial", italic=True, size=9, color="1565C0")
    ws.merge_cells("A6:I6")
    ws["A6"] = "Nota: El usuario y contrasena del estudiante seran generados automaticamente por el sistema."
    ws["A6"].font = note_font

    # Row 8: Student column headers
    student_headers = ["Nombre", "Apellido", "DNI", "Cumpleanos", "Genero", "Celular", "Correo", "Direccion", "Observaciones"]
    for col, hdr in enumerate(student_headers, 1):
        cell = ws.cell(row=8, column=col, value=hdr)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    col_widths = [20, 20, 15, 15, 14, 15, 30, 35, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=8, column=i).column_letter].width = w

    # Freeze at row 9 so headers always visible
    ws.freeze_panes = "A9"

    # Row 9: Example row (unlocked)
    example_data = ["Juan", "Perez", "78451236", "15/03/2010", "Masculino", "987654321", "juan@email.com", "Av. Lima 123", "---"]
    example_font = Font(name="Arial", italic=True, size=10, color="999999")
    for col, val in enumerate(example_data, 1):
        cell = ws.cell(row=9, column=col, value=val)
        cell.font = example_font
        cell.border = thin_border
        cell.protection = Protection(locked=False)

    # Empty rows for student data (unlocked)
    for row in range(10, 510):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if row < 15:
                cell.border = thin_border
            cell.protection = Protection(locked=False)

    # Lock rows 1-8 (title, filters, instructions, headers)
    for row in range(1, 9):
        for col in range(1, 10):
            ws.cell(row=row, column=col).protection = Protection(locked=True)

    # Apply sheet protection
    from openpyxl.worksheet.protection import SheetProtection
    ws.protection = SheetProtection(
        sheet=True, objects=True, scenarios=True,
        formatCells=False, formatColumns=False, formatRows=False,
        insertRows=True, deleteRows=True,
        selectLockedCells=True, selectUnlockedCells=False
    )

    # ── Hidden metadata sheet for verification ──
    meta_ws = wb.create_sheet("edunet_metadata")
    meta_ws.sheet_state = "hidden"
    meta_keys = ["school_id", "nivel_id", "nivel_name", "grado_id", "grado_name",
                 "seccion_id", "seccion_name", "turno_id", "turno_name",
                 "anio_escolar", "fecha_generacion"]
    meta_vals = [school_id, nivel_id, nivel_name, grado_id, grado_name,
                 seccion_id, seccion_name, turno_id or "", turno_name,
                 anio_escolar, datetime.now(timezone.utc).isoformat()]
    for i, (k, v) in enumerate(zip(meta_keys, meta_vals), 1):
        meta_ws.cell(row=i, column=1, value=k)
        meta_ws.cell(row=i, column=2, value=v)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_nivel = nivel_name.replace(" ", "_") or "todos"
    safe_grado = grado_name.replace(" ", "_") or "todos"
    safe_seccion = seccion_name.replace(" ", "_") or "todas"
    filename = f"plantilla_estudiantes_{safe_nivel}_{safe_grado}_{safe_seccion}.xlsx"

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/students/import")
async def import_students(
    file: UploadFile = File(...),
    nivel_id: str = Form(""),
    grado_id: str = Form(""),
    seccion_id: str = Form(""),
    turno_id: str = Form(""),
    use_file_config: str = Form("false"),
    current_user = Depends(get_current_user)
):
    """Import students from Excel or CSV file"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden importar estudiantes")

    school_id = user["school_id"]
    now = datetime.now(timezone.utc).isoformat()

    ext = file.filename.lower().rsplit(".", 1)[-1] if file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Formato no soportado. Use .xlsx, .xls o .csv")

    content = await file.read()
    rows = []

    # ── Metadata verification for xlsx files ──
    file_metadata = {}
    if ext == "xlsx":
        try:
            meta_wb = load_workbook(io.BytesIO(content), read_only=True)
            if "edunet_metadata" in meta_wb.sheetnames:
                meta_ws = meta_wb["edunet_metadata"]
                for row in meta_ws.iter_rows(values_only=True):
                    if row and row[0] and row[1] is not None:
                        file_metadata[str(row[0]).strip()] = str(row[1]).strip()
            meta_wb.close()
        except Exception:
            pass  # No metadata - old template or manual file

    if file_metadata and use_file_config != "true":
        mismatches = []
        meta_nivel = file_metadata.get("nivel_id", "")
        meta_grado = file_metadata.get("grado_id", "")
        meta_seccion = file_metadata.get("seccion_id", "")
        meta_turno = file_metadata.get("turno_id", "")
        meta_school = file_metadata.get("school_id", "")
        meta_year = file_metadata.get("anio_escolar", "")

        if meta_school and meta_school != school_id:
            mismatches.append("school_id")
        if meta_nivel and nivel_id and meta_nivel != nivel_id:
            mismatches.append("nivel")
        if meta_grado and grado_id and meta_grado != grado_id:
            mismatches.append("grado")
        if meta_seccion and seccion_id and meta_seccion != seccion_id:
            mismatches.append("seccion")
        if meta_turno and turno_id and meta_turno != turno_id:
            mismatches.append("turno")

        # Check academic year
        active_year_doc = await db.academic_years.find_one({"school_id": school_id, "is_active": True}, {"_id": 0})
        current_year = str(active_year_doc.get("year", "")) if active_year_doc else str(datetime.now(timezone.utc).year)
        year_mismatch = meta_year and current_year and meta_year != current_year

        if mismatches or year_mismatch:
            return {
                "metadata_mismatch": True,
                "file_config": {
                    "nivel_id": meta_nivel, "nivel_name": file_metadata.get("nivel_name", ""),
                    "grado_id": meta_grado, "grado_name": file_metadata.get("grado_name", ""),
                    "seccion_id": meta_seccion, "seccion_name": file_metadata.get("seccion_name", ""),
                    "turno_id": meta_turno, "turno_name": file_metadata.get("turno_name", ""),
                    "anio_escolar": meta_year,
                    "fecha_generacion": file_metadata.get("fecha_generacion", ""),
                },
                "current_config": {
                    "nivel_id": nivel_id, "grado_id": grado_id,
                    "seccion_id": seccion_id, "turno_id": turno_id,
                    "anio_escolar": current_year,
                },
                "mismatches": mismatches,
                "year_mismatch": year_mismatch,
            }

    # Use file metadata to fill missing filters (either explicit use_file_config or empty form params)
    should_use_file_metadata = file_metadata and (
        use_file_config == "true" or not nivel_id or not grado_id or not seccion_id or not turno_id
    )
    if should_use_file_metadata:
        # Resolve IDs by NAME in the current school (file might come from a different school)
        meta_nivel_name = file_metadata.get("nivel_name", "").strip()
        meta_grado_name = file_metadata.get("grado_name", "").strip()
        meta_seccion_name = file_metadata.get("seccion_name", "").strip()

        if meta_nivel_name:
            real_nivel = await db.academic_levels.find_one(
                {"school_id": school_id, "nombre": {"$regex": f"^{meta_nivel_name}$", "$options": "i"}},
                {"_id": 0, "id": 1}
            )
            if real_nivel:
                nivel_id = real_nivel["id"]
            else:
                nivel_id = file_metadata.get("nivel_id", nivel_id)

        if meta_grado_name and nivel_id:
            real_grado = await db.grades.find_one(
                {"school_id": school_id, "nivel_id": nivel_id, "nombre": {"$regex": f"^{meta_grado_name}$", "$options": "i"}},
                {"_id": 0, "id": 1}
            )
            if real_grado:
                grado_id = real_grado["id"]
            else:
                grado_id = file_metadata.get("grado_id", grado_id)

        if meta_seccion_name and grado_id:
            real_seccion = await db.sections.find_one(
                {"school_id": school_id, "grado_id": grado_id, "nombre": {"$regex": f"^{meta_seccion_name}$", "$options": "i"}},
                {"_id": 0, "id": 1}
            )
            if real_seccion:
                seccion_id = real_seccion["id"]
            else:
                seccion_id = file_metadata.get("seccion_id", seccion_id)

        turno_id = file_metadata.get("turno_id", turno_id)
        meta_turno_name = file_metadata.get("turno_name", "").strip()
        if meta_turno_name:
            real_turno = await db.shifts.find_one(
                {"school_id": school_id, "nombre": {"$regex": f"^{meta_turno_name}$", "$options": "i"}},
                {"_id": 0, "id": 1}
            )
            if real_turno:
                turno_id = real_turno["id"]

    # ── Validate: all 4 academic fields are required ──
    missing_fields = []
    if not nivel_id: missing_fields.append("Nivel")
    if not grado_id: missing_fields.append("Grado")
    if not seccion_id: missing_fields.append("Sección")
    if not turno_id: missing_fields.append("Turno")
    if missing_fields:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo no contiene los datos académicos requeridos: {', '.join(missing_fields)}. Descargue una plantilla nueva con todos los filtros seleccionados (Nivel, Grado, Sección y Turno)."
        )

    try:
        if ext == "csv":
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append({k.strip(): (v.strip() if v else "") for k, v in r.items()})
        elif ext == "xlsx":
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            header_row_idx = None
            for i, row in enumerate(all_rows):
                if row and any(str(c or "").strip().lower() in ("nombre", "name") for c in row):
                    header_row_idx = i
                    break
            if header_row_idx is None:
                raise HTTPException(status_code=400, detail="No se encontro la fila de encabezados (Nombre, Apellido...)")
            headers_raw = [str(c or "").strip() for c in all_rows[header_row_idx]]
            for row in all_rows[header_row_idx + 1:]:
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue
                d = {}
                for j, h in enumerate(headers_raw):
                    if j < len(row) and row[j] is not None:
                        val = row[j]
                        if isinstance(val, datetime):
                            d[h] = val.strftime("%d/%m/%Y")
                        else:
                            d[h] = str(val).strip()
                    else:
                        d[h] = ""
                rows.append(d)
            wb.close()
        elif ext == "xls":
            import xlrd
            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            header_row_idx = None
            for i in range(min(10, sheet.nrows)):
                vals = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
                if any(v.lower() in ("nombre", "name") for v in vals):
                    header_row_idx = i
                    break
            if header_row_idx is None:
                raise HTTPException(status_code=400, detail="No se encontro la fila de encabezados")
            headers_raw = [str(sheet.cell_value(header_row_idx, j)).strip() for j in range(sheet.ncols)]
            for i in range(header_row_idx + 1, sheet.nrows):
                vals = [str(sheet.cell_value(i, j)).strip() for j in range(sheet.ncols)]
                if all(v == "" for v in vals):
                    continue
                d = {}
                for j, h in enumerate(headers_raw):
                    d[h] = vals[j] if j < len(vals) else ""
                rows.append(d)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene datos de estudiantes")

    # ── Fallback: resolve turno from visible Excel metadata (row 2-3) or hidden sheet ──
    if not turno_id:
        # Try 1: resolve from hidden sheet metadata (turno_name)
        meta_turno_name = file_metadata.get("turno_name", "").strip() if file_metadata else ""
        if meta_turno_name:
            real_turno = await db.shifts.find_one(
                {"school_id": school_id, "nombre": {"$regex": f"^{meta_turno_name}$", "$options": "i"}},
                {"_id": 0, "id": 1}
            )
            if real_turno:
                turno_id = real_turno["id"]

        # Try 2: read from visible cells in xlsx (row 2=headers, row 3=values)
        if not turno_id and ext == "xlsx":
            try:
                wb_meta = load_workbook(io.BytesIO(content), read_only=True)
                ws_meta = wb_meta.active
                meta_rows = list(ws_meta.iter_rows(min_row=2, max_row=3, values_only=True))
                if len(meta_rows) >= 2:
                    meta_headers = [str(c or "").strip().lower() for c in meta_rows[0]]
                    meta_values = [str(c or "").strip() for c in meta_rows[1]]
                    for hi, hv in enumerate(meta_headers):
                        if hv == "turno" and hi < len(meta_values) and meta_values[hi]:
                            visible_turno_name = meta_values[hi]
                            real_turno = await db.shifts.find_one(
                                {"school_id": school_id, "nombre": {"$regex": f"^{visible_turno_name}$", "$options": "i"}},
                                {"_id": 0, "id": 1}
                            )
                            if real_turno:
                                turno_id = real_turno["id"]
                            break
                wb_meta.close()
            except Exception:
                pass

    COL_MAP = {
        "nombre": "name", "name": "name",
        "apellido": "last_name", "apellidos": "last_name", "last_name": "last_name",
        "dni": "dni", "documento": "dni",
        "cumpleanos": "birthday", "cumpleaños": "birthday", "birthday": "birthday", "fecha_nacimiento": "birthday",
        "genero": "gender", "género": "gender", "gender": "gender", "sexo": "gender",
        "celular": "phone", "telefono": "phone", "phone": "phone",
        "correo": "email", "email": "email",
        "direccion": "address", "address": "address",
        "observaciones": "notes", "notas": "notes", "notes": "notes",
    }

    def normalize_key(k):
        k = k.lower().strip()
        k = unicodedata.normalize("NFD", k)
        k = "".join(c for c in k if unicodedata.category(c) != "Mn")
        return COL_MAP.get(k, k)

    created = []
    pending = []

    # Check activation mode
    fin_settings = await db.school_financial_settings.find_one(
        {"school_id": school_id}, {"_id": 0, "activacion_modo": 1}
    )
    activation_mode = (fin_settings or {}).get("activacion_modo", "matricula_pension")

    last_code = await db.users.find_one(
        {"school_id": school_id, "student_code": {"$exists": True}},
        sort=[("student_code", -1)],
        projection={"student_code": 1, "_id": 0}
    )
    code_counter = 1
    if last_code and last_code.get("student_code"):
        try:
            code_counter = int(last_code["student_code"].split("-")[1]) + 1
        except (ValueError, IndexError):
            pass

    for idx, raw_row in enumerate(rows):
        row = {normalize_key(k): v for k, v in raw_row.items() if k.strip()}
        name = row.get("name", "").strip()
        last_name = row.get("last_name", "").strip()

        # Skip example row from template
        if name.lower() == "juan" and last_name.lower() == "perez" and row.get("dni", "").strip() == "78451236":
            continue

        dni = row.get("dni", "").strip()
        email = row.get("email", "").strip().lower()
        phone = row.get("phone", "").strip()
        address = row.get("address", "").strip()
        notes = row.get("notes", "").strip()
        birthday_raw = row.get("birthday", "").strip()
        gender_raw = row.get("gender", "").strip().lower()

        # Parse birthday
        birthday = ""
        if birthday_raw:
            birthday_str = str(birthday_raw).strip()
            # Handle "2010-03-15 00:00:00" format from Excel datetime stringification
            if " " in birthday_str:
                birthday_str = birthday_str.split(" ")[0]
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    birthday = datetime.strptime(birthday_str, fmt).strftime("%Y-%m-%d")
                    break
                except (ValueError, TypeError):
                    pass
            if not birthday:
                birthday = birthday_str

        # Normalize gender
        errors = []
        gender = ""
        if gender_raw:
            if gender_raw in ("masculino", "male", "m", "hombre"):
                gender = "male"
            elif gender_raw in ("femenino", "female", "f", "mujer"):
                gender = "female"
            else:
                gender = ""
                errors.append("Genero no valido")
        if not name:
            errors.append("Nombre vacio")
        if not last_name:
            errors.append("Apellido vacio")

        if dni:
            existing_dni = await db.users.find_one(
                {"school_id": school_id, "dni": dni, "student_status": {"$ne": "deleted"}},
                {"_id": 0, "id": 1}
            )
            if existing_dni:
                errors.append(f"DNI {dni} ya existe")

        # Email duplicado permitido para estudiantes (hermanos comparten correo de padres)

        student_code = f"STU-{code_counter:06d}"
        code_counter += 1

        base_username = f"{name.lower().replace(' ', '')}.{last_name.lower().replace(' ', '')}" if name and last_name else f"est{idx}"
        base_username = "".join(c for c in unicodedata.normalize("NFD", base_username) if unicodedata.category(c) != "Mn")
        username = base_username
        suffix = 1
        while await db.users.find_one({"username": username, "school_id": school_id}):
            username = f"{base_username}{suffix}"
            suffix += 1

        new_student = {
            "id": str(uuid.uuid4()),
            "username": username,
            "password": hash_password(dni if dni else "123456"),
            "password_display": dni if dni else "123456",
            "name": name,
            "last_name": last_name,
            "email": email or None,
            "phone": phone or None,
            "dni": dni or None,
            "birthday": birthday or None,
            "gender": gender or None,
            "address": address or None,
            "role": "student",
            "school_id": school_id,
            "email_verified": True,
            "nivel_id": nivel_id or None,
            "grado_id": grado_id or None,
            "seccion_id": seccion_id or None,
            "turno_id": turno_id or None,
            "student_code": student_code,
            "student_status": "pending" if errors else ("active" if activation_mode == "on_create" else "active"),
            "import_status": "pending" if errors else "imported",
            "import_errors": errors if errors else None,
            "import_notes": notes or None,
            "created_at": now,
            "updated_at": now,
        }

        # Determine if this is a duplicate error (DNI/email already exists)
        has_duplicate_error = any("ya existe" in e for e in errors)

        if has_duplicate_error:
            # DO NOT insert — just report the error. The original record already exists.
            pending.append({"row": idx + 1, "name": f"{name} {last_name}", "errors": errors, "student_code": student_code, "skipped": True})
        else:
            # Generate short QR (centralized service)
            qr_id, qr_token = await generate_user_qr(db)
            new_student["qr_id"] = qr_id
            new_student["qr_token"] = qr_token
            new_student["qr_version"] = 2

            if errors:
                pending.append({"row": idx + 1, "name": f"{name} {last_name}", "errors": errors, "student_code": student_code})
                new_student["student_status"] = "pending"
            else:
                new_student["student_status"] = "active"

            await db.users.insert_one(new_student)
            new_student.pop("_id", None)
            new_student.pop("password", None)

            if not errors:
                created.append({"name": f"{name} {last_name}", "student_code": student_code})

    skipped = [p for p in pending if p.get("skipped")]
    real_pending = [p for p in pending if not p.get("skipped")]

    logger.info(f"Student import: {len(created)} created, {len(real_pending)} pending, {len(skipped)} skipped by {user['id']}")

    return {
        "message": f"Importacion completada",
        "created_count": len(created),
        "pending_count": len(real_pending),
        "skipped_count": len(skipped),
        "created": created,
        "pending": real_pending,
        "skipped": skipped,
    }

@router.get("/students/pending")
async def get_pending_students(current_user = Depends(get_current_user)):
    """Get students with import errors"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")

    students = await db.users.find(
        {"school_id": user["school_id"], "role": "student", "import_status": "pending"},
        {"_id": 0, "password": 0}
    ).to_list(500)
    return students

@router.put("/students/pending/{student_id}/activate")
async def activate_pending_student(student_id: str, current_user = Depends(get_current_user)):
    """Activate a pending student after fixing errors"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")

    result = await db.users.update_one(
        {"id": student_id, "school_id": user["school_id"], "role": "student"},
        {"$set": {"import_status": "imported", "student_status": "active", "import_errors": None, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    return {"message": "Estudiante activado correctamente"}

@router.put("/students/pending/{student_id}/edit")
async def edit_pending_student(student_id: str, request: Request, current_user = Depends(get_current_user)):
    """Edit a pending student's data to fix import errors"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores")

    body = await request.json()
    allowed = {"name", "last_name", "dni", "email", "phone", "address", "birthday", "gender", "notes"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if not updates:
        raise HTTPException(status_code=400, detail="Sin campos para actualizar")

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()

    # Re-validate: check DNI/email uniqueness
    new_errors = []
    school_id = user["school_id"]
    if "dni" in updates and updates["dni"]:
        existing = await db.users.find_one(
            {"school_id": school_id, "dni": updates["dni"], "id": {"$ne": student_id}}, {"_id": 0, "id": 1}
        )
        if existing:
            new_errors.append(f"DNI {updates['dni']} ya existe")
    if "email" in updates and updates["email"]:
        existing = await db.users.find_one(
            {"school_id": school_id, "email": updates["email"], "id": {"$ne": student_id}}, {"_id": 0, "id": 1}
        )
        if existing:
            new_errors.append(f"Correo {updates['email']} ya existe")

    if new_errors:
        updates["import_errors"] = new_errors
    else:
        updates["import_errors"] = None
        updates["import_status"] = "imported"
        updates["student_status"] = "active"

    result = await db.users.update_one(
        {"id": student_id, "school_id": school_id, "role": "student"},
        {"$set": updates}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")

    updated = await db.users.find_one({"id": student_id}, {"_id": 0, "password": 0})
    return {"message": "Estudiante actualizado" + (" y activado" if not new_errors else ""), "student": updated, "errors": new_errors}

@router.delete("/students/pending/{student_id}")
async def delete_pending_student(student_id: str, current_user = Depends(get_current_user)):
    """Delete a pending student"""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")

    result = await db.users.delete_one(
        {"id": student_id, "school_id": user["school_id"], "role": "student", "import_status": "pending"}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Estudiante pendiente no encontrado")
    return {"message": "Estudiante eliminado"}

@router.delete("/students/pending")
async def delete_all_pending_students(current_user = Depends(get_current_user)):
    """Delete only DUPLICATE pending students (where a non-pending original exists).
    Preserves pending students that are the only copy (first import with errors)."""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=403, detail="No tienes un colegio asociado")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden limpiar pendientes")

    school_id = user["school_id"]

    # Get all pending students
    pending_students = await db.users.find(
        {"school_id": school_id, "role": "student", "import_status": "pending"},
        {"_id": 0, "id": 1, "dni": 1, "email": 1, "import_errors": 1}
    ).to_list(1000)

    if not pending_students:
        return {"message": "No hay registros pendientes", "deleted_count": 0, "preserved_count": 0}

    ids_to_delete = []
    preserved_count = 0

    for student in pending_students:
        is_safe_to_delete = False
        dni = student.get("dni")
        email = student.get("email")

        # Check if a non-pending copy with the same DNI exists
        if dni:
            original = await db.users.find_one(
                {"school_id": school_id, "role": "student", "dni": dni,
                 "id": {"$ne": student["id"]}, "import_status": {"$ne": "pending"}},
                {"_id": 0, "id": 1}
            )
            if original:
                is_safe_to_delete = True

        # If DNI didn't match, check by email
        if not is_safe_to_delete and email:
            original = await db.users.find_one(
                {"school_id": school_id, "role": "student", "email": email,
                 "id": {"$ne": student["id"]}, "import_status": {"$ne": "pending"}},
                {"_id": 0, "id": 1}
            )
            if original:
                is_safe_to_delete = True

        if is_safe_to_delete:
            ids_to_delete.append(student["id"])
        else:
            preserved_count += 1

    deleted_count = 0
    if ids_to_delete:
        result = await db.users.delete_many(
            {"id": {"$in": ids_to_delete}, "school_id": school_id}
        )
        deleted_count = result.deleted_count

    return {
        "message": f"{deleted_count} duplicados eliminados" + (f", {preserved_count} registros preservados (sin duplicado original)" if preserved_count > 0 else ""),
        "deleted_count": deleted_count,
        "preserved_count": preserved_count
    }

# ══════════════════════════════════════════════════════════════════════════════



# ══════════════════════════════════════════════════════════════════════════════
# BULK SAFE DELETE STUDENTS
# ══════════════════════════════════════════════════════════════════════════════

class BulkSafeDeleteRequest(BaseModel):
    nivel_id: str
    grado_id: str
    seccion_id: str
    turno_id: Optional[str] = None
    delete_reason: str = Field(..., min_length=3)
    confirm: bool = False  # True = execute, False = analyze only

@router.post("/students/bulk-safe-delete")
async def bulk_safe_delete_students(data: BulkSafeDeleteRequest, current_user=Depends(get_current_user)):
    """Analyze and optionally soft-delete students without academic activity. Admin/owner only."""
    user = await resolve_user_from_token(current_user)
    if not user or not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ejecutar esta accion")

    school_id = user.get("school_id")
    if not school_id:
        raise HTTPException(status_code=403, detail="No autorizado")

    student_roles = {"$in": ["student", "estudiante"]}

    # Find students - filter by grado_id + seccion_id (nivel is implicit in grado)
    student_filter = {
        "school_id": school_id,
        "role": student_roles,
        "grado_id": data.grado_id,
        "seccion_id": data.seccion_id,
        "student_status": {"$ne": "deleted"},
    }
    if data.turno_id:
        # Also match students without turno assigned
        student_filter["$or"] = [
            {"turno_id": data.turno_id},
            {"turno_id": None},
            {"turno_id": {"$exists": False}},
        ]

    students = await db.users.find(student_filter, {"_id": 0, "id": 1, "name": 1, "last_name": 1}).to_list(2000)

    # If no results, try progressively relaxed filters to diagnose
    if not students:
        # Try without turno filter
        relaxed_filter = {
            "school_id": school_id,
            "role": student_roles,
            "grado_id": data.grado_id,
            "seccion_id": data.seccion_id,
            "student_status": {"$ne": "deleted"},
        }
        found_without_turno = await db.users.count_documents(relaxed_filter)
        if found_without_turno > 0:
            raise HTTPException(status_code=404, detail=f"Hay {found_without_turno} estudiantes en esa seccion pero con un turno diferente al seleccionado.")

        # Try without seccion filter
        relaxed_filter2 = {
            "school_id": school_id,
            "role": student_roles,
            "grado_id": data.grado_id,
            "student_status": {"$ne": "deleted"},
        }
        found_in_grade = await db.users.count_documents(relaxed_filter2)
        if found_in_grade > 0:
            raise HTTPException(status_code=404, detail=f"No se encontraron estudiantes en esa seccion. Hay {found_in_grade} en el grado pero con otra seccion.")

        # Try by nivel only
        relaxed_filter3 = {
            "school_id": school_id,
            "role": student_roles,
            "nivel_id": data.nivel_id,
            "student_status": {"$ne": "deleted"},
        }
        found_in_nivel = await db.users.count_documents(relaxed_filter3)
        if found_in_nivel > 0:
            raise HTTPException(status_code=404, detail=f"No se encontraron estudiantes con ese grado/seccion. Hay {found_in_nivel} en el nivel pero en otros grados.")

        # Check if there are students without academic assignment
        orphan_filter = {
            "school_id": school_id,
            "role": student_roles,
            "student_status": {"$ne": "deleted"},
            "$or": [{"grado_id": None}, {"grado_id": {"$exists": False}}, {"grado_id": ""}],
        }
        orphan_count = await db.users.count_documents(orphan_filter)
        if orphan_count > 0:
            raise HTTPException(status_code=404, detail=f"No se encontraron estudiantes con esos filtros academicos. Hay {orphan_count} estudiantes sin nivel/grado asignado en el colegio. Usa la herramienta de Huerfanos para gestionarlos.")

        raise HTTPException(status_code=404, detail="No se encontraron estudiantes con esos filtros")

    student_ids = [s["id"] for s in students]

    # Optimized: batch check activity across all collections
    attendance_ids = set(await db.attendances.distinct("user_id", {"user_id": {"$in": student_ids}, "school_id": school_id}))
    legacy_att_ids = set(await db.student_attendance.distinct("student_id", {"student_id": {"$in": student_ids}, "school_id": school_id}))
    grade_ids = set(await db.student_grades.distinct("student_id", {"student_id": {"$in": student_ids}, "school_id": school_id}))
    task_ids = set(await db.task_submissions.distinct("student_id", {"student_id": {"$in": student_ids}, "school_id": school_id}))
    exam_ids = set(await db.exam_attempts.distinct("student_id", {"student_id": {"$in": student_ids}, "school_id": school_id}))
    payment_ids = set(await db.payments.distinct("student_id", {"student_id": {"$in": student_ids}, "school_id": school_id}))

    blocked_ids = attendance_ids | legacy_att_ids | grade_ids | task_ids | exam_ids | payment_ids

    deletable = []
    blocked = []
    for s in students:
        sid = s["id"]
        full_name = f"{s.get('name', '')} {s.get('last_name', '')}".strip()
        if sid in blocked_ids:
            reasons = []
            if sid in attendance_ids or sid in legacy_att_ids:
                reasons.append("Tiene asistencias")
            if sid in grade_ids:
                reasons.append("Tiene notas")
            if sid in task_ids:
                reasons.append("Tiene tareas")
            if sid in exam_ids:
                reasons.append("Tiene examenes")
            if sid in payment_ids:
                reasons.append("Tiene pagos")
            blocked.append({"id": sid, "name": full_name, "reason": ", ".join(reasons)})
        else:
            deletable.append({"id": sid, "name": full_name})

    # Analysis mode (confirm=false)
    if not data.confirm:
        return {
            "mode": "analysis",
            "total_found": len(students),
            "deletable_count": len(deletable),
            "blocked_count": len(blocked),
            "deletable": deletable,
            "blocked": blocked,
        }

    # Execute PERMANENT delete (remove from database completely)
    if not deletable:
        raise HTTPException(status_code=400, detail="No hay alumnos eliminables. Todos tienen actividad academica.")

    now_iso = datetime.now(timezone.utc).isoformat()
    deletable_ids = [d["id"] for d in deletable]

    await db.users.delete_many(
        {"id": {"$in": deletable_ids}, "school_id": school_id}
    )

    # Audit log
    await db.bulk_delete_logs.insert_one({
        "id": str(uuid.uuid4()),
        "school_id": school_id,
        "action": "bulk_safe_delete_students",
        "filters": {"nivel_id": data.nivel_id, "grado_id": data.grado_id, "seccion_id": data.seccion_id, "turno_id": data.turno_id},
        "deleted_count": len(deletable),
        "blocked_count": len(blocked),
        "deleted_ids": deletable_ids,
        "reason": data.delete_reason,
        "performed_by": user["id"],
        "performed_by_name": f"{user.get('name', '')} {user.get('last_name', '')}".strip(),
        "created_at": now_iso,
    })

    logger.info(f"Bulk safe delete: {len(deletable)} deleted, {len(blocked)} blocked in school {school_id}")

    return {
        "mode": "executed",
        "total_found": len(students),
        "deleted": len(deletable),
        "blocked": len(blocked),
        "blocked_students": blocked,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  GET /api/students/export-credentials — Export student credentials as Excel
# ═══════════════════════════════════════════════════════════════════════════════
@router.get("/students/export-credentials")
async def export_student_credentials(
    nivel_id: str = "",
    grado_id: str = "",
    seccion_id: str = "",
    turno_id: str = "",
    current_user=Depends(get_current_user)
):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=400, detail="school_id es requerido")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden exportar credenciales")

    if not nivel_id or not grado_id or not seccion_id or not turno_id:
        raise HTTPException(status_code=400, detail="Todos los filtros son requeridos: nivel, grado, seccion y turno")

    school_id = user["school_id"]
    query = {"role": {"$in": ["student", "estudiante"]}, "school_id": school_id, "student_status": {"$ne": "deleted"}}
    query["nivel_id"] = nivel_id
    query["grado_id"] = grado_id
    query["seccion_id"] = seccion_id
    query["turno_id"] = turno_id

    students = await db.users.find(
        query,
        {"_id": 0, "name": 1, "last_name": 1, "username": 1, "dni": 1}
    ).to_list(5000)

    if not students:
        raise HTTPException(status_code=404, detail="No se encontraron estudiantes con los filtros aplicados")

    # Resolve filter names
    nivel_doc = await db.academic_levels.find_one({"id": nivel_id, "school_id": school_id}, {"_id": 0, "name": 1, "nombre": 1})
    grado_doc = await db.grades.find_one({"id": grado_id, "school_id": school_id}, {"_id": 0, "name": 1, "nombre": 1})
    seccion_doc = await db.sections.find_one({"id": seccion_id, "school_id": school_id}, {"_id": 0, "name": 1, "nombre": 1})
    turno_doc = await db.shifts.find_one({"id": turno_id, "school_id": school_id}, {"_id": 0, "name": 1, "nombre": 1})

    nivel_name = (nivel_doc.get("name") or nivel_doc.get("nombre") or nivel_id) if nivel_doc else nivel_id
    grado_name = (grado_doc.get("name") or grado_doc.get("nombre") or grado_id) if grado_doc else grado_id
    seccion_name = (seccion_doc.get("name") or seccion_doc.get("nombre") or seccion_id) if seccion_doc else seccion_id
    turno_name = (turno_doc.get("name") or turno_doc.get("nombre") or turno_id) if turno_doc else turno_id

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime, timezone

    wb = Workbook()
    ws = wb.active
    ws.title = "Credenciales"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20

    title_font = Font(name="Arial", bold=True, size=14, color="1565C0")
    label_font = Font(name="Arial", bold=True, size=11, color="333333")
    value_font = Font(name="Arial", size=11, color="333333")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Row 1: Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "Credenciales de Estudiantes"
    ws["A1"].font = title_font

    # Row 3-8: Metadata
    meta = [
        ("Nivel:", nivel_name),
        ("Grado:", grado_name),
        ("Seccion:", seccion_name),
        ("Turno:", turno_name),
        ("Fecha de exportacion:", datetime.now(timezone.utc).strftime("%d/%m/%Y")),
        ("Total de estudiantes:", str(len(students))),
    ]
    for i, (label, value) in enumerate(meta, 3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=value).font = value_font

    # Row 10: Table headers
    data_start = 10
    headers = ["Nombre completo", "Usuario", "Contrasena"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.freeze_panes = f"A{data_start + 1}"

    # Data rows
    for idx, s in enumerate(students, data_start + 1):
        full_name = f"{s.get('name', '')} {s.get('last_name', '')}".strip()
        username = s.get("username", "")
        dni = (s.get("dni") or "").strip()
        password = dni if dni else "123456"

        ws.cell(row=idx, column=1, value=full_name).border = thin_border
        ws.cell(row=idx, column=2, value=username).border = thin_border
        ws.cell(row=idx, column=3, value=password).border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Build descriptive filename
    import re as _re
    def sanitize(s):
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        s = _re.sub(r"[^a-zA-Z0-9]+", "_", s.lower()).strip("_")
        return s or "x"

    filename = f"credenciales_{sanitize(nivel_name)}_{sanitize(grado_name)}_{sanitize(seccion_name)}.xlsx"

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  GET /api/teachers/export-credentials — Export teacher credentials as Excel
# ═══════════════════════════════════════════════════════════════════════════════
@router.get("/teachers/export-credentials")
async def export_teacher_credentials(current_user=Depends(get_current_user)):
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=400, detail="school_id es requerido")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden exportar credenciales")

    school_id = user["school_id"]

    teachers = await db.users.find(
        {"role": "teacher", "school_id": school_id},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "username": 1, "plain_password": 1}
    ).sort([("last_name", 1), ("name", 1)]).to_list(5000)

    if not teachers:
        raise HTTPException(status_code=404, detail="No hay profesores para exportar")

    # Backfill: generate new passwords for teachers missing plain_password
    import random, string
    def _gen_pwd(length=10):
        chars = string.ascii_letters + string.digits
        while True:
            pwd = "".join(random.choices(chars, k=length))
            if any(c.isupper() for c in pwd) and any(c.islower() for c in pwd) and any(c.isdigit() for c in pwd):
                return pwd

    for t in teachers:
        if not t.get("plain_password"):
            new_pwd = _gen_pwd()
            await db.users.update_one(
                {"id": t["id"], "school_id": school_id},
                {"$set": {"plain_password": new_pwd, "password_display": new_pwd, "password": hash_password(new_pwd), "updated_at": now_iso()}}
            )
            t["plain_password"] = new_pwd
            logger.info(f"Backfilled plain_password for teacher {t.get('username')} ({t['id']})")

    # Get school name for metadata
    school = await db.schools.find_one({"id": school_id}, {"_id": 0, "name": 1, "school_name": 1})
    school_name = (school.get("school_name") or school.get("name") or school_id) if school else school_id

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime, timezone

    wb = Workbook()
    ws = wb.active
    ws.title = "Credenciales"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20

    title_font = Font(name="Arial", bold=True, size=14, color="1565C0")
    label_font = Font(name="Arial", bold=True, size=11, color="333333")
    value_font = Font(name="Arial", size=11, color="333333")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Row 1: Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "Credenciales de Profesores"
    ws["A1"].font = title_font

    # Row 3-5: Metadata
    meta = [
        ("Colegio:", school_name),
        ("Fecha de exportacion:", datetime.now(timezone.utc).strftime("%d/%m/%Y")),
        ("Total de profesores:", str(len(teachers))),
    ]
    for i, (label, value) in enumerate(meta, 3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=value).font = value_font

    # Row 7: Table headers
    data_start = 7
    headers = ["Nombre del Profesor", "Nombre de Usuario", "Contrasena"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.freeze_panes = f"A{data_start + 1}"

    # Data rows
    for idx, t in enumerate(teachers, data_start + 1):
        full_name = f"{t.get('last_name', '')} {t.get('name', '')}".strip()
        username = t.get("username", "")
        password = t.get("plain_password", "")

        ws.cell(row=idx, column=1, value=full_name).border = thin_border
        ws.cell(row=idx, column=2, value=username).border = thin_border
        ws.cell(row=idx, column=3, value=password).border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    import re as _re
    def sanitize(s):
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        s = _re.sub(r"[^a-zA-Z0-9]+", "_", s.lower()).strip("_")
        return s or "x"

    date_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"credenciales_profesores_{sanitize(school_name)}_{date_str}.xlsx"

    logger.info(f"Teacher credentials exported by user {user.get('id')} — {len(teachers)} teachers, school {school_id}")

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/parents/export-credentials")
async def export_parent_credentials(current_user=Depends(get_current_user)):
    """Export an Excel with parent login credentials (name, username, email, password).

    Mirrors the teacher credentials export. Parents may log in by username OR
    email, so both are included. The plain password lives in `plain_password`
    or (for older docs) `password_display`; if neither exists a new password is
    generated and persisted (same backfill behaviour as teachers)."""
    user = await resolve_user_from_token(current_user)
    if not user or not user.get("school_id"):
        raise HTTPException(status_code=400, detail="school_id es requerido")
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden exportar credenciales")

    school_id = user["school_id"]

    parents = await db.users.find(
        {"role": "parent", "school_id": school_id, "student_status": {"$ne": "deleted"}},
        {"_id": 0, "id": 1, "name": 1, "last_name": 1, "username": 1, "email": 1,
         "plain_password": 1, "password_display": 1}
    ).sort([("last_name", 1), ("name", 1)]).to_list(None)

    if not parents:
        raise HTTPException(status_code=404, detail="No hay padres para exportar")

    import random, string
    def _gen_pwd(length=10):
        chars = string.ascii_letters + string.digits
        while True:
            pwd = "".join(random.choices(chars, k=length))
            if any(c.isupper() for c in pwd) and any(c.islower() for c in pwd) and any(c.isdigit() for c in pwd):
                return pwd

    for p in parents:
        plain = p.get("plain_password") or p.get("password_display")
        if not plain:
            new_pwd = _gen_pwd()
            await db.users.update_one(
                {"id": p["id"], "school_id": school_id},
                {"$set": {"plain_password": new_pwd, "password_display": new_pwd, "password": hash_password(new_pwd), "updated_at": now_iso()}}
            )
            plain = new_pwd
            logger.info(f"Backfilled plain_password for parent {p.get('username') or p.get('email')} ({p['id']})")
        p["_plain"] = plain

    school = await db.schools.find_one({"id": school_id}, {"_id": 0, "name": 1, "school_name": 1})
    school_name = (school.get("school_name") or school.get("name") or school_id) if school else school_id

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime, timezone

    wb = Workbook()
    ws = wb.active
    ws.title = "Credenciales"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 20

    title_font = Font(name="Arial", bold=True, size=14, color="1565C0")
    label_font = Font(name="Arial", bold=True, size=11, color="333333")
    value_font = Font(name="Arial", size=11, color="333333")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = "Credenciales de Padres de Familia"
    ws["A1"].font = title_font

    meta = [
        ("Colegio:", school_name),
        ("Fecha de exportacion:", datetime.now(timezone.utc).strftime("%d/%m/%Y")),
        ("Total de padres:", str(len(parents))),
    ]
    for i, (label, value) in enumerate(meta, 3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=value).font = value_font

    data_start = 7
    headers = ["Nombre del Apoderado", "Nombre de Usuario", "Correo", "Contrasena"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.freeze_panes = f"A{data_start + 1}"

    for idx, p in enumerate(parents, data_start + 1):
        full_name = f"{p.get('last_name', '')} {p.get('name', '')}".strip()
        ws.cell(row=idx, column=1, value=full_name).border = thin_border
        ws.cell(row=idx, column=2, value=p.get("username") or "").border = thin_border
        ws.cell(row=idx, column=3, value=p.get("email") or "").border = thin_border
        ws.cell(row=idx, column=4, value=p.get("_plain") or "").border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    import re as _re
    def sanitize(s):
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        s = _re.sub(r"[^a-zA-Z0-9]+", "_", s.lower()).strip("_")
        return s or "x"

    date_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"credenciales_padres_{sanitize(school_name)}_{date_str}.xlsx"

    logger.info(f"Parent credentials exported by user {user.get('id')} — {len(parents)} parents, school {school_id}")

    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ══════════════════════════════════════════════════════════════════════════════
# Teacher QR bulk download (PDF grid — mirrors student QR download)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/teachers/qr/bulk-download")
async def teacher_qr_bulk_download(current_user=Depends(get_current_user)):
    """Generate PDF with QR cards for all teachers in the school (3x3 grid per page)."""
    from fastapi.responses import StreamingResponse, JSONResponse
    import qrcode
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.colors import HexColor
    import httpx
    from PIL import Image as PILImage

    user = await resolve_user_from_token(current_user)
    if not user or not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden descargar QR de profesores")

    school_id = user.get("school_id")
    if not school_id:
        raise HTTPException(status_code=403, detail="No autorizado")

    try:
        logger.info(f"[QR Bulk] === Starting teacher QR bulk download for school {school_id} ===")

        # Phase 1: Fetch teachers from DB
        teachers = await db.users.find(
            {"role": "teacher", "school_id": school_id},
            {"_id": 0, "id": 1, "name": 1, "last_name": 1, "qr_token": 1, "username": 1, "photo_url": 1}
        ).sort([("last_name", 1), ("name", 1)]).to_list(5000)
        logger.info(f"[QR Bulk] Phase 1 - Fetch teachers: Found {len(teachers)} teachers")

        if not teachers:
            raise HTTPException(status_code=404, detail="No hay profesores para generar QR")

        # Phase 2: Backfill qr_token for teachers that don't have one
        logger.info("[QR Bulk] Phase 2 - Backfill qr_tokens: Starting...")
        backfill_count = 0
        for t in teachers:
            if not t.get("qr_token"):
                try:
                    qr_id, qr_token = await generate_user_qr(db)
                    await db.users.update_one(
                        {"id": t["id"]},
                        {"$set": {"qr_id": qr_id, "qr_token": qr_token}}
                    )
                    t["qr_token"] = qr_token
                    backfill_count += 1
                except Exception as e:
                    logger.warning(f"[QR Bulk] Failed to backfill qr_token for teacher {t.get('id')}: {e}")
        logger.info(f"[QR Bulk] Phase 2 complete: Backfilled {backfill_count} new tokens")

        # Get school info
        school = await db.schools.find_one({"id": school_id}, {"_id": 0, "name": 1, "school_name": 1, "nombre": 1, "logo_url": 1})
        school_name = (school or {}).get("name") or (school or {}).get("school_name") or (school or {}).get("nombre") or "Colegio"
        school_logo_url = (school or {}).get("logo_url")

        # Pre-fetch school logo once (with resize to save memory)
        logo_img = None
        if school_logo_url:
            try:
                async with httpx.AsyncClient(timeout=httpx.Timeout(5.0)) as client:
                    resp = await client.get(school_logo_url)
                    if resp.status_code == 200:
                        pil_logo = PILImage.open(BytesIO(resp.content))
                        if pil_logo.mode in ('RGBA', 'P', 'LA'):
                            pil_logo = pil_logo.convert('RGBA')
                            _bg = PILImage.new('RGB', pil_logo.size, (255, 255, 255))
                            _bg.paste(pil_logo, mask=pil_logo.split()[-1])
                            pil_logo = _bg
                        else:
                            pil_logo = pil_logo.convert('RGB')
                        pil_logo.thumbnail((200, 200))
                        logo_img = BytesIO()
                        pil_logo.save(logo_img, format='JPEG', quality=75)
                        logo_img.seek(0)
                        del pil_logo
                        logger.info("[QR Bulk] School logo downloaded and resized OK")
            except Exception as e:
                logger.warning(f"[QR Bulk] School logo download failed: {e}")

        def make_qr_image(token_data, size=250):
            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=2)
            qr.add_data(token_data)
            qr.make(fit=True)
            return qr.make_image(fill_color="black", back_color="white").resize((size, size))

        buf = BytesIO()
        c = pdf_canvas.Canvas(buf, pagesize=A4)
        w, h = A4

        cols, rows = 3, 3
        card_w = 60 * mm
        card_h = 88 * mm
        margin_x = (w - cols * card_w) / (cols + 1)
        margin_y = (h - rows * card_h) / (rows + 1)

        navy = HexColor("#001f4b")
        teal = HexColor("#0d9488")
        gray = HexColor("#64748b")
        light_bg = HexColor("#f1f5f9")
        border_color = HexColor("#d1d5db")

        total_teachers = len([t for t in teachers if t.get("qr_token")])
        logger.info(f"[QR Bulk] Phase 3 - PDF generation: {total_teachers} teachers with QR tokens")
        card_idx = 0
        for t in teachers:
            if not t.get("qr_token"):
                continue
            if card_idx > 0 and card_idx % (cols * rows) == 0:
                c.showPage()

            pos = card_idx % (cols * rows)
            col = pos % cols
            row = pos // cols
            x = margin_x + col * (card_w + margin_x)
            y = h - margin_y - (row + 1) * card_h - row * margin_y

            # Card border
            c.setFillColor(HexColor("#ffffff"))
            c.setStrokeColor(border_color)
            c.setLineWidth(0.5)
            c.roundRect(x, y, card_w, card_h, 2 * mm, fill=1, stroke=1)

            # Top teal bar
            c.setFillColor(teal)
            c.rect(x + 0.5, y + card_h - 4 * mm, card_w - 1, 4 * mm, fill=1, stroke=0)

            # Logo + School name header
            logo_y = y + card_h - 19 * mm
            if logo_img:
                try:
                    logo_img.seek(0)
                    c.drawImage(ImageReader(logo_img), x + (card_w - 10 * mm) / 2, logo_y + 2 * mm, 10 * mm, 10 * mm, preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass

            c.setFillColor(navy)
            c.setFont("Helvetica-Bold", 6)
            display_name = "COLEGIO" + school_name[7:] if school_name.lower().startswith("colegio") else f"COLEGIO {school_name}"
            name_trunc = display_name[:30]
            tw = c.stringWidth(name_trunc, "Helvetica-Bold", 6)
            c.drawString(x + (card_w - tw) / 2, logo_y - 2 * mm, name_trunc)

            # Divider
            c.setStrokeColor(HexColor("#e2e8f0"))
            c.setLineWidth(0.4)
            c.line(x + 4 * mm, logo_y - 4 * mm, x + card_w - 4 * mm, logo_y - 4 * mm)

            # Sequential photo download + resize for this teacher
            photo_size = 20 * mm
            photo_x = x + (card_w - photo_size) / 2
            photo_y = logo_y - 5 * mm - photo_size
            teacher_photo_buf = None
            photo_url = t.get("photo_url")
            if photo_url:
                try:
                    logger.info(f"[QR Bulk] Downloading photo {card_idx + 1}/{total_teachers}: teacher {t.get('id')}")
                    async with httpx.AsyncClient(timeout=httpx.Timeout(5.0)) as photo_client:
                        resp = await photo_client.get(photo_url)
                        if resp.status_code == 200:
                            pil_img = PILImage.open(BytesIO(resp.content))
                            if pil_img.mode in ('RGBA', 'P', 'LA'):
                                pil_img = pil_img.convert('RGB')
                            pil_img.thumbnail((200, 200))
                            teacher_photo_buf = BytesIO()
                            pil_img.save(teacher_photo_buf, format='JPEG', quality=75)
                            teacher_photo_buf.seek(0)
                            del pil_img
                except Exception as photo_err:
                    logger.warning(f"[QR Bulk] Photo download failed for teacher {t.get('id')}: {photo_err}")
                    teacher_photo_buf = None

            # Draw teacher photo or initials fallback
            if teacher_photo_buf:
                try:
                    c.saveState()
                    path = c.beginPath()
                    cx_p = photo_x + photo_size / 2
                    cy_p = photo_y + photo_size / 2
                    path.circle(cx_p, cy_p, photo_size / 2)
                    path.close()
                    c.clipPath(path, stroke=0)
                    c.drawImage(ImageReader(teacher_photo_buf), photo_x, photo_y, photo_size, photo_size, preserveAspectRatio=True, mask='auto')
                    c.restoreState()
                    c.setStrokeColor(HexColor("#cbd5e1"))
                    c.setLineWidth(0.8)
                    c.circle(cx_p, cy_p, photo_size / 2, fill=0, stroke=1)
                except Exception:
                    try:
                        c.restoreState()
                    except Exception:
                        pass
                    c.setFillColor(light_bg)
                    c.circle(photo_x + photo_size / 2, photo_y + photo_size / 2, photo_size / 2, fill=1, stroke=0)
                    c.setFillColor(navy)
                    c.setFont("Helvetica-Bold", 16)
                    c.drawCentredString(photo_x + photo_size / 2, photo_y + photo_size / 2 - 3, (t.get("name", "?")[:1]).upper())
            else:
                c.setFillColor(light_bg)
                c.circle(photo_x + photo_size / 2, photo_y + photo_size / 2, photo_size / 2, fill=1, stroke=0)
                c.setFillColor(navy)
                c.setFont("Helvetica-Bold", 16)
                c.drawCentredString(photo_x + photo_size / 2, photo_y + photo_size / 2 - 3, (t.get("name", "?")[:1]).upper())
            content_top = photo_y - 4 * mm

            # Explicitly free photo buffer to release memory
            if teacher_photo_buf:
                try:
                    teacher_photo_buf.close()
                except Exception:
                    pass
                del teacher_photo_buf

            # Teacher name
            info_y = content_top
            c.setFillColor(navy)
            c.setFont("Helvetica-Bold", 7)
            full_name = f"{t.get('name', '')} {t.get('last_name', '')}".strip()
            if len(full_name) > 22:
                full_name = full_name[:21] + "."
            tw = c.stringWidth(full_name, "Helvetica-Bold", 7)
            c.drawString(x + (card_w - tw) / 2, info_y, full_name)

            # Role label
            c.setFillColor(gray)
            c.setFont("Helvetica", 5.5)
            role_line = "PROFESOR"
            tw2 = c.stringWidth(role_line, "Helvetica", 5.5)
            c.drawString(x + (card_w - tw2) / 2, info_y - 4 * mm, role_line)

            # QR code
            footer_y = y + 2 * mm
            qr_top = info_y - 7 * mm
            qr_bottom = footer_y + 4 * mm
            available = qr_top - qr_bottom
            qr_size_px = min(available, 32 * mm)
            qr_size_px = max(qr_size_px, 18 * mm)

            qr_img = make_qr_image(t["qr_token"], 250)
            qr_buf = BytesIO()
            qr_img.save(qr_buf, format="PNG")
            qr_buf.seek(0)
            qr_x = x + (card_w - qr_size_px) / 2
            qr_y = qr_bottom + (available - qr_size_px) / 2
            c.drawImage(ImageReader(qr_buf), qr_x, qr_y, qr_size_px, qr_size_px)

            # Footer
            c.setFillColor(HexColor("#94a3b8"))
            c.setFont("Helvetica", 4)
            c.drawCentredString(x + card_w / 2, footer_y, "Personal e intransferible")

            # Free QR buffer
            del qr_buf

            card_idx += 1

        c.save()
        buf.seek(0)
        logger.info(f"[QR Bulk] Phase 3 complete: PDF generated with {card_idx} cards")

        # Free logo buffer
        if logo_img:
            try:
                logo_img.close()
            except Exception:
                pass
            del logo_img

        now_str = datetime.now(timezone.utc).strftime("%Y%m%d")
        filename = f"qr_profesores_{now_str}.pdf"

        logger.info(f"[QR Bulk] === SUCCESS: PDF exported by user {user.get('id')} — {card_idx} teachers, school {school_id} ===")

        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"})
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[QR Bulk] CRITICAL: Error generating teacher QR PDF for school {school_id}: {e}")
        return JSONResponse(status_code=500, content={"detail": f"Error generando PDF de QR: {str(e)}"})


# ══════════════════════════════════════════════════════════════════════════════
# ASSIGN DNI AS PASSWORD FOR PARENTS
# ══════════════════════════════════════════════════════════════════════════════

class AsignarClaveDniRequest(BaseModel):
    sobrescribir: bool = False

@router.post("/admin/padres/asignar-clave-dni")
async def asignar_clave_dni_padres(data: AsignarClaveDniRequest, user=Depends(require_admin())):
    """Assign DNI as password to parents. Only for admin/owner."""
    if not user:
        raise HTTPException(status_code=403, detail="Usuario no encontrado")
    school_id = user.get("school_id")
    if not school_id:
        raise HTTPException(status_code=400, detail="No se encontro el colegio")

    query = {"school_id": school_id, "role": "parent"}
    parents = await db.users.find(query, {"_id": 0, "id": 1, "dni": 1, "password": 1, "name": 1, "last_name": 1}).to_list(5000)

    actualizados = 0
    sin_dni = 0
    omitidos_con_clave = 0

    for p in parents:
        dni = (p.get("dni") or "").strip()
        if not dni:
            sin_dni += 1
            continue

        has_password = bool(p.get("password"))
        if has_password and not data.sobrescribir:
            omitidos_con_clave += 1
            continue

        hashed = hash_password(dni)
        await db.users.update_one(
            {"id": p["id"]},
            {"$set": {"password": hashed, "password_display": dni, "updated_at": now_iso()}}
        )
        actualizados += 1

    logger.info(f"[ASIGNAR-DNI] school={school_id} actualizados={actualizados} sin_dni={sin_dni} omitidos={omitidos_con_clave} sobrescribir={data.sobrescribir}")

    return {
        "actualizados": actualizados,
        "sin_dni": sin_dni,
        "omitidos_con_clave": omitidos_con_clave,
        "total_procesados": len(parents),
    }
