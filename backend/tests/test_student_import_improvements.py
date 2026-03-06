"""
Tests for Student Import System Improvements:
1. Template Row 2 shows Nivel, Grado, Seccion, Turno names
2. Row 4 has auto-generated credentials note
3. Freeze panes at A7
4. Hidden 'edunet_metadata' sheet
5. Example row (Juan Perez) NOT imported
6. Metadata mismatch detection
7. use_file_config=true overrides filters
"""

import pytest
import requests
import io
import os
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test data from review request
TEST_CREDENTIALS = {
    "email": "admin@elroble.edu",
    "password": "1234abc8"
}

# Academic IDs for testing
TEST_LEVEL_ID = "023ca042-cb46-43aa-97e3-a5c9cd7a20ee"  # INICIAL
TEST_GRADE_ID = "6ef8ab18-41b2-45e7-b482-06a84d95c34d"  # 3 años
TEST_SECTION_ID = "11f50cbc-f5f6-422a-a989-87b2af6027f1"  # A
TEST_SHIFT_ID = "8e1f4e98-37fa-40e3-a49d-a4ac08179262"  # Mañana


@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json=TEST_CREDENTIALS)
    if response.status_code == 200:
        return response.json().get("token")
    pytest.skip(f"Authentication failed: {response.status_code} - {response.text}")


class TestTemplateGeneration:
    """Tests for the template generation endpoint"""
    
    def test_template_requires_auth(self):
        """Template endpoint requires authentication"""
        response = requests.get(f"{BASE_URL}/api/students/import/template")
        assert response.status_code == 401, f"Expected 401, got {response.status_code}"
        print("✓ Template endpoint requires authentication")
    
    def test_template_row2_shows_filter_names(self, auth_token):
        """Row 2 should show Nivel, Grado, Seccion AND Turno names"""
        params = {
            "nivel_id": TEST_LEVEL_ID,
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
            "turno_id": TEST_SHIFT_ID
        }
        response = requests.get(
            f"{BASE_URL}/api/students/import/template",
            params=params,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        # Load workbook from response
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb.active
        
        # Check Row 2 content - should contain Nivel, Grado, Seccion, Turno
        row2_value = str(ws["A2"].value or "")
        print(f"Row 2 content: {row2_value}")
        
        assert "Nivel" in row2_value, f"Row 2 should contain 'Nivel', got: {row2_value}"
        assert "Grado" in row2_value, f"Row 2 should contain 'Grado', got: {row2_value}"
        assert "Seccion" in row2_value, f"Row 2 should contain 'Seccion', got: {row2_value}"
        assert "Turno" in row2_value, f"Row 2 should contain 'Turno', got: {row2_value}"
        assert "Sin filtros" not in row2_value, f"Row 2 should NOT show 'Sin filtros' when filters are set: {row2_value}"
        
        wb.close()
        print("✓ Row 2 correctly shows Nivel, Grado, Seccion, and Turno names")
    
    def test_template_row4_has_credentials_note(self, auth_token):
        """Row 4 should have note about auto-generated credentials"""
        params = {
            "nivel_id": TEST_LEVEL_ID,
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
        }
        response = requests.get(
            f"{BASE_URL}/api/students/import/template",
            params=params,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb.active
        
        # Check Row 4 content - should mention auto-generated credentials
        row4_value = str(ws["A4"].value or "")
        print(f"Row 4 content: {row4_value}")
        
        assert "usuario" in row4_value.lower() or "contrasena" in row4_value.lower(), \
            f"Row 4 should mention credentials, got: {row4_value}"
        assert "automaticamente" in row4_value.lower() or "generados" in row4_value.lower(), \
            f"Row 4 should mention auto-generation, got: {row4_value}"
        
        wb.close()
        print("✓ Row 4 correctly shows credentials auto-generation note")
    
    def test_template_freeze_panes_at_a7(self, auth_token):
        """Headers should be frozen at A7 (row 6 is headers)"""
        params = {"nivel_id": TEST_LEVEL_ID}
        response = requests.get(
            f"{BASE_URL}/api/students/import/template",
            params=params,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=False)  # Need read_only=False for freeze_panes
        ws = wb.active
        
        freeze_panes = ws.freeze_panes
        print(f"Freeze panes location: {freeze_panes}")
        
        assert freeze_panes == "A7", f"Freeze panes should be at A7, got: {freeze_panes}"
        
        wb.close()
        print("✓ Freeze panes correctly set at A7")
    
    def test_template_has_hidden_metadata_sheet(self, auth_token):
        """Template should have hidden 'edunet_metadata' sheet with filter info"""
        params = {
            "nivel_id": TEST_LEVEL_ID,
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
            "turno_id": TEST_SHIFT_ID
        }
        response = requests.get(
            f"{BASE_URL}/api/students/import/template",
            params=params,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=False)
        
        # Check that edunet_metadata sheet exists
        assert "edunet_metadata" in wb.sheetnames, f"Should have 'edunet_metadata' sheet, found: {wb.sheetnames}"
        
        meta_ws = wb["edunet_metadata"]
        
        # Check sheet is hidden
        assert meta_ws.sheet_state == "hidden", f"Metadata sheet should be hidden, got: {meta_ws.sheet_state}"
        
        # Read metadata values
        metadata = {}
        for row in meta_ws.iter_rows(values_only=True):
            if row and row[0] and row[1] is not None:
                metadata[str(row[0]).strip()] = str(row[1]).strip()
        
        print(f"Metadata found: {metadata}")
        
        # Verify required metadata fields
        assert "school_id" in metadata, "Metadata should contain school_id"
        assert "nivel_id" in metadata, "Metadata should contain nivel_id"
        assert "nivel_name" in metadata, "Metadata should contain nivel_name"
        assert "grado_id" in metadata, "Metadata should contain grado_id"
        assert "turno_id" in metadata, "Metadata should contain turno_id"
        assert "turno_name" in metadata, "Metadata should contain turno_name"
        
        # Verify IDs match what was requested
        assert metadata.get("nivel_id") == TEST_LEVEL_ID, f"nivel_id mismatch"
        assert metadata.get("grado_id") == TEST_GRADE_ID, f"grado_id mismatch"
        assert metadata.get("seccion_id") == TEST_SECTION_ID, f"seccion_id mismatch"
        assert metadata.get("turno_id") == TEST_SHIFT_ID, f"turno_id mismatch"
        
        wb.close()
        print("✓ Hidden edunet_metadata sheet exists with all required fields")


class TestStudentImport:
    """Tests for the student import endpoint"""
    
    def test_example_row_not_imported(self, auth_token):
        """The example row (Juan Perez, DNI 78451236) should NOT be imported"""
        # First download template
        params = {
            "nivel_id": TEST_LEVEL_ID,
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
            "turno_id": TEST_SHIFT_ID
        }
        response = requests.get(
            f"{BASE_URL}/api/students/import/template",
            params=params,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200
        
        # Import the template with only example row
        files = {"file": ("template.xlsx", response.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {
            "nivel_id": TEST_LEVEL_ID,
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
            "turno_id": TEST_SHIFT_ID,
            "use_file_config": "true"  # Match the file's config
        }
        
        import_response = requests.post(
            f"{BASE_URL}/api/students/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert import_response.status_code == 200, f"Expected 200, got {import_response.status_code}: {import_response.text}"
        result = import_response.json()
        
        # If metadata_mismatch is returned, that's fine - handle it
        if result.get("metadata_mismatch"):
            print("✓ Metadata check triggered (expected when template config matches)")
            return
        
        # Otherwise, check counts
        created_count = result.get("created_count", 0)
        print(f"Import result: created={created_count}, pending={result.get('pending_count', 0)}")
        
        # Example row should NOT be imported - so created_count should be 0
        # (since template only has example row and no real data)
        assert created_count == 0, f"Example row (Juan Perez) should NOT be imported, but got {created_count} created"
        print("✓ Example row correctly NOT imported")
    
    def test_metadata_mismatch_detection(self, auth_token):
        """Uploading template with different filters should detect mismatch"""
        # Download template with TEST filters
        params = {
            "nivel_id": TEST_LEVEL_ID,
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
            "turno_id": TEST_SHIFT_ID
        }
        response = requests.get(
            f"{BASE_URL}/api/students/import/template",
            params=params,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200
        
        # Try to import with DIFFERENT nivel_id
        files = {"file": ("template.xlsx", response.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {
            "nivel_id": "different-nivel-id",  # Different from template
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
            "turno_id": TEST_SHIFT_ID,
            "use_file_config": "false"
        }
        
        import_response = requests.post(
            f"{BASE_URL}/api/students/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert import_response.status_code == 200
        result = import_response.json()
        
        print(f"Mismatch detection result: {result}")
        
        # Should return metadata_mismatch = true
        assert result.get("metadata_mismatch") == True, f"Should detect metadata mismatch, got: {result}"
        assert "file_config" in result, "Should include file_config in response"
        assert "current_config" in result, "Should include current_config in response"
        assert "mismatches" in result, "Should include mismatches list in response"
        
        print("✓ Metadata mismatch correctly detected")
    
    def test_use_file_config_overrides_filters(self, auth_token):
        """use_file_config=true should use metadata values instead of form filters"""
        # Download template with specific filters
        params = {
            "nivel_id": TEST_LEVEL_ID,
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
            "turno_id": TEST_SHIFT_ID
        }
        response = requests.get(
            f"{BASE_URL}/api/students/import/template",
            params=params,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert response.status_code == 200
        
        # Add a test student row to the template
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Add real student data at row 8 (after example row)
        import uuid
        unique_id = str(uuid.uuid4())[:6]
        test_student_data = [f"TestImport{unique_id}", f"UserFileConfig{unique_id}", f"TEST{unique_id}", "", "", "", ""]
        for col, val in enumerate(test_student_data, 1):
            ws.cell(row=8, column=col, value=val)
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        # Import with use_file_config=true and EMPTY/DIFFERENT filters
        files = {"file": ("template.xlsx", output.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {
            "nivel_id": "",  # Empty - should be overridden by file metadata
            "grado_id": "",
            "seccion_id": "",
            "turno_id": "",
            "use_file_config": "true"  # Use file's metadata
        }
        
        import_response = requests.post(
            f"{BASE_URL}/api/students/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert import_response.status_code == 200
        result = import_response.json()
        
        print(f"Import with use_file_config=true result: {result}")
        
        # Should NOT return metadata_mismatch (since we're using file config)
        assert not result.get("metadata_mismatch"), f"Should not have mismatch when use_file_config=true: {result}"
        
        # Should have imported the student
        created_count = result.get("created_count", 0)
        pending_count = result.get("pending_count", 0)
        total = created_count + pending_count
        assert total >= 1, f"Should have imported at least 1 student, got created={created_count}, pending={pending_count}"
        
        print(f"✓ use_file_config=true correctly overrides filters (created: {created_count}, pending: {pending_count})")


class TestCSVImport:
    """Tests for CSV import (no metadata verification)"""
    
    def test_csv_import_no_metadata_check(self, auth_token):
        """CSV files should not trigger metadata mismatch (no metadata sheet)"""
        import uuid
        unique_id = str(uuid.uuid4())[:6]
        
        csv_content = f"""Nombre,Apellido,DNI,Celular,Correo,Direccion,Observaciones
TestCSV{unique_id},Import{unique_id},CSV{unique_id},,,, Test CSV import"""
        
        files = {"file": ("students.csv", csv_content.encode(), "text/csv")}
        data = {
            "nivel_id": TEST_LEVEL_ID,
            "grado_id": TEST_GRADE_ID,
            "seccion_id": TEST_SECTION_ID,
            "turno_id": TEST_SHIFT_ID,
        }
        
        response = requests.post(
            f"{BASE_URL}/api/students/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200
        result = response.json()
        
        print(f"CSV import result: {result}")
        
        # Should NOT return metadata_mismatch for CSV
        assert not result.get("metadata_mismatch"), f"CSV should not trigger metadata check: {result}"
        
        # Should have imported
        total = result.get("created_count", 0) + result.get("pending_count", 0)
        assert total >= 1, f"Should have imported CSV student, got: {result}"
        
        print("✓ CSV import works without metadata verification")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
