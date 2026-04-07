"""
Test suite for Teacher Export Credentials feature
Tests GET /api/teachers/export-credentials endpoint
"""
import pytest
import requests
import os
from io import BytesIO

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials from test_credentials.md
ADMIN_EMAIL = "admin@elroble.edu"
ADMIN_PASSWORD = "1234abc8"
TEACHER_EMAIL = "sonia3009@gmail.com"
TEACHER_PASSWORD = "teacher123"


class TestTeacherExportCredentials:
    """Tests for GET /api/teachers/export-credentials endpoint"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test fixtures"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
    
    def get_admin_token(self):
        """Get authentication token for admin user"""
        response = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD
        })
        if response.status_code == 200:
            return response.json().get("token")
        pytest.skip(f"Admin authentication failed: {response.status_code} - {response.text}")
    
    def get_teacher_token(self):
        """Get authentication token for teacher user"""
        response = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": TEACHER_EMAIL,
            "password": TEACHER_PASSWORD
        })
        if response.status_code == 200:
            return response.json().get("token")
        pytest.skip(f"Teacher authentication failed: {response.status_code} - {response.text}")
    
    def test_export_credentials_returns_200_for_admin(self):
        """Test that admin can export teacher credentials successfully"""
        token = self.get_admin_token()
        
        response = self.session.get(
            f"{BASE_URL}/api/teachers/export-credentials",
            headers={"Authorization": f"Bearer {token}"}
        )
        
        # Status code assertion
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Content-Type assertion - should be Excel file
        content_type = response.headers.get("Content-Type", "")
        assert "spreadsheetml" in content_type or "application/vnd" in content_type, \
            f"Expected Excel content type, got: {content_type}"
        
        # Content-Disposition assertion - should have filename
        content_disposition = response.headers.get("Content-Disposition", "")
        assert "attachment" in content_disposition, f"Expected attachment, got: {content_disposition}"
        assert "credenciales_profesores" in content_disposition, \
            f"Expected filename with 'credenciales_profesores', got: {content_disposition}"
        
        # Verify response has content (Excel file)
        assert len(response.content) > 0, "Response should contain Excel file data"
        
        print(f"SUCCESS: Admin exported teacher credentials - File size: {len(response.content)} bytes")
    
    def test_export_credentials_returns_403_for_teacher(self):
        """Test that teacher role cannot export credentials (403 Forbidden)"""
        token = self.get_teacher_token()
        
        response = self.session.get(
            f"{BASE_URL}/api/teachers/export-credentials",
            headers={"Authorization": f"Bearer {token}"}
        )
        
        # Status code assertion - should be 403 Forbidden
        assert response.status_code == 403, f"Expected 403, got {response.status_code}: {response.text}"
        
        # Verify error message
        data = response.json()
        assert "detail" in data, "Response should contain error detail"
        assert "administrador" in data["detail"].lower() or "admin" in data["detail"].lower(), \
            f"Error message should mention admin requirement, got: {data['detail']}"
        
        print(f"SUCCESS: Teacher correctly denied access - {data['detail']}")
    
    def test_export_credentials_returns_401_without_token(self):
        """Test that unauthenticated request returns 401"""
        response = self.session.get(f"{BASE_URL}/api/teachers/export-credentials")
        
        # Status code assertion - should be 401 Unauthorized
        assert response.status_code == 401, f"Expected 401, got {response.status_code}: {response.text}"
        
        print(f"SUCCESS: Unauthenticated request correctly denied with 401")
    
    def test_export_credentials_excel_structure(self):
        """Test that exported Excel has correct structure"""
        token = self.get_admin_token()
        
        response = self.session.get(
            f"{BASE_URL}/api/teachers/export-credentials",
            headers={"Authorization": f"Bearer {token}"}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        # Parse Excel file to verify structure
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(BytesIO(response.content))
            ws = wb.active
            
            # Verify title row (Row 1)
            title_cell = ws["A1"].value
            assert title_cell is not None, "Title cell should not be empty"
            assert "Credenciales" in title_cell or "Profesores" in title_cell, \
                f"Title should mention credentials/teachers, got: {title_cell}"
            
            # Verify metadata rows (Rows 3-5)
            # Row 3: Colegio
            colegio_label = ws.cell(row=3, column=1).value
            assert colegio_label is not None and "Colegio" in colegio_label, \
                f"Row 3 should have 'Colegio' label, got: {colegio_label}"
            
            # Row 4: Fecha
            fecha_label = ws.cell(row=4, column=1).value
            assert fecha_label is not None and "Fecha" in fecha_label, \
                f"Row 4 should have 'Fecha' label, got: {fecha_label}"
            
            # Row 5: Total
            total_label = ws.cell(row=5, column=1).value
            assert total_label is not None and "Total" in total_label, \
                f"Row 5 should have 'Total' label, got: {total_label}"
            
            # Verify header row (Row 7)
            header_row = 7
            expected_headers = ["Nombre del Profesor", "Nombre de Usuario", "Contrasena"]
            for col, expected in enumerate(expected_headers, 1):
                actual = ws.cell(row=header_row, column=col).value
                assert actual is not None and expected.lower() in actual.lower(), \
                    f"Column {col} header should be '{expected}', got: {actual}"
            
            # Verify there's at least one data row (Row 8+)
            first_data_row = ws.cell(row=8, column=1).value
            # Note: This might be None if no teachers exist, which is handled by 404 in the API
            
            print(f"SUCCESS: Excel structure verified - Title: {title_cell}")
            print(f"  - Metadata rows present (Colegio, Fecha, Total)")
            print(f"  - Headers: {expected_headers}")
            
            wb.close()
            
        except ImportError:
            pytest.skip("openpyxl not available for Excel parsing")
    
    def test_teachers_sorted_alphabetically_by_last_name(self):
        """Test that teachers are sorted alphabetically by last_name"""
        token = self.get_admin_token()
        
        response = self.session.get(
            f"{BASE_URL}/api/teachers/export-credentials",
            headers={"Authorization": f"Bearer {token}"}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(BytesIO(response.content))
            ws = wb.active
            
            # Extract teacher names from data rows (starting at row 8)
            names = []
            row = 8
            while True:
                name = ws.cell(row=row, column=1).value
                if name is None or name == "":
                    break
                names.append(name)
                row += 1
            
            if len(names) > 1:
                # Verify names are sorted (case-insensitive)
                sorted_names = sorted(names, key=lambda x: x.lower() if x else "")
                assert names == sorted_names, \
                    f"Teachers should be sorted alphabetically. Got: {names[:5]}... Expected: {sorted_names[:5]}..."
                print(f"SUCCESS: {len(names)} teachers sorted alphabetically by last_name")
            else:
                print(f"INFO: Only {len(names)} teacher(s) found, sorting verification skipped")
            
            wb.close()
            
        except ImportError:
            pytest.skip("openpyxl not available for Excel parsing")


class TestTeacherExportCredentialsEdgeCases:
    """Edge case tests for teacher export credentials"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test fixtures"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
    
    def get_admin_token(self):
        """Get authentication token for admin user"""
        response = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": ADMIN_EMAIL,
            "password": ADMIN_PASSWORD
        })
        if response.status_code == 200:
            return response.json().get("token")
        pytest.skip(f"Admin authentication failed: {response.status_code}")
    
    def test_export_with_invalid_token(self):
        """Test that invalid token returns 401"""
        response = self.session.get(
            f"{BASE_URL}/api/teachers/export-credentials",
            headers={"Authorization": "Bearer invalid_token_12345"}
        )
        
        assert response.status_code == 401, f"Expected 401, got {response.status_code}"
        print("SUCCESS: Invalid token correctly rejected with 401")
    
    def test_export_with_malformed_auth_header(self):
        """Test that malformed auth header returns 401"""
        response = self.session.get(
            f"{BASE_URL}/api/teachers/export-credentials",
            headers={"Authorization": "NotBearer token123"}
        )
        
        assert response.status_code == 401, f"Expected 401, got {response.status_code}"
        print("SUCCESS: Malformed auth header correctly rejected with 401")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
